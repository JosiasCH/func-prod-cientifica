import csv
import hashlib
import io
import json
import logging
import os
import re
import time
import unicodedata
from datetime import datetime, timezone

import azure.functions as func

app = func.FunctionApp(http_auth_level=func.AuthLevel.FUNCTION)


# =========================
# CONFIG / GENERAL HELPERS
# =========================
def get_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise RuntimeError(f"Missing required setting: {name}")
    return value


def parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on", "si", "sí"}


def get_bool_setting(name: str, default: bool = False) -> bool:
    return parse_bool(os.environ.get(name), default=default)


def parse_positive_int(value: str | None, field_name: str) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        parsed = int(str(value).strip())
    except Exception:
        raise ValueError(f"{field_name} must be an integer.")
    if parsed < 1:
        raise ValueError(f"{field_name} must be >= 1.")
    return parsed


def resolve_llm_enabled(request_value: str | None = None) -> bool:
    if request_value is not None and str(request_value).strip() != "":
        return parse_bool(request_value, default=False)
    return get_bool_setting("THEMATIC_LLM_ENABLED", default=False)


def resolve_career_ambiguity_ai_enabled(request_value: str | None = None) -> bool:
    """
    Controla la IA SOLO para resolver carrera objetivo en casos ambiguos:
    ULima + Facultad/Escuela de Ingeniería genérica, pero sin carrera explícita
    Industrial/Civil/Sistemas.

    Es independiente de THEMATIC_LLM_ENABLED para evitar activar LLM en todos
    los campos temáticos y disparar costos/429.
    """
    if request_value is not None and str(request_value).strip() != "":
        return parse_bool(request_value, default=False)
    return get_bool_setting("CAREER_AMBIGUITY_LLM_ENABLED", default=False)


def get_career_ambiguity_llm_min_confidence() -> float:
    raw = os.environ.get("CAREER_AMBIGUITY_LLM_MIN_CONFIDENCE", "0.78")
    try:
        value = float(raw)
    except Exception:
        value = 0.78
    return max(0.0, min(1.0, value))


def get_openai_max_retries() -> int:
    raw = os.environ.get("AZURE_OPENAI_MAX_RETRIES", "4")
    try:
        value = int(raw)
    except Exception:
        value = 4
    return max(0, min(8, value))


def get_openai_retry_base_seconds() -> float:
    raw = os.environ.get("AZURE_OPENAI_RETRY_BASE_SECONDS", "2.0")
    try:
        value = float(raw)
    except Exception:
        value = 2.0
    return max(0.25, min(30.0, value))


def parse_nonnegative_float(value: str | None, default: float = 0.0) -> float:
    if value is None or str(value).strip() == "":
        return default
    try:
        parsed = float(str(value).strip())
    except Exception:
        return default
    return max(0.0, parsed)


def parse_optional_positive_int_setting(name: str) -> int | None:
    raw = os.environ.get(name)
    if raw is None or str(raw).strip() == "":
        return None
    try:
        parsed = int(str(raw).strip())
    except Exception:
        return None
    return parsed if parsed >= 1 else None


def get_career_ai_call_delay_seconds() -> float:
    """
    Pausa entre llamadas exitosas a Azure OpenAI para carrera ambigua.

    El run con career_ai=true puede generar muchas llamadas seguidas a /responses;
    sin throttling aparecen 429 constantes aunque el SDK reintente.
    Default conservador: 8.0 segundos.
    """
    raw = os.environ.get("CAREER_AI_CALL_DELAY_SECONDS")
    if raw is None or str(raw).strip() == "":
        raw = os.environ.get("AZURE_OPENAI_CALL_DELAY_SECONDS", "8.0")
    return min(60.0, parse_nonnegative_float(raw, default=8.0))


def get_career_ai_max_calls_per_run() -> int | None:
    """Límite opcional para evitar timeouts/costos en corridas masivas."""
    return parse_optional_positive_int_setting("CAREER_AI_MAX_CALLS_PER_RUN")


def get_career_ai_throttle_after_429_seconds() -> float:
    raw = os.environ.get("CAREER_AI_EXTRA_SLEEP_AFTER_429_SECONDS", "8.0")
    return min(120.0, parse_nonnegative_float(raw, default=8.0))


def resolve_thematic_review_ai_enabled(request_value: str | None = None) -> bool:
    """
    IA selectiva para revisar/corregir SOLO clasificaciones temáticas sospechosas
    de área/línea de carrera e IDIC. No reemplaza al clasificador completo.
    """
    if request_value is not None and str(request_value).strip() != "":
        return parse_bool(request_value, default=False)
    return get_bool_setting("THEMATIC_REVIEW_AI_ENABLED", default=False)


def get_thematic_review_ai_min_confidence() -> float:
    raw = os.environ.get("THEMATIC_REVIEW_AI_MIN_CONFIDENCE", "0.82")
    try:
        value = float(raw)
    except Exception:
        value = 0.82
    return max(0.0, min(1.0, value))


def get_thematic_review_ai_call_delay_seconds() -> float:
    raw = os.environ.get("THEMATIC_REVIEW_AI_CALL_DELAY_SECONDS", "8.0")
    return min(60.0, parse_nonnegative_float(raw, default=8.0))


def get_thematic_review_ai_max_calls_per_run() -> int | None:
    return parse_optional_positive_int_setting("THEMATIC_REVIEW_AI_MAX_CALLS_PER_RUN")


def should_send_thematic_review_to_rejected() -> bool:
    """
    Si true, una decisión IA REVIEW en clasificación temática manda el caso a REVIEW
    y no entra a curated. Default false para no reducir cobertura sin validar.
    """
    return get_bool_setting("THEMATIC_REVIEW_SEND_UNSAFE_TO_REVIEW", default=False)


def resolve_ai_taxonomy_classifier_enabled(request_value: str | None = None) -> bool:
    """
    IA principal para clasificar la taxonomía completa de registros ya aceptados
    como Ingeniería ULima: área/línea de carrera + categoría/área/línea IDIC.

    Es independiente de THEMATIC_LLM_ENABLED. Cuando está activo, la IA clasifica
    todos los registros válidos del lote con catálogo cerrado; las reglas quedan
    como fallback solo cuando esta opción está apagada.
    """
    if request_value is not None and str(request_value).strip() != "":
        return parse_bool(request_value, default=False)
    return get_bool_setting("AI_TAXONOMY_CLASSIFIER_ENABLED", default=False)


def get_ai_taxonomy_min_confidence() -> float:
    raw = os.environ.get("AI_TAXONOMY_MIN_CONFIDENCE", "0.78")
    try:
        value = float(raw)
    except Exception:
        value = 0.78
    return max(0.0, min(1.0, value))


def get_ai_taxonomy_call_delay_seconds() -> float:
    raw = os.environ.get("AI_TAXONOMY_CALL_DELAY_SECONDS", "8.0")
    return min(60.0, parse_nonnegative_float(raw, default=8.0))


def get_ai_taxonomy_max_calls_per_run() -> int | None:
    return parse_optional_positive_int_setting("AI_TAXONOMY_MAX_CALLS_PER_RUN")


def should_ai_taxonomy_review_reject() -> bool:
    """
    Si la IA taxonómica no puede clasificar con seguridad, el registro no debe
    pasar a curated. Default true: preferimos REVIEW antes que mala taxonomía.
    """
    return get_bool_setting("AI_TAXONOMY_REVIEW_TO_REJECTED", default=True)


def get_ingest_singleton_lock_enabled() -> bool:
    """Evita ejecuciones simultáneas del ingest con un lock distribuido en SQL."""
    return get_bool_setting("SCOPUS_INGEST_SINGLETON_LOCK_ENABLED", default=True)


def get_ingest_singleton_lock_timeout_ms() -> int:
    raw = os.environ.get("SCOPUS_INGEST_LOCK_TIMEOUT_MS", "0")
    try:
        parsed = int(str(raw).strip())
    except Exception:
        parsed = 0
    return max(0, min(600000, parsed))


def acquire_scopus_ingest_singleton_lock():
    """
    Toma un candado distribuido en Azure SQL usando sp_getapplock.

    El throttle de Python solo funciona dentro de un worker/proceso. Este lock
    bloquea ejecuciones concurrentes entre workers, instancias o clicks repetidos
    desde Azure Portal. Mantiene la conexión abierta durante todo el run porque
    el lock owner es la sesión SQL.
    """
    if not get_ingest_singleton_lock_enabled():
        return None

    from mssql_python import connect

    conn = connect(get_sql_connection_string())
    cursor = conn.cursor()
    lock_timeout_ms = get_ingest_singleton_lock_timeout_ms()

    try:
        cursor.execute(
            """
            DECLARE @lock_result INT;
            EXEC @lock_result = sp_getapplock
                @Resource = ?,
                @LockMode = 'Exclusive',
                @LockOwner = 'Session',
                @LockTimeout = ?;
            SELECT @lock_result AS lock_result;
            """,
            ("scopus_ingest_pipeline_singleton", lock_timeout_ms),
        )
        row = cursor.fetchone()
        lock_result = int(row_attr(row, "lock_result", 0))
        cursor.close()

        if lock_result < 0:
            try:
                conn.close()
            except Exception:
                pass
            raise RuntimeError(
                "Another Scopus ingestion run is already in progress. "
                "Wait for it to finish before starting a new run. "
                f"sp_getapplock_result={lock_result}; timeout_ms={lock_timeout_ms}."
            )

        logging.info("Acquired Scopus ingest singleton lock. sp_getapplock_result=%s", lock_result)
        return conn

    except Exception:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        raise


def release_scopus_ingest_singleton_lock(lock_conn) -> None:
    if lock_conn is None:
        return

    try:
        cursor = lock_conn.cursor()
        cursor.execute(
            """
            DECLARE @release_result INT;
            EXEC @release_result = sp_releaseapplock
                @Resource = ?,
                @LockOwner = 'Session';
            SELECT @release_result AS release_result;
            """,
            ("scopus_ingest_pipeline_singleton",),
        )
        row = cursor.fetchone()
        release_result = int(row_attr(row, "release_result", 0))
        lock_conn.commit()
        cursor.close()
        logging.info("Released Scopus ingest singleton lock. sp_releaseapplock_result=%s", release_result)
    except Exception as exc:
        logging.warning("Could not release Scopus ingest singleton lock cleanly: %s", str(exc))
    finally:
        try:
            lock_conn.close()
        except Exception:
            pass


_LAST_AZURE_OPENAI_CALL_AT = 0.0
_CAREER_AI_CALLS_THIS_RUN = 0
_THEMATIC_REVIEW_AI_CALLS_THIS_RUN = 0
_AI_TAXONOMY_CALLS_THIS_RUN = 0


def reset_ai_runtime_counters() -> None:
    global _CAREER_AI_CALLS_THIS_RUN, _THEMATIC_REVIEW_AI_CALLS_THIS_RUN, _AI_TAXONOMY_CALLS_THIS_RUN
    _CAREER_AI_CALLS_THIS_RUN = 0
    _THEMATIC_REVIEW_AI_CALLS_THIS_RUN = 0
    _AI_TAXONOMY_CALLS_THIS_RUN = 0


def get_career_ai_calls_this_run() -> int:
    return int(_CAREER_AI_CALLS_THIS_RUN)


def get_thematic_review_ai_calls_this_run() -> int:
    return int(_THEMATIC_REVIEW_AI_CALLS_THIS_RUN)


def get_ai_taxonomy_calls_this_run() -> int:
    return int(_AI_TAXONOMY_CALLS_THIS_RUN)


def reserve_ai_taxonomy_call_slot() -> dict:
    global _AI_TAXONOMY_CALLS_THIS_RUN

    max_calls = get_ai_taxonomy_max_calls_per_run()
    if max_calls is not None and _AI_TAXONOMY_CALLS_THIS_RUN >= max_calls:
        return {
            "allowed": False,
            "calls_used": _AI_TAXONOMY_CALLS_THIS_RUN,
            "max_calls": max_calls,
        }

    _AI_TAXONOMY_CALLS_THIS_RUN += 1
    return {
        "allowed": True,
        "calls_used": _AI_TAXONOMY_CALLS_THIS_RUN,
        "max_calls": max_calls,
    }


def reserve_thematic_review_ai_call_slot() -> dict:
    global _THEMATIC_REVIEW_AI_CALLS_THIS_RUN

    max_calls = get_thematic_review_ai_max_calls_per_run()
    if max_calls is not None and _THEMATIC_REVIEW_AI_CALLS_THIS_RUN >= max_calls:
        return {
            "allowed": False,
            "calls_used": _THEMATIC_REVIEW_AI_CALLS_THIS_RUN,
            "max_calls": max_calls,
        }

    _THEMATIC_REVIEW_AI_CALLS_THIS_RUN += 1
    return {
        "allowed": True,
        "calls_used": _THEMATIC_REVIEW_AI_CALLS_THIS_RUN,
        "max_calls": max_calls,
    }



def reserve_career_ai_call_slot() -> dict:
    """
    Reserva una llamada IA para el run actual.
    Si CAREER_AI_MAX_CALLS_PER_RUN se supera, el caso se manda a REVIEW
    en lugar de seguir saturando Azure OpenAI.
    """
    global _CAREER_AI_CALLS_THIS_RUN

    max_calls = get_career_ai_max_calls_per_run()
    if max_calls is not None and _CAREER_AI_CALLS_THIS_RUN >= max_calls:
        return {
            "allowed": False,
            "calls_used": _CAREER_AI_CALLS_THIS_RUN,
            "max_calls": max_calls,
        }

    _CAREER_AI_CALLS_THIS_RUN += 1
    return {
        "allowed": True,
        "calls_used": _CAREER_AI_CALLS_THIS_RUN,
        "max_calls": max_calls,
    }


def throttle_azure_openai_call(delay_seconds: float, label: str = "azure_openai") -> None:
    """
    Throttle global simple dentro del worker. Evita ráfagas de requests a /responses.
    No intenta paralelizar: prioriza estabilidad del pipeline sobre velocidad.
    """
    global _LAST_AZURE_OPENAI_CALL_AT

    delay_seconds = max(0.0, float(delay_seconds or 0.0))
    if delay_seconds <= 0:
        _LAST_AZURE_OPENAI_CALL_AT = time.monotonic()
        return

    now = time.monotonic()
    if _LAST_AZURE_OPENAI_CALL_AT > 0:
        elapsed = now - _LAST_AZURE_OPENAI_CALL_AT
        remaining = delay_seconds - elapsed
        if remaining > 0:
            logging.info("Throttling %s Azure OpenAI call for %.2fs", label, remaining)
            time.sleep(remaining)

    _LAST_AZURE_OPENAI_CALL_AT = time.monotonic()


def extract_retry_after_seconds_from_exception(exc: Exception) -> float | None:
    """Intenta leer Retry-After / retry-after-ms de una excepción OpenAI/Azure."""
    response = getattr(exc, "response", None)
    headers = getattr(response, "headers", None) if response is not None else None
    if not headers:
        return None

    try:
        retry_ms = headers.get("retry-after-ms") or headers.get("x-ms-retry-after-ms")
        if retry_ms:
            return max(0.0, float(retry_ms) / 1000.0)
    except Exception:
        pass

    try:
        retry_after = headers.get("retry-after") or headers.get("Retry-After")
        if retry_after:
            return max(0.0, float(retry_after))
    except Exception:
        pass

    return None


def get_default_max_rows_per_run() -> int | None:
    return parse_positive_int(os.environ.get("SCOPUS_MAX_ROWS_PER_RUN"), "SCOPUS_MAX_ROWS_PER_RUN")


def get_hard_max_rows_per_run() -> int:
    """Run84: límite duro de seguridad; recomendación operativa = 40 filas."""
    parsed = parse_positive_int(os.environ.get("SCOPUS_HARD_MAX_ROWS_PER_RUN", "40"), "SCOPUS_HARD_MAX_ROWS_PER_RUN")
    return parsed or 40


def resolve_save_rejected_enabled(request_value: str | None = None) -> bool:
    """
    Controla si los registros rechazados se guardan en stg.scopus_raw_load para auditoría.

    Por defecto queda ACTIVADO porque en cargas masivas necesitamos revisar falsos negativos
    por afiliación/carrera/tipo documental sin contaminar curated.publications.
    """
    if request_value is not None and str(request_value).strip() != "":
        return parse_bool(request_value, default=True)
    return get_bool_setting("SCOPUS_SAVE_REJECTED_TO_STAGING", default=True)


def get_request_json_payload(req: func.HttpRequest) -> dict:
    """
    Lee payload JSON si el endpoint se ejecuta por POST desde Azure Portal/Test/Run.

    Azure Functions expone query string en req.params, pero en el portal muchas veces
    los parámetros se envían por body. Sin esta lectura, career_ai=true puede quedar
    ignorado y el run reporta career_ambiguity_ai_enabled=false.
    """
    try:
        payload = req.get_json()
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def get_request_value(req: func.HttpRequest, payload: dict, aliases: list[str]) -> str | None:
    """
    Devuelve el primer valor encontrado en query string o JSON body.
    Soporta alias para evitar fallos por nombres como career_ai, careerAi,
    use_career_ai, etc.
    """
    for alias in aliases:
        value = req.params.get(alias)
        if value is not None and str(value).strip() != "":
            return str(value).strip()

    for alias in aliases:
        value = payload.get(alias) if isinstance(payload, dict) else None
        if value is not None and str(value).strip() != "":
            return str(value).strip()

    return None


def resolve_request_bool(
    req: func.HttpRequest,
    payload: dict,
    aliases: list[str],
    env_name: str | None = None,
    default: bool = False,
) -> bool:
    """
    Boolean robusto para parámetros HTTP. Primero toma query/body; si no existe,
    cae a variable de entorno opcional.
    """
    request_value = get_request_value(req, payload, aliases)
    if request_value is not None:
        return parse_bool(request_value, default=default)

    if env_name:
        return get_bool_setting(env_name, default=default)

    return default


def validate_career_ai_runtime_configuration(use_career_ai: bool) -> None:
    """
    Si se solicita career_ai=true, no se debe continuar silenciosamente sin Azure OpenAI.
    Antes el pipeline podía seguir con heurística y el usuario veía ai_inferred=0.
    """
    if not use_career_ai:
        return

    missing: list[str] = []
    if not os.environ.get("AZURE_OPENAI_API_KEY"):
        missing.append("AZURE_OPENAI_API_KEY")
    if not (os.environ.get("AZURE_OPENAI_BASE_URL") or os.environ.get("AZURE_OPENAI_ENDPOINT")):
        missing.append("AZURE_OPENAI_ENDPOINT or AZURE_OPENAI_BASE_URL")
    if not os.environ.get("AZURE_OPENAI_RESPONSES_MODEL"):
        missing.append("AZURE_OPENAI_RESPONSES_MODEL")

    if missing:
        raise RuntimeError(
            "career_ai=true was requested, but Azure OpenAI is not fully configured. "
            "Missing settings: " + ", ".join(missing)
        )


def validate_ai_taxonomy_runtime_configuration(use_ai_taxonomy: bool) -> None:
    """
    Si se solicita ai_taxonomy=true, la clasificación semántica no debe caer
    silenciosamente al clasificador por reglas.
    """
    if not use_ai_taxonomy:
        return

    missing: list[str] = []
    if not os.environ.get("AZURE_OPENAI_API_KEY"):
        missing.append("AZURE_OPENAI_API_KEY")
    if not (os.environ.get("AZURE_OPENAI_BASE_URL") or os.environ.get("AZURE_OPENAI_ENDPOINT")):
        missing.append("AZURE_OPENAI_ENDPOINT or AZURE_OPENAI_BASE_URL")
    if not os.environ.get("AZURE_OPENAI_RESPONSES_MODEL"):
        missing.append("AZURE_OPENAI_RESPONSES_MODEL")

    if missing:
        raise RuntimeError(
            "ai_taxonomy=true was requested, but Azure OpenAI is not fully configured. "
            "Missing settings: " + ", ".join(missing)
        )


def slice_rows_by_1_based_range(
    rows: list[dict],
    start_row: int | None = None,
    end_row: int | None = None,
    max_rows: int | None = None,
) -> tuple[list[dict], int, int | None]:
    """
    Devuelve un subconjunto de filas manteniendo numeración 1-based respecto al CSV original.

    start_row/end_row se refieren a filas de datos, no a la cabecera.
    Ejemplo: start_row=41&end_row=80 procesa las filas de datos 41..80 del CSV.
    """
    total = len(rows)
    start = start_row or 1

    if end_row is not None and max_rows is not None:
        raise ValueError("Use either end_row or max_rows, not both.")

    if end_row is not None and end_row < start:
        raise ValueError("end_row must be >= start_row.")

    if max_rows is not None:
        end = min(total, start + max_rows - 1)
    elif end_row is not None:
        end = min(total, end_row)
    else:
        default_max = get_default_max_rows_per_run()
        if default_max is not None:
            end = min(total, start + default_max - 1)
        else:
            end = total

    if start > total:
        return [], start, None

    selected = rows[start - 1 : end]
    return selected, start, end


def get_sql_connection_string() -> str:
    return get_env("SQL_CONNECTION_STRING")


def get_blob_service():
    from azure.storage.blob import BlobServiceClient

    conn_str = get_env("DATA_STORAGE_CONNECTION_STRING")
    return BlobServiceClient.from_connection_string(conn_str)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def safe_get(row: dict, aliases: list[str]) -> str | None:
    for alias in aliases:
        if alias in row and row[alias] not in (None, ""):
            return str(row[alias]).strip()
    return None


def compute_record_hash(eid: str | None, doi: str | None, title: str | None) -> str:
    raw = f"{eid or ''}|{doi or ''}|{title or ''}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def normalize_generic_text(value: str | None) -> str:
    if not value:
        return ""

    value = str(value).lower().strip()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.replace("\\", " ")
    value = value.replace("–", " ").replace("—", " ").replace("-", " ")
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalize_person_name(value: str | None) -> str:
    if not value:
        return ""
    norm = normalize_generic_text(value)
    norm = norm.replace(",", " ")
    norm = norm.replace("/", " ")
    norm = norm.replace(".", " ")
    norm = re.sub(r"\s+", " ", norm).strip()
    return norm


def split_semicolon_values(value: str | None) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in str(value).split(";") if part and str(part).strip()]


def clean_author_full_name(value: str) -> str:
    return re.sub(r"\s*\(\d+\)\s*$", "", str(value)).strip()


def is_ulima_text(value: str | None) -> bool:
    """
    Detecta afiliación REAL a Universidad de Lima / University of Lima / ULIMA.

    Parche run80:
    - Evita falsos positivos por instituciones externas que contienen "University of Lima"
      como parte de otro nombre institucional.
    - Normaliza puntuación para capturar casos como:
      "National Technological University of Lima, Sur Lima".
    - No acepta "Lima" como ciudad; solo la universidad exacta.
    """
    norm = normalize_generic_text(value)
    if not norm:
        return False

    # Versión sin puntuación para que "Lima, Sur" y "Lima Sur" sean equivalentes.
    norm_flat = re.sub(r"[^a-z0-9]+", " ", norm).strip()
    norm_flat = re.sub(r"\s+", " ", norm_flat)

    negative_patterns = [
        "universidad de lima sur",
        "university of lima sur",
        "technological university of lima",
        "technological university of lima sur",
        "national technological university of lima",
        "national technological university of lima sur",
        "universidad tecnologica de lima",
        "universidad tecnologica de lima sur",
        "universidad nacional tecnologica de lima",
        "universidad nacional tecnologica de lima sur",
        "universidad nacional de ingenieria",
        "national university of engineering",
        "universidad tecnologica del peru",
        "technological university of peru",
        "universidad tecnologica de los andes",
        "universidad nacional mayor de san marcos",
        "national university of san marcos",
        "universidad nacional de san agustin",
        "national university of san agustin",
    ]

    if any(pattern in norm_flat for pattern in negative_patterns):
        return False

    # ULIMA explícito como token independiente.
    if re.search(r"\bulima\b", norm_flat):
        return True

    # Aceptar solo la forma exacta "Universidad de Lima" / "University of Lima".
    positive_regexes = [
        r"(?<!national technological )(?<!technological )\buniversity of lima\b",
        r"(?<!universidad nacional tecnologica )(?<!universidad tecnologica )\buniversidad de lima\b",
    ]

    return any(re.search(pattern, norm_flat) for pattern in positive_regexes)

def row_attr(row, attr: str, idx: int):
    try:
        return getattr(row, attr)
    except AttributeError:
        return row[idx]


def normalize_url_or_doi(value: str | None) -> str | None:
    if not value:
        return None

    raw = str(value).strip()
    if not raw:
        return None

    if raw.lower().startswith("http://") or raw.lower().startswith("https://"):
        return raw

    if raw.lower().startswith("doi:"):
        raw = raw[4:].strip()

    return raw


def build_doi_url(doi_value: str | None) -> str | None:
    normalized = normalize_url_or_doi(doi_value)
    if not normalized:
        return None

    if normalized.lower().startswith("http://") or normalized.lower().startswith("https://"):
        return normalized

    return f"https://doi.org/{normalized}"


def is_conference_document_type(document_type_value: str | None) -> bool:
    norm = normalize_generic_text(document_type_value)
    return "conference" in norm


def derive_conference_journal_value(document_type_value: str | None) -> str | None:
    if not document_type_value:
        return None
    return "Conference" if is_conference_document_type(document_type_value) else "Journal"


def derive_revista_and_conference(
    source_title_value: str | None,
    document_type_value: str | None,
) -> tuple[str | None, str | None]:
    if not source_title_value:
        return None, None

    if is_conference_document_type(document_type_value):
        return None, source_title_value

    return source_title_value, None


def build_authors_display(
    author_full_names_value: str | None,
    authors_value: str | None,
) -> str | None:
    full_names = [clean_author_full_name(x) for x in split_semicolon_values(author_full_names_value)]
    full_names = [x for x in full_names if x]
    if full_names:
        return "; ".join(full_names)

    short_names = [clean_author_full_name(x) for x in split_semicolon_values(authors_value)]
    short_names = [x for x in short_names if x]
    if short_names:
        return "; ".join(short_names)

    return None


def unique_keep_order(items: list[str]) -> list[str]:
    seen = set()
    result = []
    for item in items:
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result


def build_text_corpus(*values: str | None) -> str:
    parts: list[str] = []
    for value in values:
        norm = normalize_generic_text(value)
        if norm:
            parts.append(norm)
    return " | ".join(parts)


def phrase_in_text(text: str, phrase: str | None) -> bool:
    norm_phrase = normalize_generic_text(phrase)
    if not norm_phrase:
        return False
    return norm_phrase in text


def choose_best_scored_candidate(score_map: dict[str, int], min_score: int = 1) -> str | None:
    positives = [(k, v) for k, v in score_map.items() if v >= min_score]
    if not positives:
        return None

    positives.sort(key=lambda x: x[1], reverse=True)
    best_key, best_score = positives[0]

    tied = [k for k, v in positives if v == best_score]
    if len(tied) > 1:
        return None

    return best_key


def choose_best_scored_candidate_relaxed(score_map: dict[str, int], min_score: int = 1) -> str | None:
    positives = [(k, v) for k, v in score_map.items() if v >= min_score]
    if not positives:
        return None

    positives.sort(key=lambda x: (-x[1], normalize_generic_text(x[0])))
    return positives[0][0]


def choose_best_scored_candidate_with_margin(
    score_map: dict[str, int],
    min_score: int = 1,
    min_margin: int = 1,
) -> str | None:
    positives = [(k, v) for k, v in score_map.items() if v >= min_score]
    if not positives:
        return None

    positives.sort(key=lambda x: (-x[1], normalize_generic_text(x[0])))
    best_key, best_score = positives[0]

    if len(positives) == 1:
        return best_key

    second_score = positives[1][1]
    if (best_score - second_score) < min_margin:
        return None

    return best_key


THEMATIC_FIELD_WEIGHTS = {
    "abstract": 10,
    "title": 4,
    "author_keywords": 3,
    "index_keywords": 2,
    "source_title": 1,
}

THEMATIC_STRICT_MIN_SCORE = 7
THEMATIC_STRICT_MIN_MARGIN = 2
THEMATIC_APPROX_MIN_SCORE = 4


def build_thematic_text_fields(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict[str, str]:
    return {
        "title": normalize_generic_text(title_value),
        "abstract": normalize_generic_text(abstract_value),
        "author_keywords": normalize_generic_text(author_keywords_value),
        "index_keywords": normalize_generic_text(index_keywords_value),
        "source_title": normalize_generic_text(source_title_value),
    }


def score_alias_in_thematic_fields(
    text_fields: dict[str, str],
    alias: str | None,
    bonus: int = 0,
) -> int:
    alias_norm = normalize_generic_text(alias)
    if not alias_norm:
        return 0

    score = 0
    for field_name, text_value in text_fields.items():
        if text_value and phrase_in_text(text_value, alias_norm):
            score += THEMATIC_FIELD_WEIGHTS.get(field_name, 1) + bonus
    return score


def build_weighted_candidate_score(
    text_fields: dict[str, str],
    primary_aliases: list[str],
    support_aliases: list[str],
) -> int:
    score = 0
    seen: set[str] = set()

    for alias in primary_aliases:
        alias_norm = normalize_generic_text(alias)
        if alias_norm and alias_norm not in seen:
            score += score_alias_in_thematic_fields(text_fields, alias_norm, bonus=2)
            seen.add(alias_norm)

    for alias in support_aliases:
        alias_norm = normalize_generic_text(alias)
        if alias_norm and alias_norm not in seen:
            score += score_alias_in_thematic_fields(text_fields, alias_norm, bonus=0)
            seen.add(alias_norm)

    return score


def clip_text(value: str | None, max_chars: int = 7000) -> str | None:
    if not value:
        return None
    value = str(value).strip()
    if len(value) <= max_chars:
        return value
    return value[:max_chars]

CSV_DELIMITER_CANDIDATES = [",", ";", "\t", "|"]
CSV_PARSE_PREVIEW_ROWS = 25

SQL_SAFE_MAX_LENGTHS = {
    "eid": 64,
    "publication_year_raw": 16,
    "scopus_year_raw": 16,
    "conference_journal_raw": 32,
    "indexation_raw": 64,
    "issn_raw": 64,
    "isbn_raw": 64,
    "es_scopus_raw": 16,
    "retractado_raw": 16,
    "descontinuado_scopus_raw": 64,
    "first_author_ulima_raw": 16,
    "metodo_cruce_scopus_raw": 64,
}

ISSN_REGEX = re.compile(r"\b\d{4}-?\d{3}[\dXx]\b")
ISBN_REGEX = re.compile(r"\b(?:97[89][-\s]?)?(?:\d[-\s]?){9,12}[\dXx]\b")
EID_REGEX = re.compile(r"\b2-s2\.0-\d+\b", flags=re.IGNORECASE)
YEAR_REGEX = re.compile(r"^\d{4}$")


def clean_csv_header_value(value: str | None) -> str:
    if value is None:
        return ""
    return str(value).replace("\ufeff", "").strip()


def strip_excel_separator_hint(csv_text: str) -> tuple[str, str | None]:
    if not csv_text:
        return csv_text, None

    lines = csv_text.splitlines()
    if not lines:
        return csv_text, None

    first_line = lines[0].strip()
    match = re.match(r"^sep=(.)$", first_line, flags=re.IGNORECASE)
    if not match:
        return csv_text, None

    delimiter = match.group(1)
    remaining_text = "\n".join(lines[1:])
    return remaining_text, delimiter


def parse_csv_rows_with_delimiter(csv_text: str, delimiter: str) -> list[list[str]]:
    reader = csv.reader(
        io.StringIO(csv_text),
        delimiter=delimiter,
        quotechar='"',
        doublequote=True,
        skipinitialspace=False,
    )

    rows: list[list[str]] = []
    for row in reader:
        cleaned_row = [str(cell) if cell is not None else "" for cell in row]
        if any(str(cell).strip() for cell in cleaned_row):
            rows.append(cleaned_row)

    return rows


def get_expected_csv_header_aliases() -> set[str]:
    aliases: set[str] = set()
    for alias_list in COLUMN_ALIASES.values():
        for alias in alias_list:
            aliases.add(normalize_generic_text(alias))

    extra_headers = [
        "Author full names",
        "Authors with affiliations",
        "Publisher",
        "Publication Stage",
        "Source title",
        "Document Type",
        "Affiliations",
    ]
    for header in extra_headers:
        aliases.add(normalize_generic_text(header))

    return aliases


def score_csv_header_match(headers: list[str]) -> int:
    expected_aliases = get_expected_csv_header_aliases()
    normalized_headers = [normalize_generic_text(clean_csv_header_value(h)) for h in headers]
    return sum(1 for h in normalized_headers if h in expected_aliases)


def score_csv_candidate(rows: list[list[str]]) -> int:
    if not rows:
        return -10_000

    headers = rows[0]
    header_count = len(headers)
    if header_count == 0:
        return -10_000

    score = 0
    score += score_csv_header_match(headers) * 100

    preview_rows = rows[1: 1 + CSV_PARSE_PREVIEW_ROWS]
    if not preview_rows:
        return score

    for row in preview_rows:
        if len(row) == header_count:
            score += 8
        elif abs(len(row) - header_count) == 1:
            score += 2
        else:
            score -= 10

    return score


def normalize_csv_row_length(row: list[str], headers: list[str]) -> list[str]:
    expected = len(headers)
    values = list(row)

    if len(values) < expected:
        values.extend([""] * (expected - len(values)))
        return values

    if len(values) > expected:
        prefix = values[: expected - 1]
        merged_tail = ",".join(v for v in values[expected - 1:] if v is not None)
        return prefix + [merged_tail]

    return values


def build_dict_rows_from_csv_rows(rows: list[list[str]]) -> list[dict]:
    if not rows:
        return []

    headers = [clean_csv_header_value(h) for h in rows[0]]
    dict_rows: list[dict] = []

    for raw_row in rows[1:]:
        normalized_row = normalize_csv_row_length(raw_row, headers)
        row_dict = {headers[idx]: normalized_row[idx].strip() for idx in range(len(headers))}
        dict_rows.append(row_dict)

    return dict_rows


def extract_first_regex_match(value: str | None, pattern: re.Pattern[str]) -> str | None:
    if not value:
        return None
    match = pattern.search(str(value))
    if not match:
        return None
    return match.group(0).strip()


def extract_all_regex_matches(value: str | None, pattern: re.Pattern[str]) -> list[str]:
    if not value:
        return []
    matches = [m.group(0).strip() for m in pattern.finditer(str(value))]
    return unique_keep_order(matches)


def sanitize_short_field(value: str | None, max_length: int) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    if not cleaned:
        return None
    if len(cleaned) <= max_length:
        return cleaned
    return cleaned[:max_length]


def sanitize_year_field(value: str | None) -> str | None:
    if not value:
        return None
    cleaned = str(value).strip()
    if YEAR_REGEX.match(cleaned):
        return cleaned

    extracted = extract_first_regex_match(cleaned, re.compile(r"\b\d{4}\b"))
    return extracted


def sanitize_identifier_fields(mapped: dict) -> dict:
    sanitized = dict(mapped)

    sanitized["eid"] = extract_first_regex_match(sanitized.get("eid"), EID_REGEX) or sanitize_short_field(
        sanitized.get("eid"), SQL_SAFE_MAX_LENGTHS["eid"]
    )

    sanitized["publication_year_raw"] = sanitize_year_field(sanitized.get("publication_year_raw"))
    sanitized["scopus_year_raw"] = sanitize_year_field(sanitized.get("scopus_year_raw"))

    issn_matches = extract_all_regex_matches(sanitized.get("issn_raw"), ISSN_REGEX)
    sanitized["issn_raw"] = "; ".join(issn_matches) if issn_matches else sanitize_short_field(
        sanitized.get("issn_raw"), SQL_SAFE_MAX_LENGTHS["issn_raw"]
    )

    isbn_matches = extract_all_regex_matches(sanitized.get("isbn_raw"), ISBN_REGEX)
    sanitized["isbn_raw"] = "; ".join(isbn_matches) if isbn_matches else sanitize_short_field(
        sanitized.get("isbn_raw"), SQL_SAFE_MAX_LENGTHS["isbn_raw"]
    )

    for field_name, max_length in SQL_SAFE_MAX_LENGTHS.items():
        if field_name in {"eid", "publication_year_raw", "scopus_year_raw", "issn_raw", "isbn_raw"}:
            continue
        sanitized[field_name] = sanitize_short_field(sanitized.get(field_name), max_length)

    return sanitized



def coerce_choice(value: str | None, allowed_values: list[str]) -> str | None:
    if not value:
        return None

    norm_value = normalize_generic_text(value)
    if not norm_value:
        return None

    for allowed in allowed_values:
        if normalize_generic_text(allowed) == norm_value:
            return allowed

    return None


def parse_float_or_none(value) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def extract_json_object(text: str | None) -> dict | None:
    if not text:
        return None

    candidate = str(text).strip()
    if candidate.startswith("```"):
        candidate = re.sub(r"^```(?:json)?", "", candidate.strip(), flags=re.IGNORECASE).strip()
        candidate = re.sub(r"```$", "", candidate.strip()).strip()

    try:
        loaded = json.loads(candidate)
        return loaded if isinstance(loaded, dict) else None
    except Exception:
        pass

    match = re.search(r"\{.*\}", candidate, flags=re.DOTALL)
    if not match:
        return None

    try:
        loaded = json.loads(match.group(0))
        return loaded if isinstance(loaded, dict) else None
    except Exception:
        return None


def merge_non_null_fields(base: dict, filler: dict, fields: list[str]) -> dict:
    result = dict(base)
    for field in fields:
        if not result.get(field) and filler.get(field):
            result[field] = filler[field]
    return result


def has_any_thematic_field(payload: dict) -> bool:
    return any(
        [
            payload.get("area_carrera_raw"),
            payload.get("linea_carrera_raw"),
            payload.get("category_tematica_raw"),
            payload.get("area_idic_raw"),
            payload.get("linea_idic_raw"),
        ]
    )


def missing_thematic_fields(payload: dict) -> list[str]:
    fields = [
        "area_carrera_raw",
        "linea_carrera_raw",
        "category_tematica_raw",
        "area_idic_raw",
        "linea_idic_raw",
    ]
    return [f for f in fields if not payload.get(f)]


# =========================
# CAREER HINTS FROM TEXT
# =========================
CAREER_PATTERNS = {
    "Ingeniería Industrial": [
        "carrera de ingenieria industrial",
        "programa de ingenieria industrial",
        "industrial engineering career",
        "industrial engineering program",
        "ingenieria industrial",
        "ingenieria industiral",
        "ingenieria indutrial",
        "ingenieria industrial y",
        "industrial engineering",
        "industrial and systems engineering",
    ],
    "Ingeniería de Sistemas": [
        "carrera de ingenieria de sistemas",
        "programa de ingenieria de sistemas",
        "systems engineering career",
        "systems engineering program",
        "ingenieria de sistemas",
        "ingenieria de sistemas computacionales",
        "ingenieria informatica",
        "systems engineering",
        "computer systems engineering",
        "information systems engineering",
        "software engineering",
    ],
    "Ingeniería Civil": [
        "carrera de ingenieria civil",
        "programa de ingenieria civil",
        "civil engineering career",
        "civil engineering program",
        "ingenieria civil",
        "civil engineering",
    ],
}


def infer_careers_from_text(value: str | None) -> list[str]:
    norm = normalize_generic_text(value)
    found: list[str] = []
    for career, patterns in CAREER_PATTERNS.items():
        if any(pattern in norm for pattern in patterns):
            found.append(career)
    return found


def infer_career_from_text(value: str | None) -> str | None:
    careers = infer_careers_from_text(value)
    return careers[0] if careers else None


VALID_ENGINEERING_CAREERS = [
    "Ingeniería Industrial",
    "Ingeniería Civil",
    "Ingeniería de Sistemas",
]

VALID_SCOPUS_DOCUMENT_TYPES = {
    "article",
    "conference paper",
}

ULIMA_ENGINEERING_CONTEXT_HINTS = [
    "facultad de ingenieria",
    "facultad de ingenieria y arquitectura",
    "facultad de lngenieria",
    "facultad de ingeneria",
    "facultad de ingenieri",
    "faculty of engineering",
    "faculty of engineering and architecture",
    "engineering faculty",
    "school of engineering",
    "school of engineering and architecture",
    "engineering school",
    "escuela de ingenieria",
    "department of engineering",
    "departamento de ingenieria",
    "carrera de ingenieria",
    "programa de ingenieria",
    "engineering department",
]

EXTERNAL_INSTITUTION_BOUNDARY_HINTS = [
    "universidad",
    "university",
    "college",
    "polytechnic",
    "politecnico",
]


def filter_valid_engineering_careers(careers: list[str] | None) -> list[str]:
    if not careers:
        return []
    return unique_keep_order([c for c in careers if c in VALID_ENGINEERING_CAREERS])


def normalize_document_type_for_filter(value: str | None) -> str:
    return normalize_generic_text(value)


def is_valid_scopus_document_type(value: str | None) -> bool:
    return normalize_document_type_for_filter(value) in VALID_SCOPUS_DOCUMENT_TYPES


def split_affiliation_clauses(value: str | None) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in str(value).split(",") if part and str(part).strip()]



def has_external_institution_boundary(value: str | None) -> bool:
    norm = normalize_generic_text(value)
    if not norm or is_ulima_text(value):
        return False
    return any(hint in norm for hint in EXTERNAL_INSTITUTION_BOUNDARY_HINTS)



def extract_ulima_local_contexts(
    value: str | None,
    max_lookback: int = 3,
    max_lookahead: int = 6,
) -> list[str]:
    """
    Extrae el contexto local alrededor de una afiliación ULima.

    Importante para Scopus: algunas afiliaciones vienen como:
        "Universidad de Lima, ..., Carrera de Ingeniería Industrial, Facultad de Ingeniería"
    Es decir, la carrera/facultad aparece DESPUÉS de "Universidad de Lima".

    La versión anterior solo miraba hacia atrás, por lo que podía rechazar falsamente
    artículos válidos de Ingeniería ULima cuando la carrera aparecía hacia adelante
    dentro del mismo bloque de autor/afiliación.
    """
    if not value or not is_ulima_text(value):
        return []

    clauses = split_affiliation_clauses(value)
    if not clauses:
        return []

    contexts: list[str] = []
    for idx, clause in enumerate(clauses):
        if not is_ulima_text(clause):
            continue

        # Mirada hacia atrás: útil para patrones tipo
        # "Carrera de Ingeniería Industrial, Universidad de Lima".
        start_idx = idx
        steps = 0
        while start_idx > 0 and steps < max_lookback:
            previous_clause = clauses[start_idx - 1]
            if has_external_institution_boundary(previous_clause):
                break
            start_idx -= 1
            steps += 1

        # Mirada hacia adelante: necesaria para patrones tipo
        # "Universidad de Lima, ..., Carrera de Ingeniería Industrial, Facultad de Ingeniería".
        end_idx = idx
        steps = 0
        while end_idx + 1 < len(clauses) and steps < max_lookahead:
            next_clause = clauses[end_idx + 1]
            if has_external_institution_boundary(next_clause):
                break
            end_idx += 1
            steps += 1

        context = ", ".join(clauses[start_idx : end_idx + 1]).strip(" ,")
        if context:
            contexts.append(context)

    return unique_keep_order(contexts)



def extract_ulima_local_contexts_from_values(*values: str | None) -> list[str]:
    contexts: list[str] = []
    for value in values:
        if not value:
            continue
        parts = split_semicolon_values(value)
        if not parts:
            parts = [str(value)]
        for part in parts:
            contexts.extend(extract_ulima_local_contexts(part))
    return unique_keep_order(contexts)



def is_ulima_engineering_context(value: str | None) -> bool:
    if not value or not is_ulima_text(value):
        return False

    if filter_valid_engineering_careers(infer_careers_from_text(value)):
        return True

    norm = normalize_generic_text(value)
    return any(hint in norm for hint in ULIMA_ENGINEERING_CONTEXT_HINTS)



def resolve_engineering_affiliation_details(*values: str | None) -> dict:
    ulima_contexts = extract_ulima_local_contexts_from_values(*values)
    detected_careers: list[str] = []
    has_engineering_context = False

    for context in ulima_contexts:
        detected_careers.extend(filter_valid_engineering_careers(infer_careers_from_text(context)))
        if is_ulima_engineering_context(context):
            has_engineering_context = True

    return {
        "ulima_contexts": ulima_contexts,
        "careers": unique_keep_order(detected_careers),
        "has_ulima_affiliation": bool(ulima_contexts),
        "has_ulima_engineering_context": has_engineering_context,
    }



def resolve_engineering_careers_from_affiliation_text(*values: str | None) -> list[str]:
    return resolve_engineering_affiliation_details(*values).get("careers", [])




# =========================
# GENERIC ENGINEERING CAREER INFERENCE
# =========================
# Fallback heurístico conservador. Cuando career_ai=true, la IA decide los
# casos ambiguos de Facultad de Ingeniería genérica. Si la IA está apagada,
# estas constantes evitan rescates automáticos demasiado agresivos.
CAREER_INFERENCE_MIN_SCORE = int(os.environ.get("CAREER_INFERENCE_MIN_SCORE", "12"))
CAREER_INFERENCE_MIN_MARGIN = int(os.environ.get("CAREER_INFERENCE_MIN_MARGIN", "4"))

CAREER_INFERENCE_SIGNALS = {
    "Ingeniería Industrial": [
        "lean", "lean manufacturing", "5s", "smed", "tpm", "slp", "kanban", "kaizen",
        "bpm", "dmaic", "tqm", "oee", "abc classification", "standard work",
        "standardized work", "warehouse", "warehousing", "inventory", "inventory management",
        "supply chain", "logistics", "transportation", "transport", "production management",
        "production efficiency", "production planning", "manufacturing", "productivity",
        "productivity improvement", "process improvement", "continuous improvement",
        "quality control", "quality management", "operations management", "operational efficiency",
        "eoq", "mrp", "lot sizing", "service level", "otif", "fill rate", "picking",
        "slotting", "layout", "facility layout", "work study", "ergonomic risk",
    ],
    "Ingeniería Civil": [
        "bim", "building information modelling", "building information modeling",
        "construction", "building", "structural", "structure", "seismic", "earthquake",
        "concrete", "reinforced concrete", "beam", "beams", "masonry", "asphalt",
        "pavement", "stone mastic asphalt", "sma", "soil", "geotechnical",
        "geotechnics", "hydraulic", "hydrology", "drainage", "irrigation", "sediment",
        "water resources", "bridge", "housing", "infrastructure", "urban regeneration",
        "photogrammetry", "uav", "road", "road infrastructure", "traffic",
    ],
    "Ingeniería de Sistemas": [
        "machine learning", "deep learning", "artificial intelligence", "computer vision",
        "image processing", "object detection", "human detection", "nlp",
        "natural language processing", "large language model", "llm", "software",
        "software engineering", "information systems", "data mining", "process mining",
        "algorithm", "algorithms", "neural network", "neural networks", "iot",
        "internet of things", "cybersecurity", "cloud", "app", "mobile application",
        "web application", "database", "data science", "predictive model", "classification model",
    ],
}

INDUSTRIAL_CONTEXT_FOR_ML_SIGNALS = [
    "pyrolysis", "crude oil yield", "production", "manufacturing", "supply chain",
    "inventory", "warehouse", "logistics", "operations", "process", "quality",
    "productivity", "yield", "optimization", "operational",
]

SYSTEMS_CONTEXT_FOR_ML_SIGNALS = [
    "computer vision", "image", "images", "video", "object detection", "human detection",
    "nlp", "software", "cybersecurity", "iot", "cloud", "information systems",
]


def score_career_inference_from_text_fields(text_fields: dict[str, str]) -> dict[str, int]:
    """
    Puntúa carrera objetivo cuando la afiliación ULima dice Ingeniería genérica
    pero no especifica Industrial/Civil/Sistemas.

    No reemplaza una carrera explícita de afiliación/docente. Solo rescata casos
    que antes se rechazaban injustamente por falta de carrera explícita.
    """
    score_map = {career: 0 for career in VALID_ENGINEERING_CAREERS}

    field_weights = {
        "title": 4,
        "abstract": 2,
        "author_keywords": 3,
        "index_keywords": 2,
        "source_title": 1,
    }

    for career, signals in CAREER_INFERENCE_SIGNALS.items():
        for field_name, text_value in text_fields.items():
            if not text_value:
                continue
            weight = field_weights.get(field_name, 1)
            score_map[career] += weight * count_phrase_hits_in_text(text_value, signals)

    corpus = " | ".join(v for v in text_fields.values() if v)
    ml_present = count_phrase_hits_in_text(corpus, ["machine learning", "deep learning", "predictive model", "predictive models"]) > 0
    if ml_present:
        if count_phrase_hits_in_text(corpus, INDUSTRIAL_CONTEXT_FOR_ML_SIGNALS) > 0:
            score_map["Ingeniería Industrial"] += 4
        if count_phrase_hits_in_text(corpus, SYSTEMS_CONTEXT_FOR_ML_SIGNALS) > 0:
            score_map["Ingeniería de Sistemas"] += 4

    return score_map


def infer_target_career_from_generic_engineering_context(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict:
    text_fields = build_thematic_text_fields(
        title_value,
        abstract_value,
        author_keywords_value,
        index_keywords_value,
        source_title_value,
    )

    if not any(text_fields.values()):
        return {
            "careers": [],
            "score_map": {career: 0 for career in VALID_ENGINEERING_CAREERS},
            "reason": "NO_THEMATIC_TEXT",
        }

    score_map = score_career_inference_from_text_fields(text_fields)
    ranked = sorted(score_map.items(), key=lambda x: (-x[1], normalize_generic_text(x[0])))

    if not ranked or ranked[0][1] < CAREER_INFERENCE_MIN_SCORE:
        return {
            "careers": [],
            "score_map": score_map,
            "reason": "INSUFFICIENT_SCORE",
        }

    best_career, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0

    if (best_score - second_score) < CAREER_INFERENCE_MIN_MARGIN:
        return {
            "careers": [],
            "score_map": score_map,
            "reason": "AMBIGUOUS_SCORE",
        }

    return {
        "careers": [best_career],
        "score_map": score_map,
        "reason": "INFERRED_FROM_THEMATIC_EVIDENCE",
    }


def serialize_score_map_for_reason(score_map: dict[str, int] | None) -> str:
    score_map = score_map or {}
    return "; ".join(f"{career}={score_map.get(career, 0)}" for career in VALID_ENGINEERING_CAREERS)


def build_generic_engineering_career_ai_prompt(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    authors_value: str | None,
    ulima_docentes_value: str | None,
    ulima_contexts: list[str] | None,
    heuristic_score_map: dict[str, int] | None,
) -> str:
    """
    Prompt especializado para decidir carrera objetivo SOLO en casos ambiguos:
    ULima + Facultad/Escuela de Ingeniería genérica, sin carrera explícita.
    """
    payload = {
        "task": "resolve_target_engineering_career_for_ulima_generic_engineering_affiliation",
        "allowed_careers": VALID_ENGINEERING_CAREERS,
        "allowed_decisions": ["ACCEPT", "REVIEW", "REJECT_NOT_TARGET"],
        "article": {
            "title": clip_text(title_value, 2200),
            "abstract_scopus": clip_text(abstract_value, 9000),
            "author_keywords": clip_text(author_keywords_value, 2500),
            "index_keywords": clip_text(index_keywords_value, 2500),
            "source_title": clip_text(source_title_value, 1200),
            "authors": clip_text(authors_value, 2200),
            "ulima_docentes_detected": clip_text(ulima_docentes_value, 2200),
            "ulima_affiliation_contexts": [clip_text(x, 2500) for x in (ulima_contexts or [])],
            "heuristic_scores": heuristic_score_map or {},
        },
        "output_schema": {
            "decision": "ACCEPT|REVIEW|REJECT_NOT_TARGET",
            "career": "Ingeniería Industrial|Ingeniería Civil|Ingeniería de Sistemas|null",
            "confidence": "number 0..1",
            "evidence_level": "explicit_affiliation|docente_ref|strong_thematic|weak_or_ambiguous|out_of_scope",
            "rationale": "short Spanish explanation, max 35 words",
            "red_flags": ["short strings"],
        },
        "rules": [
            "El caso ya tiene afiliación real a Universidad de Lima y contexto genérico de Facultad/Escuela de Ingeniería, pero NO tiene carrera explícita.",
            "Debes decidir si corresponde a una de estas carreras: Ingeniería Industrial, Ingeniería Civil o Ingeniería de Sistemas.",
            "Usa el abstract como evidencia principal; usa título, keywords, source title, autores y afiliación como apoyo.",
            "La frase 'Facultad de Ingeniería' o 'Faculty of Engineering' por sí sola NO basta para aceptar una carrera.",
            "ACCEPT solo si hay evidencia temática fuerte y claramente dominante para una carrera.",
            "REVIEW si la evidencia es débil, mixta, genérica, educativa, metodológica, estadística o no permite distinguir carrera con seguridad.",
            "REJECT_NOT_TARGET si el artículo claramente pertenece a física, cosmología, medicina, arquitectura pura, educación general, diseño, química/materiales genéricos u otro campo sin vínculo defendible con Industrial/Civil/Sistemas.",
            "Ingeniería Industrial: lean, 5S, TPM, SMED, SLP, warehouse, inventory, supply chain, logistics, production, operations, productivity, quality, safety/ergonomics, commercial/CRM/process improvement, plant feasibility.",
            "Ingeniería Civil: BIM/VDC para construcción, structural/concrete/seismic/building/pavement/asphalt/geotechnics/water resources/drainage/hydraulics/construction/housing/infrastructure.",
            "Ingeniería de Sistemas: machine learning/deep learning/computer vision/NLP/software engineering/cybersecurity/IoT/cloud/databases/algorithms/information systems. No uses Sistemas solo por decir digital/online/interface/education.",
            "Si la mejor carrera y la segunda están muy cercanas, devuelve REVIEW, no ACCEPT.",
            "Devuelve SOLO JSON válido, sin markdown. No inventes campos.",
        ],
    }
    return (
        "Eres un revisor académico senior de producción científica de la Facultad de Ingeniería de la Universidad de Lima. "
        "Tu tarea es resolver carreras ambiguas con criterio conservador y auditable.\n\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )


def call_openai_responses_json(
    prompt: str,
    throttle_seconds: float = 0.0,
    call_label: str = "azure_openai",
) -> dict | None:
    if not is_thematic_llm_configured():
        return None

    max_retries = get_openai_max_retries()
    base_seconds = get_openai_retry_base_seconds()
    last_error: Exception | None = None

    for attempt in range(max_retries + 1):
        try:
            throttle_azure_openai_call(throttle_seconds, label=call_label)

            client = get_azure_openai_client()
            response = client.responses.create(
                model=get_env("AZURE_OPENAI_RESPONSES_MODEL"),
                input=prompt,
            )
            return extract_json_object(getattr(response, "output_text", None))
        except Exception as exc:
            last_error = exc
            message = str(exc).lower()
            retryable = any(
                token in message
                for token in ["429", "rate limit", "too many requests", "timeout", "temporarily", "503", "502"]
            )
            if attempt >= max_retries or not retryable:
                break

            retry_after = extract_retry_after_seconds_from_exception(exc)
            if retry_after is not None:
                sleep_seconds = min(120.0, retry_after + 1.0)
            else:
                sleep_seconds = min(120.0, base_seconds * (2 ** attempt))

            if "429" in message or "rate limit" in message or "too many requests" in message:
                sleep_seconds = max(sleep_seconds, get_career_ai_throttle_after_429_seconds())

            logging.warning(
                "Azure OpenAI retryable error for %s on attempt %s/%s: %s. Sleeping %.1fs",
                call_label,
                attempt + 1,
                max_retries + 1,
                str(exc),
                sleep_seconds,
            )
            time.sleep(sleep_seconds)

    logging.warning(
        "Azure OpenAI JSON call failed for %s after retries: %s",
        call_label,
        str(last_error) if last_error else "unknown",
    )
    return None


def classify_generic_engineering_target_career_with_ai(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    authors_value: str | None,
    ulima_docentes_value: str | None,
    ulima_contexts: list[str] | None,
    heuristic_score_map: dict[str, int] | None,
) -> dict:
    """
    Usa IA solo para decidir carrera objetivo en afiliaciones ULima + Ingeniería genérica.
    Resultado conservador: si no hay ACCEPT con confianza suficiente, queda para REVIEW.
    """
    empty = {
        "careers": [],
        "decision": "NO_AI_RESULT",
        "confidence": None,
        "reason": "NO_AI_RESULT",
        "rationale": None,
        "evidence_level": None,
        "raw": None,
    }

    if not is_thematic_llm_configured():
        empty["decision"] = "AI_NOT_CONFIGURED"
        empty["reason"] = "AI_NOT_CONFIGURED"
        return empty

    call_slot = reserve_career_ai_call_slot()
    if not call_slot.get("allowed"):
        empty["decision"] = "AI_SKIPPED_MAX_CALLS_PER_RUN"
        empty["reason"] = (
            f"AI_SKIPPED_MAX_CALLS_PER_RUN_{call_slot.get('calls_used')}_OF_{call_slot.get('max_calls')}"
        )
        return empty

    prompt = build_generic_engineering_career_ai_prompt(
        title_value=title_value,
        abstract_value=abstract_value,
        author_keywords_value=author_keywords_value,
        index_keywords_value=index_keywords_value,
        source_title_value=source_title_value,
        authors_value=authors_value,
        ulima_docentes_value=ulima_docentes_value,
        ulima_contexts=ulima_contexts,
        heuristic_score_map=heuristic_score_map,
    )
    raw_output = call_openai_responses_json(
        prompt,
        throttle_seconds=get_career_ai_call_delay_seconds(),
        call_label="career_ambiguity_ai",
    )
    if not raw_output:
        empty["decision"] = "AI_FAILED_OR_NON_JSON"
        empty["reason"] = "AI_FAILED_OR_NON_JSON"
        return empty

    decision = str(raw_output.get("decision") or "").strip().upper()
    career = coerce_choice(raw_output.get("career"), VALID_ENGINEERING_CAREERS)
    confidence = parse_float_or_none(raw_output.get("confidence"))
    rationale = str(raw_output.get("rationale") or "").strip() or None
    evidence_level = str(raw_output.get("evidence_level") or "").strip() or None

    result = {
        "careers": [],
        "decision": decision or "INVALID_DECISION",
        "confidence": confidence,
        "reason": decision or "INVALID_DECISION",
        "rationale": rationale,
        "evidence_level": evidence_level,
        "raw": raw_output,
    }

    min_confidence = get_career_ambiguity_llm_min_confidence()
    if decision == "ACCEPT" and career and confidence is not None and confidence >= min_confidence:
        result["careers"] = [career]
        result["reason"] = "AI_ACCEPTED_TARGET_CAREER"
        return result

    if decision == "ACCEPT" and career:
        result["reason"] = f"AI_LOW_CONFIDENCE_ACCEPT_BELOW_THRESHOLD_{min_confidence:.2f}"
        return result

    if decision in {"REVIEW", "REJECT_NOT_TARGET"}:
        result["reason"] = f"AI_{decision}"
        return result

    result["reason"] = "AI_INVALID_OR_INCOMPLETE_OUTPUT"
    return result


def format_ai_review_reason(prefix: str, ai_result: dict, score_map: dict[str, int] | None) -> str:
    confidence = ai_result.get("confidence")
    confidence_text = "NULL" if confidence is None else f"{float(confidence):.2f}"
    rationale = ai_result.get("rationale") or "No rationale returned"
    evidence = ai_result.get("evidence_level") or "unknown"
    return (
        f"{prefix}: AI decision did not produce a safe accepted career. "
        f"ai_reason={ai_result.get('reason')}; confidence={confidence_text}; evidence_level={evidence}; "
        f"scores={serialize_score_map_for_reason(score_map)}; rationale={rationale[:500]}"
    )


GENERIC_ENGINEERING_PHYSICS_REVIEW_SIGNALS = [
    "magnetogenesis", "electroweak", "cosmology", "cosmological", "supercooled",
    "phase transition", "primordial", "particle physics", "high energy physics",
    "journal of high energy physics", "string theory",
]

GENERIC_ENGINEERING_EDUCATION_REVIEW_SIGNALS = [
    "research training", "methodological design", "engineering students",
    "early childhood", "online education", "preschool", "curriculum",
    "education", "educational", "higher education", "learning outcomes",
]

GENERIC_ENGINEERING_MEDICAL_REVIEW_SIGNALS = [
    "brachytherapy", "radiotherapy", "clinical", "patient", "patients",
    "cancer", "dose distribution", "treatment planning", "medical device",
]

GENERIC_ENGINEERING_MATERIAL_WATER_REVIEW_SIGNALS = [
    "photoelectrocatalytic", "photoelectrocatalysis", "water disinfection",
    "water treatment", "arsenic removal", "adsorption", "activated carbon",
    "chitosan", "carbon nanotubes", "electrode", "photoanode", "azo dyes",
]


def has_strong_systems_evidence(text_fields: dict[str, str]) -> bool:
    return contains_any_phrase_in_text_fields(
        text_fields,
        SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS
        + SYSTEMS_ML_STRONG_SIGNALS
        + SYSTEMS_CYBER_STRONG_SIGNALS
        + SYSTEMS_GAME_STRONG_SIGNALS
        + SYSTEMS_ARVR_STRONG_SIGNALS
        + SYSTEMS_AGENT_STRONG_SIGNALS
        + SYSTEMS_SOFTWARE_STRONG_SIGNALS
        + SYSTEMS_GENDER_TECH_STRONG_SIGNALS,
    )


def has_strong_industrial_evidence(text_fields: dict[str, str]) -> bool:
    return contains_any_phrase_in_text_fields(
        text_fields,
        INDUSTRIAL_OEM_STRONG_SIGNALS
        + INDUSTRIAL_SCM_STRONG_SIGNALS
        + INDUSTRIAL_ORA_STRONG_SIGNALS
        + INDUSTRIAL_PDD_STRONG_SIGNALS
        + INDUSTRIAL_EXTENDED_STRONG_SIGNALS
        + CHEM_MATERIAL_PRODUCT_STRONG_SIGNALS,
    )


def has_strong_civil_evidence(text_fields: dict[str, str]) -> bool:
    return contains_any_phrase_in_text_fields(
        text_fields,
        CIVIL_STRUCTURAL_STRONG_SIGNALS
        + CIVIL_CONSTRUCTION_MATERIALS_STRONG_SIGNALS
        + CIVIL_BIM_VDC_STRONG_SIGNALS
        + CIVIL_WATER_STRONG_SIGNALS
        + CIVIL_TRANSPORT_GEOTECH_STRONG_SIGNALS,
    )


def generic_engineering_heuristic_review_reason(
    inferred_careers: list[str],
    text_fields: dict[str, str],
    score_map: dict[str, int] | None = None,
) -> str | None:
    """
    Filtro conservador para afiliaciones ULima + Facultad/Escuela de Ingeniería
    sin carrera explícita cuando NO se usa IA.

    Objetivo: evitar que la heurística fuerce Sistemas/VR/Visión/Software o carreras
    débiles en papers de educación, física, medicina, materiales o agua.
    """
    careers = filter_valid_engineering_careers(inferred_careers)
    if not careers:
        return "NO_VALID_INFERRED_CAREER"

    score_detail = serialize_score_map_for_reason(score_map or {})
    systems_evidence = has_strong_systems_evidence(text_fields)
    industrial_evidence = has_strong_industrial_evidence(text_fields)
    civil_evidence = has_strong_civil_evidence(text_fields)

    # Señales claramente fuera del alcance o demasiado ambiguas para decidir carrera por reglas.
    if contains_any_phrase_in_text_fields(text_fields, GENERIC_ENGINEERING_PHYSICS_REVIEW_SIGNALS):
        return f"UNSAFE_HEURISTIC_PHYSICS_OR_COSMOLOGY; scores={score_detail}"

    if contains_any_phrase_in_text_fields(text_fields, GENERIC_ENGINEERING_MEDICAL_REVIEW_SIGNALS):
        return f"UNSAFE_HEURISTIC_MEDICAL_OR_CLINICAL_DEVICE; scores={score_detail}"

    # Educación puede ser de Sistemas/Civil/Industrial, pero la carrera no se debe forzar
    # salvo que haya evidencia técnica fuerte adicional.
    if contains_any_phrase_in_text_fields(text_fields, GENERIC_ENGINEERING_EDUCATION_REVIEW_SIGNALS):
        if not (systems_evidence or industrial_evidence or civil_evidence):
            return f"UNSAFE_HEURISTIC_EDUCATION_GENERIC; scores={score_detail}"

    # Materiales/agua no debe terminar como Sistemas por palabras sueltas de diseño/dispositivo.
    if "Ingeniería de Sistemas" in careers and contains_any_phrase_in_text_fields(text_fields, GENERIC_ENGINEERING_MATERIAL_WATER_REVIEW_SIGNALS):
        if not systems_evidence:
            return f"UNSAFE_HEURISTIC_SYSTEMS_ASSIGNED_TO_MATERIALS_OR_WATER; scores={score_detail}"

    if "Ingeniería de Sistemas" in careers and not systems_evidence:
        return f"UNSAFE_HEURISTIC_SYSTEMS_WITHOUT_STRONG_SYSTEMS_EVIDENCE; scores={score_detail}"

    if "Ingeniería Industrial" in careers and not industrial_evidence:
        return f"UNSAFE_HEURISTIC_INDUSTRIAL_WITHOUT_STRONG_INDUSTRIAL_EVIDENCE; scores={score_detail}"

    if "Ingeniería Civil" in careers and not civil_evidence:
        return f"UNSAFE_HEURISTIC_CIVIL_WITHOUT_STRONG_CIVIL_EVIDENCE; scores={score_detail}"

    return None


def choose_primary_career_for_classification(carrera_raw: str | None, text_fields: dict[str, str]) -> str | None:
    """
    La publicación puede tener multicarrera en carrera_raw. Para no duplicar registros,
    se conserva carrera_raw como lista separada por ';'. Esta función solo elige una
    carrera primaria para clasificar area_carrera/linea_carrera de la fila principal.

    El análisis por carrera en Power BI debe resolverse después con una tabla puente,
    no duplicando autores ni publicaciones en curated.publications.
    """
    careers = filter_valid_engineering_careers(split_semicolon_values(carrera_raw))
    if not careers:
        return None
    if len(careers) == 1:
        return careers[0]

    score_map = score_career_inference_from_text_fields(text_fields)
    ranked = sorted(careers, key=lambda c: (-score_map.get(c, 0), careers.index(c)))
    return ranked[0] if ranked else careers[0]




def career_has_strong_thematic_evidence_for_docente_ref(career: str, text_fields: dict[str, str]) -> bool:
    """
    Evidencia mínima para rescatar casos ULima + DOCENTES_REF_ONLY cuando la
    afiliación local no declara Facultad/Carrera de Ingeniería.

    Esta regla NO reemplaza la afiliación explícita. Solo evita falsos negativos
    cuando un docente de Industrial/Civil/Sistemas aparece en ref.docentes_ulima
    y el título/abstract/keywords respaldan claramente su carrera.
    """
    if career == "Ingeniería Industrial":
        return has_strong_industrial_evidence(text_fields)
    if career == "Ingeniería Civil":
        return has_strong_civil_evidence(text_fields)
    if career == "Ingeniería de Sistemas":
        return has_strong_systems_evidence(text_fields)
    return False


def resolve_docentes_ref_strong_thematic_rescue(
    docente_ref_careers: list[str] | None,
    text_fields: dict[str, str],
) -> dict:
    """
    Decide si un registro rechazado por ausencia de contexto explícito de
    ingeniería puede recuperarse por DOCENTES_REF + evidencia temática fuerte.

    Resultado conservador:
    - ACCEPT si hay una carrera docente con evidencia fuerte y dominante.
    - REVIEW si hay empate/multicarrera ambigua o evidencia insuficiente.
    """
    careers = filter_valid_engineering_careers(docente_ref_careers or [])
    if not careers:
        return {
            "accepted": False,
            "careers": [],
            "reason": "NO_DOCENTE_REF_TARGET_CAREER",
            "score_map": {career: 0 for career in VALID_ENGINEERING_CAREERS},
        }

    score_map = score_career_inference_from_text_fields(text_fields)
    strong_careers = [
        career for career in careers
        if career_has_strong_thematic_evidence_for_docente_ref(career, text_fields)
    ]

    if not strong_careers:
        return {
            "accepted": False,
            "careers": [],
            "reason": "DOCENTES_REF_WITHOUT_STRONG_THEMATIC_EVIDENCE",
            "score_map": score_map,
        }

    if len(strong_careers) == 1:
        return {
            "accepted": True,
            "careers": strong_careers,
            "reason": "DOCENTES_REF_STRONG_THEMATIC_SINGLE_CAREER",
            "score_map": score_map,
        }

    ranked = sorted(strong_careers, key=lambda c: (-score_map.get(c, 0), strong_careers.index(c)))
    best = ranked[0]
    best_score = score_map.get(best, 0)
    second_score = score_map.get(ranked[1], 0) if len(ranked) > 1 else 0

    # Si hay varias carreras docentes, solo aceptar la dominante; si no, REVIEW.
    if (best_score - second_score) >= max(6, CAREER_INFERENCE_MIN_MARGIN):
        return {
            "accepted": True,
            "careers": [best],
            "reason": "DOCENTES_REF_STRONG_THEMATIC_DOMINANT_CAREER",
            "score_map": score_map,
        }

    return {
        "accepted": False,
        "careers": strong_careers,
        "reason": "DOCENTES_REF_STRONG_THEMATIC_AMBIGUOUS_MULTI_CAREER",
        "score_map": score_map,
    }


def normalize_docentes_ref_careers_from_enrichment(enrichment: dict) -> list[str]:
    raw = enrichment.get("docentes_ref_careers")
    if isinstance(raw, list):
        return filter_valid_engineering_careers(raw)
    if isinstance(raw, str):
        return filter_valid_engineering_careers(split_semicolon_values(raw))
    return []

def determine_row_engineering_eligibility(
    document_type_value: str | None,
    authors_with_affiliations_value: str | None,
    affiliations_value: str | None,
    enrichment: dict,
    title_value: str | None = None,
    abstract_value: str | None = None,
    author_keywords_value: str | None = None,
    index_keywords_value: str | None = None,
    source_title_value: str | None = None,
    authors_value: str | None = None,
    use_career_ai: bool = False,
) -> dict:
    """
    Determina si una fila Scopus debe pasar a curated.

    Corrección principal:
    - No rechazar automáticamente afiliaciones ULima + Ingeniería genérica.
    - Si la afiliación es ULima + Ingeniería pero no trae carrera explícita, inferir
      Industrial/Civil/Sistemas usando título, abstract, keywords y source title.
    - Mantener segmentación por afiliación para evitar mezclar Ingeniería de otra
      universidad con Universidad de Lima.
    """
    normalized_doc_type = normalize_document_type_for_filter(document_type_value)
    if normalized_doc_type not in VALID_SCOPUS_DOCUMENT_TYPES:
        return {
            "eligible": False,
            "carrera_raw": None,
            "reason": f"REJECT_INVALID_DOCUMENT_TYPE: {document_type_value or 'UNKNOWN'}",
        }

    careers_from_enrichment = filter_valid_engineering_careers(
        split_semicolon_values(enrichment.get("carrera_raw"))
    )
    if careers_from_enrichment and enrichment.get("has_ulima_engineering_affiliation_raw"):
        return {
            "eligible": True,
            "carrera_raw": "; ".join(careers_from_enrichment),
            "reason": None,
            "metodo_cruce_scopus_raw": enrichment.get("metodo_cruce_scopus_raw") or "AFFILIATION_EXPLICIT_CAREER",
        }

    affiliation_details = resolve_engineering_affiliation_details(
        authors_with_affiliations_value,
        affiliations_value,
    )
    careers_from_affiliation = filter_valid_engineering_careers(affiliation_details.get("careers"))
    has_ulima_affiliation = bool(affiliation_details.get("has_ulima_affiliation"))
    has_engineering_context = bool(affiliation_details.get("has_ulima_engineering_context"))

    if careers_from_affiliation and has_engineering_context:
        return {
            "eligible": True,
            "carrera_raw": "; ".join(careers_from_affiliation),
            "reason": None,
            "metodo_cruce_scopus_raw": "AFFILIATION_EXPLICIT_CAREER",
        }

    if has_engineering_context and not careers_from_affiliation:
        inferred = infer_target_career_from_generic_engineering_context(
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
        )
        score_map = inferred.get("score_map") or {}

        # IA especializada para los casos realmente ambiguos:
        # ULima + Facultad/Escuela de Ingeniería genérica, sin Industrial/Civil/Sistemas explícito.
        # Si la IA está activa, su decisión conservadora prevalece sobre el rescate heurístico.
        if use_career_ai:
            ai_result = classify_generic_engineering_target_career_with_ai(
                title_value=title_value,
                abstract_value=abstract_value,
                author_keywords_value=author_keywords_value,
                index_keywords_value=index_keywords_value,
                source_title_value=source_title_value,
                authors_value=authors_value,
                ulima_docentes_value=enrichment.get("ulima_docentes_raw"),
                ulima_contexts=affiliation_details.get("ulima_contexts"),
                heuristic_score_map=score_map,
            )
            ai_careers = filter_valid_engineering_careers(ai_result.get("careers"))

            if ai_careers:
                method_suffix_map = {
                    "Ingeniería Industrial": "INDUSTRIAL",
                    "Ingeniería Civil": "CIVIL",
                    "Ingeniería de Sistemas": "SISTEMAS",
                }
                inferred_method_suffix = "_".join(method_suffix_map.get(career, "UNKNOWN") for career in ai_careers)
                return {
                    "eligible": True,
                    "carrera_raw": "; ".join(ai_careers),
                    "reason": None,
                    "metodo_cruce_scopus_raw": f"AFFIL_GENERIC_ENG_AI_{inferred_method_suffix}",
                }

            prefix = "REVIEW_ULIMA_ENGINEERING_AI_NO_TARGET_CAREER"
            if ai_result.get("decision") == "REJECT_NOT_TARGET" or ai_result.get("reason") == "AI_REJECT_NOT_TARGET":
                prefix = "REJECT_ULIMA_ENGINEERING_AI_NOT_TARGET_CAREER"

            return {
                "eligible": False,
                "carrera_raw": None,
                "reason": format_ai_review_reason(prefix, ai_result, score_map),
            }

        inferred_careers = filter_valid_engineering_careers(inferred.get("careers"))

        if inferred_careers:
            text_fields_for_safety = build_thematic_text_fields(
                title_value,
                abstract_value,
                author_keywords_value,
                index_keywords_value,
                source_title_value,
            )
            unsafe_reason = generic_engineering_heuristic_review_reason(
                inferred_careers=inferred_careers,
                text_fields=text_fields_for_safety,
                score_map=score_map,
            )
            if unsafe_reason:
                return {
                    "eligible": False,
                    "carrera_raw": None,
                    "reason": (
                        "REVIEW_ULIMA_ENGINEERING_GENERIC_HEURISTIC_UNSAFE: "
                        "ULima engineering context detected, but heuristic career inference is not safe enough. "
                        f"inference_reason={unsafe_reason}"
                    ),
                }

            method_suffix_map = {
                "Ingeniería Industrial": "INDUSTRIAL",
                "Ingeniería Civil": "CIVIL",
                "Ingeniería de Sistemas": "SISTEMAS",
            }
            inferred_method_suffix = "_".join(
                method_suffix_map.get(career, normalize_generic_text(career).replace("ingenieria ", "").replace(" ", "_").upper())
                for career in inferred_careers
            )
            return {
                "eligible": True,
                "carrera_raw": "; ".join(inferred_careers),
                "reason": None,
                "metodo_cruce_scopus_raw": f"AFFILIATION_GENERIC_ENGINEERING_INFERRED_{inferred_method_suffix}",
            }

        reason_detail = inferred.get("reason") or "NO_INFERENCE"
        score_detail = serialize_score_map_for_reason(score_map)
        return {
            "eligible": False,
            "carrera_raw": None,
            "reason": (
                "REVIEW_ULIMA_ENGINEERING_GENERIC_NO_TARGET_CAREER: "
                "ULima engineering context detected, but no explicit target career and thematic inference was not strong enough. "
                f"inference_reason={reason_detail}; scores={score_detail}"
            ),
        }

    # Rescate conservador para falsos negativos:
    # ULima real + docente en ref.docentes_ulima + evidencia temática fuerte,
    # aunque la afiliación local no declare Facultad/Carrera de Ingeniería.
    # Esto corrige casos como Lean/Industrial, ML/Sistemas y sísmica/Civil
    # cuando Scopus solo reporta "Universidad de Lima" o "Instituto de Investigación Científica".
    docente_ref_careers = normalize_docentes_ref_careers_from_enrichment(enrichment)
    if (has_ulima_affiliation or enrichment.get("es_ulima_raw_detected")) and docente_ref_careers:
        text_fields_for_docente_rescue = build_thematic_text_fields(
            title_value,
            abstract_value,
            author_keywords_value,
            index_keywords_value,
            source_title_value,
        )
        rescue = resolve_docentes_ref_strong_thematic_rescue(
            docente_ref_careers=docente_ref_careers,
            text_fields=text_fields_for_docente_rescue,
        )
        rescue_score_map = rescue.get("score_map") or {}

        if rescue.get("accepted") and rescue.get("careers"):
            method_suffix_map = {
                "Ingeniería Industrial": "INDUSTRIAL",
                "Ingeniería Civil": "CIVIL",
                "Ingeniería de Sistemas": "SISTEMAS",
            }
            rescued_careers = filter_valid_engineering_careers(rescue.get("careers"))
            inferred_method_suffix = "_".join(method_suffix_map.get(career, "UNKNOWN") for career in rescued_careers)
            return {
                "eligible": True,
                "carrera_raw": "; ".join(rescued_careers),
                "reason": None,
                "metodo_cruce_scopus_raw": f"DOCENTES_REF_STRONG_THEMATIC_EVIDENCE_{inferred_method_suffix}",
            }

        return {
            "eligible": False,
            "carrera_raw": None,
            "reason": (
                "REVIEW_ULIMA_DOCENTES_REF_NO_ENGINEERING_CONTEXT: "
                "ULima affiliation and docente_ref detected, but explicit engineering affiliation is missing "
                "and thematic evidence was not safe enough for automatic acceptance. "
                f"rescue_reason={rescue.get('reason')}; "
                f"docente_ref_careers={'; '.join(docente_ref_careers)}; "
                f"scores={serialize_score_map_for_reason(rescue_score_map)}"
            ),
        }

    if not has_ulima_affiliation and not enrichment.get("es_ulima_raw_detected"):
        return {
            "eligible": False,
            "carrera_raw": None,
            "reason": "REJECT_NO_ULIMA_AFFILIATION: no Universidad de Lima / University of Lima affiliation detected.",
        }

    if has_ulima_affiliation and not has_engineering_context:
        return {
            "eligible": False,
            "carrera_raw": None,
            "reason": "REJECT_REAL_ULIMA_BUT_NO_ENGINEERING_CONTEXT: ULima affiliation detected, but no Faculty/Career/School/Department of Engineering context detected in the same ULima segment.",
        }

    if careers_from_affiliation and not has_engineering_context:
        return {
            "eligible": False,
            "carrera_raw": "; ".join(careers_from_affiliation),
            "reason": "REJECT_TARGET_CAREER_WITHOUT_ULIMA_ENGINEERING_CONTEXT: target career text detected, but not inside a valid ULima engineering affiliation context.",
        }

    return {
        "eligible": False,
        "carrera_raw": None,
        "reason": "REJECT_NO_VALID_ULIMA_ENGINEERING_AFFILIATION: no valid ULima engineering affiliation for Industrial/Civil/Systems.",
    }


# =========================
# CATALOGS
# =========================
CAREER_AREA_LINE_CATALOG = {
    "Ingeniería Industrial": {
        "Work Design & Human Factors": [
            "Diseño de sistemas de trabajo",
            "Evaluación de factores físicos",
            "Evaluación ergonómica",
        ],
        "Operations Research & Analysis": [
            "Modelamiento matemático a la mejora de procesos como soporte a la toma de decisiones",
            "Simulación para la mejora del diseño de procesos",
            "Diseño y desarrollo de modelos para el análisis y predicción de las variables de un proceso",
        ],
        "Operations Engineering & Management": [
            "Planeamiento y Gestión de Operaciones",
            "Planeamiento, programación y control de proyectos",
            "Gestión de mantenimiento",
        ],
        "Supply Chain Management": [
            "Gestión de la cadena de suministro",
            "Gestión de Logística Inversa",
            "Gestión de Inventarios, Almacenes y Transportes",
            "Gestión de compras y proveedores, Nivel de servicio y Satisfacción al cliente",
        ],
        "Safety": [
            "Gestión de riesgos ocupacionales",
            "Identificación, análisis, evaluación y control de riesgos en seguridad y salud ocupacional",
        ],
        "Product Design & Development": [
            "Diseño de producto",
            "Desarrollo de producto",
        ],
    },
    "Ingeniería Civil": {
        "Ciencias de la Ingeniería": [
            "Ecuaciones diferenciales aplicadas al análisis estructural",
            "Física de materiales",
        ],
        "Construcción": [
            "Metodología BIM",
            "Normativa BIM",
            "Materiales de Construcción",
            "Sostenibilidad",
            "Innovación en Proceso constructivos",
            "Tecnología",
        ],
        "Estructuras": [
            "Técnicas de experimentación en estructuras",
            "Vulnerabilidad sísmica de estructuras",
            "Sistemas de protección sísmica de estructuras",
        ],
        "Innovación Empresarial": [
            "Obras de Construcción y su relación con el medio ambiente",
        ],
        "Hidráulica": [
            "Hidrología e Hidráulica",
            "Riego y Drenaje",
            "Calidad del Agua",
            "Cambio Climático en Recursos Hídricos",
            "Transporte de Sedimentos",
            "Hidrogeología",
        ],
        "Transporte y Geotecnia": [
            "Geotecnia computacional",
            "Geotecnia ambiental e hidrogeología",
            "Geotecnia experimental",
            "Geotecnia minera",
            "Mecánica de rocas e ingeniería geológica",
            "Tecnologia de mezclas asfalticas",
            "Diseño estructural de Pavimentos",
            "Mejora y estabilización del suelo",
            "Cambio climatico en la infraestructura vial",
            "Gestion de riesgos geológicos",
        ],
        "Gestión de Proyectos": [
            "Gestión de riesgos",
            "Gestión estratégica de contratos",
            "Gestión de las comunicaciones",
            "Gestión de recursos",
        ],
    },
    "Ingeniería de Sistemas": {
        "Aplicaciones en inteligencia artificial": [
            "Visión computacional",
            "NLP",
            "Aprendizaje automático",
            "Minería de datos",
        ],
        "Sistemas de Tecnologías de Información (TI)": [
            "Seguridad de sistemas y aplicaciones",
            "IoT",
            "Computación de alto rendimiento",
            "Redes y ciberseguridad",
            "Sostenibilidad en TI",
        ],
        "Interacción humano-computadora": [
            "HCI",
            "Realidad virtual y aumentada",
            "Construcción de juegos y gamificación",
            "Agentes virtuales",
        ],
        "Algoritmos y sistemas computacionales": [
            "Optimización computacional",
            "Ingeniería de software",
            "Diseño de algoritmos",
        ],
        "Tecnologías y gestión de la información": [
            "Gestión de procesos tecnológicos",
            "Liderazgo, género y tecnología",
            "Sistemas de gestión del conocimiento",
            "Minería de procesos",
            "Computación aplicada",
            "Simulación de procesos",
        ],
    },
}

IDIC_CATEGORY_AREA_LINE_CATALOG = {
    "Innovación y tecnología digital": {
        "Inteligencia artificial y computación avanzada": [
            "Machine learning y deep learning",
            "Procesamiento de lenguaje natural",
            "Visión computacional",
            "Sistemas autónomos y robótica",
        ],
        "Transformación digital": [
            "Tecnologías emergentes",
            "Ciberseguridad y privacidad",
            "Internet de las cosas (IoT)",
            "Computación cuántica",
            "Diseño y construcción virtual",
        ],
        "Experiencia digital humana": [
            "Interacción humano-computadora",
            "Realidad virtual y aumentada",
            "Diseño de interfaces adaptativas",
        ],
    },
    "Desarrollo sostenible y medioambiente": {
        "Sostenibilidad y cambio climático": [
            "Energías renovables",
            "Economía circular",
            "Gestión sostenible de recursos",
            "Adaptación al cambio climático",
        ],
        "Ciudades inteligentes y sostenibles": [
            "Urbanismo sostenible",
            "Movilidad urbana",
            "Infraestructura sostenible",
            "Gestión inteligente de recursos",
        ],
        "Tecnología y ecosistemas": [
            "Tecnologías limpias",
            "Biodiversidad y conservación",
            "Gestión de residuos",
            "Materiales avanzados",
        ],
    },
    "Sociedad y comportamiento humano": {
        "Bienestar y desarrollo humano": [
            "Salud mental y bienestar",
            "Educación, desarrollo cognitivo y socioafectivo",
            "Comportamiento social",
            "Mujer, cultura y sociedad",
            "Pobreza e informalidad",
        ],
        "Comunicación y cultura digital": [
            "Medios digitales y sociedad",
            "Comunicación intercultural",
            "Narrativas transmedia",
            "Comportamiento digital",
        ],
        "Ética, gobernanza y responsabilidad social": [
            "Ética y gobernanza",
            "Responsabilidad social",
            "Derechos humanos y tecnología",
        ],
    },
    "Gestión y economía del conocimiento": {
        "Innovación empresarial": [
            "Modelos de negocio digitales",
            "Emprendimiento tecnológico",
            "Gestión de la innovación",
            "Transformación organizacional",
        ],
        "Economía digital": [
            "Fintech y servicios financieros",
            "Mercados globales",
            "Análisis de datos económicos",
            "Economía de plataformas",
        ],
        "Gestión del conocimiento": [
            "Gestión del capital intelectual",
            "Aprendizaje organizacional",
            "Transferencia de conocimiento",
            "Inteligencia de negocios",
        ],
    },
}


def build_career_line_to_area_map() -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for carrera, areas in CAREER_AREA_LINE_CATALOG.items():
        result[carrera] = {}
        for area, lineas in areas.items():
            for linea in lineas:
                result[carrera][normalize_generic_text(linea)] = area
    return result


def build_idic_maps() -> tuple[dict[str, str], dict[str, str]]:
    linea_to_area: dict[str, str] = {}
    area_to_category: dict[str, str] = {}

    for category, areas in IDIC_CATEGORY_AREA_LINE_CATALOG.items():
        for area, lineas in areas.items():
            area_to_category[normalize_generic_text(area)] = category
            for linea in lineas:
                linea_to_area[normalize_generic_text(linea)] = area

    return linea_to_area, area_to_category


CAREER_LINE_TO_AREA_MAP = build_career_line_to_area_map()
IDIC_LINE_TO_AREA_MAP, IDIC_AREA_TO_CATEGORY_MAP = build_idic_maps()


def get_allowed_career_areas(carrera: str | None) -> list[str]:
    if not carrera:
        return []
    return list(CAREER_AREA_LINE_CATALOG.get(carrera, {}).keys())


def get_allowed_career_lines(carrera: str | None, area_carrera: str | None = None) -> list[str]:
    if not carrera:
        return []

    areas = CAREER_AREA_LINE_CATALOG.get(carrera, {})
    if not areas:
        return []

    if area_carrera:
        return list(areas.get(area_carrera, []))

    result: list[str] = []
    for lineas in areas.values():
        result.extend(lineas)
    return result


def get_allowed_idic_categories() -> list[str]:
    return list(IDIC_CATEGORY_AREA_LINE_CATALOG.keys())


def get_allowed_idic_areas(category_tematica: str | None = None) -> list[str]:
    if category_tematica:
        return list(IDIC_CATEGORY_AREA_LINE_CATALOG.get(category_tematica, {}).keys())

    result: list[str] = []
    for areas in IDIC_CATEGORY_AREA_LINE_CATALOG.values():
        result.extend(list(areas.keys()))
    return result


def get_allowed_idic_lines(
    category_tematica: str | None = None,
    area_idic: str | None = None,
) -> list[str]:
    if category_tematica and area_idic:
        return list(IDIC_CATEGORY_AREA_LINE_CATALOG.get(category_tematica, {}).get(area_idic, []))

    if category_tematica:
        result: list[str] = []
        for lineas in IDIC_CATEGORY_AREA_LINE_CATALOG.get(category_tematica, {}).values():
            result.extend(lineas)
        return result

    result: list[str] = []
    for areas in IDIC_CATEGORY_AREA_LINE_CATALOG.values():
        for lineas in areas.values():
            result.extend(lineas)
    return result


def coerce_area_carrera_from_linea(carrera: str | None, linea_carrera: str | None) -> str | None:
    if not carrera or not linea_carrera:
        return None
    return CAREER_LINE_TO_AREA_MAP.get(carrera, {}).get(normalize_generic_text(linea_carrera))


def coerce_category_tematica_from_area(area_idic: str | None) -> str | None:
    if not area_idic:
        return None
    return IDIC_AREA_TO_CATEGORY_MAP.get(normalize_generic_text(area_idic))


def coerce_area_idic_from_linea(linea_idic: str | None) -> str | None:
    if not linea_idic:
        return None
    return IDIC_LINE_TO_AREA_MAP.get(normalize_generic_text(linea_idic))


def is_valid_career_area_line(
    carrera: str | None,
    area_carrera: str | None,
    linea_carrera: str | None,
) -> bool:
    if not carrera or not area_carrera or not linea_carrera:
        return False
    return linea_carrera in CAREER_AREA_LINE_CATALOG.get(carrera, {}).get(area_carrera, [])


def is_valid_idic_triplet(
    category_tematica: str | None,
    area_idic: str | None,
    linea_idic: str | None,
) -> bool:
    if not category_tematica or not area_idic or not linea_idic:
        return False
    return linea_idic in IDIC_CATEGORY_AREA_LINE_CATALOG.get(category_tematica, {}).get(area_idic, [])


# =========================
# CLASSIFICATION HINTS (fallback)
# =========================
CAREER_LINE_HINTS = {
    "Ingeniería Industrial": {
        "Diseño de sistemas de trabajo": [
            "diseno de sistemas de trabajo",
            "work design",
            "work system design",
        ],
        "Evaluación de factores físicos": [
            "factores fisicos",
            "physical factors",
        ],
        "Evaluación ergonómica": [
            "ergonomia",
            "ergonomic",
            "ergonomics",
        ],
        "Modelamiento matemático a la mejora de procesos como soporte a la toma de decisiones": [
            "modelamiento matematico",
            "mathematical modeling",
            "decision making",
        ],
        "Simulación para la mejora del diseño de procesos": [
            "simulacion",
            "simulation",
            "process simulation",
        ],
        "Diseño y desarrollo de modelos para el análisis y predicción de las variables de un proceso": [
            "prediccion",
            "prediction",
            "forecasting",
            "analisis de variables",
            "benchmarking",
            "multi model",
            "multi-model",
            "cross validation",
            "cross-validation",
            "particle swarm optimization",
            "pso",
            "machine learning",
        ],
        "Planeamiento y Gestión de Operaciones": [
            "operations management",
            "gestion de operaciones",
            "planeamiento de operaciones",
            "operational efficiency",
            "process improvement",
            "lean",
            "5s",
            "standard work",
            "otif",
            "productivity improvement",
        ],
        "Planeamiento, programación y control de proyectos": [
            "project scheduling",
            "project control",
            "programacion y control de proyectos",
        ],
        "Gestión de mantenimiento": [
            "mantenimiento",
            "maintenance management",
        ],
        "Gestión de la cadena de suministro": [
            "supply chain",
            "cadena de suministro",
        ],
        "Gestión de Logística Inversa": [
            "logistica inversa",
            "reverse logistics",
        ],
        "Gestión de Inventarios, Almacenes y Transportes": [
            "inventarios",
            "almacenes",
            "transportes",
            "inventory",
            "warehouse",
            "transportation",
            "inventory management",
            "warehouse management",
            "storage center",
            "packing process",
            "otif",
        ],
        "Gestión de compras y proveedores, Nivel de servicio y Satisfacción al cliente": [
            "compras",
            "proveedores",
            "nivel de servicio",
            "satisfaccion al cliente",
            "supplier",
            "service level",
        ],
        "Gestión de riesgos ocupacionales": [
            "riesgos ocupacionales",
            "occupational risks",
        ],
        "Identificación, análisis, evaluación y control de riesgos en seguridad y salud ocupacional": [
            "seguridad y salud ocupacional",
            "occupational health",
            "occupational safety",
        ],
        "Diseño de producto": [
            "diseno de producto",
            "product design",
        ],
        "Desarrollo de producto": [
            "desarrollo de producto",
            "product development",
            "nanoparticles",
            "silver nanoparticles",
            "green synthesis",
            "antimicrobial applications",
            "stone mastic asphalt",
            "sma mixtures",
            "asphalt mixtures",
            "sustainable fibres",
            "waste derived fibres",
            "material formulation",
            "composite materials",
        ],
    },
    "Ingeniería Civil": {
        "Ecuaciones diferenciales aplicadas al análisis estructural": [
            "analisis estructural",
            "structural analysis",
            "ecuaciones diferenciales",
        ],
        "Física de materiales": [
            "fisica de materiales",
            "materials physics",
        ],
        "Metodología BIM": [
            "bim",
            "building information modeling",
        ],
        "Normativa BIM": [
            "normativa bim",
            "bim standard",
            "bim standards",
        ],
        "Materiales de Construcción": [
            "materiales de construccion",
            "construction materials",
        ],
        "Sostenibilidad": [
            "sostenibilidad",
            "sustainability",
        ],
        "Innovación en Proceso constructivos": [
            "proceso constructivo",
            "construction process innovation",
        ],
        "Tecnología": [
            "tecnologia en construccion",
            "construction technology",
        ],
        "Técnicas de experimentación en estructuras": [
            "experimental structures",
            "experimentacion en estructuras",
        ],
        "Vulnerabilidad sísmica de estructuras": [
            "vulnerabilidad sismica",
            "seismic vulnerability",
        ],
        "Sistemas de protección sísmica de estructuras": [
            "proteccion sismica",
            "seismic protection",
            "base isolation",
        ],
        "Obras de Construcción y su relación con el medio ambiente": [
            "medio ambiente",
            "environmental impact",
            "construction and environment",
        ],
        "Hidrología e Hidráulica": [
            "hidrologia",
            "hidraulica",
            "hydrology",
            "hydraulics",
        ],
        "Riego y Drenaje": [
            "riego",
            "drenaje",
            "irrigation",
            "drainage",
        ],
        "Calidad del Agua": [
            "calidad del agua",
            "water quality",
        ],
        "Cambio Climático en Recursos Hídricos": [
            "cambio climatico",
            "climate change",
            "recursos hidricos",
            "water resources",
        ],
        "Transporte de Sedimentos": [
            "sedimentos",
            "sediment transport",
        ],
        "Hidrogeología": [
            "hidrogeologia",
            "hydrogeology",
        ],
        "Geotecnia computacional": [
            "geotecnia computacional",
            "computational geotechnics",
        ],
        "Geotecnia ambiental e hidrogeología": [
            "geotecnia ambiental",
            "environmental geotechnics",
        ],
        "Geotecnia experimental": [
            "geotecnia experimental",
            "experimental geotechnics",
        ],
        "Geotecnia minera": [
            "geotecnia minera",
            "mining geotechnics",
        ],
        "Mecánica de rocas e ingeniería geológica": [
            "mecanica de rocas",
            "rock mechanics",
            "ingenieria geologica",
        ],
        "Tecnologia de mezclas asfalticas": [
            "mezclas asfalticas",
            "asphalt mixtures",
        ],
        "Diseño estructural de Pavimentos": [
            "pavimentos",
            "pavement design",
        ],
        "Mejora y estabilización del suelo": [
            "estabilizacion del suelo",
            "soil stabilization",
            "mejora del suelo",
        ],
        "Cambio climatico en la infraestructura vial": [
            "infraestructura vial",
            "road infrastructure",
            "climate change in transport infrastructure",
        ],
        "Gestion de riesgos geológicos": [
            "riesgos geologicos",
            "geological risks",
        ],
        "Gestión de riesgos": [
            "gestion de riesgos",
            "risk management",
        ],
        "Gestión estratégica de contratos": [
            "contratos",
            "contract management",
        ],
        "Gestión de las comunicaciones": [
            "communications management",
            "gestion de las comunicaciones",
        ],
        "Gestión de recursos": [
            "resource management",
            "gestion de recursos",
        ],
    },
    "Ingeniería de Sistemas": {
        "Visión computacional": [
            "vision computacional",
            "computer vision",
        ],
        "NLP": [
            "nlp",
            "natural language processing",
            "procesamiento de lenguaje natural",
        ],
        "Aprendizaje automático": [
            "aprendizaje automatico",
            "machine learning",
        ],
        "Minería de datos": [
            "mineria de datos",
            "data mining",
        ],
        "Seguridad de sistemas y aplicaciones": [
            "seguridad de sistemas",
            "application security",
            "systems security",
        ],
        "IoT": [
            "iot",
            "internet of things",
        ],
        "Computación de alto rendimiento": [
            "high performance computing",
            "computacion de alto rendimiento",
            "hpc",
        ],
        "Redes y ciberseguridad": [
            "ciberseguridad",
            "cybersecurity",
            "networks",
            "redes",
        ],
        "Sostenibilidad en TI": [
            "sostenibilidad en ti",
            "green it",
        ],
        "HCI": [
            "hci",
            "human computer interaction",
            "interaccion humano computadora",
        ],
        "Realidad virtual y aumentada": [
            "realidad virtual",
            "realidad aumentada",
            "virtual reality",
            "augmented reality",
        ],
        "Construcción de juegos y gamificación": [
            "gamification",
            "gamificacion",
            "game development",
            "juegos",
        ],
        "Agentes virtuales": [
            "virtual agents",
            "agentes virtuales",
            "chatbot",
            "chatbots",
        ],
        "Optimización computacional": [
            "optimizacion computacional",
            "computational optimization",
        ],
        "Ingeniería de software": [
            "ingenieria de software",
            "software engineering",
        ],
        "Diseño de algoritmos": [
            "diseno de algoritmos",
            "algorithm design",
        ],
        "Gestión de procesos tecnológicos": [
            "procesos tecnologicos",
            "technology process management",
        ],
        "Liderazgo, género y tecnología": [
            "genero y tecnologia",
            "gender and technology",
            "liderazgo y tecnologia",
            "women in stem",
            "stem leadership",
            "gender gap",
        ],
        "Sistemas de gestión del conocimiento": [
            "gestion del conocimiento",
            "knowledge management systems",
        ],
        "Minería de procesos": [
            "mineria de procesos",
            "process mining",
        ],
        "Computación aplicada": [
            "computacion aplicada",
            "applied computing",
        ],
        "Simulación de procesos": [
            "simulacion de procesos",
            "process simulation",
        ],
    },
}

IDIC_LINE_HINTS = {
    "Machine learning y deep learning": [
        "machine learning",
        "deep learning",
        "aprendizaje automatico",
    ],
    "Procesamiento de lenguaje natural": [
        "natural language processing",
        "procesamiento de lenguaje natural",
        "nlp",
    ],
    "Visión computacional": [
        "computer vision",
        "vision computacional",
    ],
    "Sistemas autónomos y robótica": [
        "robotica",
        "robotics",
        "autonomous systems",
    ],
    "Tecnologías emergentes": [
        "tecnologias emergentes",
        "emerging technologies",
    ],
    "Ciberseguridad y privacidad": [
        "ciberseguridad",
        "cybersecurity",
        "privacy",
        "privacidad",
    ],
    "Internet de las cosas (IoT)": [
        "internet of things",
        "iot",
    ],
    "Computación cuántica": [
        "computacion cuantica",
        "quantum computing",
    ],
    "Diseño y construcción virtual": [
        "construccion virtual",
        "virtual construction",
        "bim",
    ],
    "Interacción humano-computadora": [
        "human computer interaction",
        "interaccion humano computadora",
        "hci",
    ],
    "Realidad virtual y aumentada": [
        "virtual reality",
        "augmented reality",
        "realidad virtual",
        "realidad aumentada",
    ],
    "Diseño de interfaces adaptativas": [
        "adaptive interfaces",
        "interfaces adaptativas",
    ],
    "Energías renovables": [
        "energias renovables",
        "renewable energy",
    ],
    "Economía circular": [
        "economia circular",
        "circular economy",
        "valorization",
        "waste valorization",
        "recycling",
        "reuse",
    ],
    "Gestión sostenible de recursos": [
        "gestion sostenible de recursos",
        "sustainable resource management",
        "resource optimization",
        "inventory management",
        "warehouse management",
        "storage center",
    ],
    "Adaptación al cambio climático": [
        "adaptacion al cambio climatico",
        "climate adaptation",
    ],
    "Urbanismo sostenible": [
        "urbanismo sostenible",
        "sustainable urbanism",
    ],
    "Movilidad urbana": [
        "movilidad urbana",
        "urban mobility",
    ],
    "Infraestructura sostenible": [
        "infraestructura sostenible",
        "sustainable infrastructure",
    ],
    "Gestión inteligente de recursos": [
        "smart resource management",
        "gestion inteligente de recursos",
        "inventory management",
        "resource allocation",
        "storage center",
        "warehouse",
        "inventory",
    ],
    "Tecnologías limpias": [
        "tecnologias limpias",
        "clean technologies",
        "green synthesis",
        "eco friendly",
        "eco-friendly",
        "sustainable valorization",
        "clean technology",
    ],
    "Biodiversidad y conservación": [
        "biodiversidad",
        "conservacion",
        "biodiversity",
        "conservation",
    ],
    "Gestión de residuos": [
        "gestion de residuos",
        "waste management",
        "waste",
        "paper industry waste",
        "agro-industrial waste",
        "sludge",
        "residuos",
    ],
    "Materiales avanzados": [
        "materiales avanzados",
        "advanced materials",
        "nanoparticles",
        "silver nanoparticles",
    ],
    "Salud mental y bienestar": [
        "salud mental",
        "mental health",
        "wellbeing",
        "bienestar",
    ],
    "Educación, desarrollo cognitivo y socioafectivo": [
        "educacion",
        "education",
        "desarrollo cognitivo",
        "socioafectivo",
    ],
    "Comportamiento social": [
        "comportamiento social",
        "social behavior",
    ],
    "Mujer, cultura y sociedad": [
        "mujer",
        "woman",
        "women",
        "culture and society",
        "cultura y sociedad",
        "women in stem",
        "gender gap",
        "leadership in stem",
    ],
    "Pobreza e informalidad": [
        "pobreza",
        "poverty",
        "informalidad",
        "informality",
    ],
    "Medios digitales y sociedad": [
        "medios digitales",
        "digital media",
        "society",
    ],
    "Comunicación intercultural": [
        "comunicacion intercultural",
        "intercultural communication",
    ],
    "Narrativas transmedia": [
        "narrativas transmedia",
        "transmedia narratives",
    ],
    "Comportamiento digital": [
        "comportamiento digital",
        "digital behavior",
    ],
    "Ética y gobernanza": [
        "etica",
        "ethics",
        "governance",
        "gobernanza",
    ],
    "Responsabilidad social": [
        "responsabilidad social",
        "social responsibility",
    ],
    "Derechos humanos y tecnología": [
        "derechos humanos",
        "human rights",
        "technology",
    ],
    "Modelos de negocio digitales": [
        "modelos de negocio digitales",
        "digital business models",
    ],
    "Emprendimiento tecnológico": [
        "emprendimiento tecnologico",
        "technology entrepreneurship",
    ],
    "Gestión de la innovación": [
        "gestion de la innovacion",
        "innovation management",
        "improvement model",
        "continuous improvement",
        "process improvement",
        "5s",
        "standard work",
        "operational efficiency",
        "productivity improvement",
        "otif",
    ],
    "Transformación organizacional": [
        "transformacion organizacional",
        "organizational transformation",
        "organizational improvement",
        "process redesign",
        "workflow redesign",
    ],
    "Fintech y servicios financieros": [
        "fintech",
        "financial services",
        "servicios financieros",
    ],
    "Mercados globales": [
        "mercados globales",
        "global markets",
    ],
    "Análisis de datos económicos": [
        "analisis de datos economicos",
        "economic data analysis",
    ],
    "Economía de plataformas": [
        "economia de plataformas",
        "platform economy",
    ],
    "Gestión del capital intelectual": [
        "capital intelectual",
        "intellectual capital",
    ],
    "Aprendizaje organizacional": [
        "aprendizaje organizacional",
        "organizational learning",
    ],
    "Transferencia de conocimiento": [
        "transferencia de conocimiento",
        "knowledge transfer",
    ],
    "Inteligencia de negocios": [
        "inteligencia de negocios",
        "business intelligence",
    ],
}


def build_classification_score(corpus: str, aliases: list[str]) -> int:
    score = 0
    for alias in aliases:
        if phrase_in_text(corpus, alias):
            score += 1
    return score


INDUSTRIAL_ORA_STRONG_SIGNALS = [
    "optimization",
    "optimisation",
    "optimizing",
    "optimising",
    "mathematical model",
    "mathematical modeling",
    "mathematical modelling",
    "modelamiento matematico",
    "prediction",
    "predictive",
    "forecast",
    "forecasting",
    "simulation",
    "simulacion",
    "benchmarking",
    "cross validation",
    "cross-validation",
    "particle swarm optimization",
    "genetic algorithm",
    "decision support",
    "machine learning",
    "deep learning",
    "algorithm",
    "algoritmo",
]

INDUSTRIAL_SCM_STRONG_SIGNALS = [
    "inventory",
    "inventory management",
    "warehouse",
    "warehouse management",
    "storage center",
    "logistics",
    "transportation",
    "transport",
    "supply chain",
    "inventarios",
    "almacenes",
    "almacen",
    "logistica",
    "transportes",
    "cadena de suministro",
    "supplier",
    "proveedores",
    "otif",
]

INDUSTRIAL_OEM_STRONG_SIGNALS = [
    "5s",
    "standard work",
    "lean",
    "kaizen",
    "operations management",
    "gestion de operaciones",
    "planeamiento de operaciones",
    "process improvement",
    "continuous improvement",
    "operational efficiency",
    "productivity improvement",
    "maintenance",
    "mantenimiento",
]

INDUSTRIAL_PDD_STRONG_SIGNALS = [
    "product development",
    "desarrollo de producto",
    "product design",
    "diseno de producto",
    "prototype",
    "prototipo",
    "materials",
    "material",
    "advanced material",
    "composite",
    "green synthesis",
    "nanoparticles",
    "silver nanoparticles",
    "fibres",
    "fibers",
    "asphalt",
    "stone mastic asphalt",
    "valorization",
    "valorisation",
]


IDIC_INNOVATION_MANAGEMENT_STRONG_SIGNALS = [
    "5s",
    "standard work",
    "lean",
    "kaizen",
    "continuous improvement",
    "process improvement",
    "improvement model",
    "operational efficiency",
    "productivity improvement",
    "innovation management",
    "gestion de la innovacion",
    "workflow improvement",
    "mejora de procesos",
]

IDIC_ORG_TRANSFORMATION_STRONG_SIGNALS = [
    "organizational transformation",
    "organisational transformation",
    "transformacion organizacional",
    "organizational change",
    "organisational change",
    "change management",
    "business transformation",
    "process redesign",
    "workflow redesign",
    "organizational redesign",
    "restructuring",
    "digital transformation",
]


IDIC_MACHINE_LEARNING_STRONG_SIGNALS = [
    "machine learning",
    "deep learning",
    "aprendizaje automatico",
    "aprendizaje automático",
    "predictive model",
    "predictive models",
    "prediction model",
    "prediction models",
    "benchmarking",
    "multi model",
    "multi-model",
    "model benchmarking",
]

IDIC_CLIMATE_ADAPTATION_EXACT_SIGNALS = [
    "adaptacion al cambio climatico",
    "adaptación al cambio climático",
    "climate change adaptation",
    "adaptation to climate change",
    "climate adaptation",
]

IDIC_CLIMATE_TERMS = [
    "climate change",
    "cambio climatico",
    "cambio climático",
    "climatic change",
    "climate variability",
    "variabilidad climatica",
    "variabilidad climática",
    "global warming",
    "calentamiento global",
]

IDIC_ADAPTATION_CONTEXT_TERMS = [
    "adaptation",
    "adaptacion",
    "adaptación",
    "resilience",
    "resiliencia",
    "vulnerability",
    "vulnerabilidad",
    "climate risk",
    "riesgo climatico",
    "riesgo climático",
    "climate hazard",
    "amenaza climatica",
    "amenaza climática",
    "climate impact",
    "impacto climatico",
    "impacto climático",
    "exposure",
    "exposicion",
    "exposición",
]

IDIC_CIRCULAR_ECONOMY_STRONG_SIGNALS = [
    "circular economy",
    "economia circular",
    "economía circular",
    "valorization",
    "valorisation",
    "valorizacion",
    "valorización",
    "waste valorization",
    "recycling",
    "reciclaje",
    "reuse",
    "reutilizacion",
    "reutilización",
]

IDIC_CLEAN_TECH_STRONG_SIGNALS = [
    "clean technology",
    "clean technologies",
    "tecnologias limpias",
    "tecnologías limpias",
    "green synthesis",
    "sintesis verde",
    "síntesis verde",
    "eco friendly",
    "eco-friendly",
    "sustainable valorization",
    "sustainable valorisation",
    "valorizacion sostenible",
    "valorización sostenible",
]

IDIC_WASTE_MANAGEMENT_STRONG_SIGNALS = [
    "waste management",
    "gestion de residuos",
    "gestión de residuos",
    "waste",
    "residuos",
    "sludge",
    "lodo",
    "adsorption",
    "adsorcion",
    "adsorción",
    "removal",
    "remocion",
    "remoción",
]

IDIC_ADVANCED_MATERIALS_STRONG_SIGNALS = [
    "advanced materials",
    "materiales avanzados",
    "nanoparticles",
    "nanoparticulas",
    "nanopartículas",
    "silver nanoparticles",
    "composite",
    "composites",
    "material formulation",
    "fiber reinforced",
    "fibre reinforced",
    "fiber-reinforced",
    "fibre-reinforced",
    "concrete",
    "concreto",
    "beam",
    "beams",
    "vigas",
    "cyclic loading",
    "carga ciclica",
    "carga cíclica",
    "structural response",
    "respuesta estructural",
    "asphalt",
    "asfalto",
]

IDIC_SUSTAINABLE_INFRASTRUCTURE_STRONG_SIGNALS = [
    "sustainable infrastructure",
    "infraestructura sostenible",
    "urban infrastructure",
    "road infrastructure",
    "infraestructura vial",
    "pavement",
    "pavimento",
    "construction infrastructure",
]

IDIC_EMERGING_TECH_STRONG_SIGNALS = [
    "emerging technologies",
    "tecnologias emergentes",
    "tecnologías emergentes",
    "computational characterization",
    "caracterizacion computacional",
    "caracterización computacional",
    "density functional theory",
    "dft",
    "molecular docking",
    "molecular dynamics",
    "spectroscopic",
    "spectroscopy",
    "spectrometric",
    "mass spectrometry",
    "esi ms",
    "esi-ms",
    "hplc",
    "uhplc",
    "orbitrap",
    "chromatography",
    "chemical characterization",
    "bioactive compounds",
    "secondary metabolites",
    "antitubercular",
    "isoniazid",
    "hydrazones",
    "drug derivatives",
    "novel derivatives",
]

# =========================
# PATCH GUARDRAILS V2 — curated classification quality
# =========================
CIVIL_STRUCTURAL_STRONG_SIGNALS = [
    "structural", "structure", "structures", "beam", "beams", "column", "columns",
    "cyclic loading", "reversed cyclic", "reinforced concrete", "fiber reinforced concrete",
    "fibre reinforced concrete", "concrete", "masonry", "seismic", "earthquake",
    "stiffness", "ductility", "modal analysis", "hysteretic", "vulnerability",
    "self built housing", "self-built housing", "structural degradation",
]

CIVIL_CONSTRUCTION_MATERIALS_STRONG_SIGNALS = [
    "construction materials", "materiales de construccion", "self compacting concrete",
    "self-compacting concrete", "high performance concrete", "high-strength concrete",
    "heat of hydration", "hydration heat", "mass concrete", "massive concrete",
    "polypropylene fibers", "cement", "mortar", "aggregate", "construction material",
]

CIVIL_BIM_VDC_STRONG_SIGNALS = [
    "bim", "building information modelling", "building information modeling",
    "hbim", "historic building information modeling", "historic building information modelling",
    "vdc", "virtual design and construction", "digital twin", "digital twins",
    "point cloud", "uav", "gnss", "photogrammetry", "geomatic", "dynamo",
]

CIVIL_WATER_STRONG_SIGNALS = [
    "water quality", "calidad del agua", "hydrology", "hydraulic", "hydraulics",
    "water resources", "recursos hidricos", "drainage", "irrigation", "groundwater",
    "hydrogeology", "aquifer", "river", "sediment transport", "rainfall", "flood",
    "wastewater", "arsenic", "pollutant", "contaminant", "adsorption", "water treatment",
]

CIVIL_TRANSPORT_GEOTECH_STRONG_SIGNALS = [
    "asphalt", "pavement", "stone mastic asphalt", "sma", "road", "soil", "geotechnical",
    "geotechnics", "rock mechanics", "slope", "stabilization", "transport infrastructure",
]

SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS = [
    "computer vision", "vision computacional", "image enhancement", "image processing",
    "object detection", "human detection", "activity recognition", "keypoint detection",
    "homography", "monocular video", "surveillance camera", "surveillance cameras",
    "yolo", "gan", "generative adversarial", "low light", "low-light", "image",
    "images", "video", "pose", "detection",
]

SYSTEMS_ML_STRONG_SIGNALS = [
    "machine learning", "deep learning", "aprendizaje automatico", "classification",
    "classifier", "predictive model", "predictive models", "xgboost", "lightgbm",
    "catboost", "random forest", "neural network", "neural networks", "lstm",
    "cnn", "transformer", "data mining", "clustering",
]

SYSTEMS_CYBER_STRONG_SIGNALS = [
    "cybersecurity", "ciberseguridad", "ddos", "botnet", "malware", "phishing",
    "intrusion detection", "network security", "application security", "security of systems",
    "seguridad de sistemas", "seguridad de aplicaciones",
]

SYSTEMS_GAME_STRONG_SIGNALS = [
    "serious game", "serious games", "gamification", "gamificacion", "game-based",
    "game based", "game", "games", "video game", "videogame", "juegos",
]

SYSTEMS_ARVR_STRONG_SIGNALS = [
    "augmented reality", "virtual reality", "mixed reality", "realidad aumentada",
    "realidad virtual", "ar platform", "vr", "ar",
]

SYSTEMS_AGENT_STRONG_SIGNALS = [
    "virtual agent", "virtual agents", "agente virtual", "agentes virtuales",
    "chatbot", "chatbots", "conversational agent", "conversational agents",
    "assistant", "avatar", "human agent interaction",
]

SYSTEMS_SOFTWARE_STRONG_SIGNALS = [
    "software engineering", "ingenieria de software", "software architecture",
    "clean architecture", "mobile application", "android application", "web application",
    "application", "app", "platform", "database", "information system", "information systems",
]

SYSTEMS_GENDER_TECH_STRONG_SIGNALS = [
    "women in stem", "gender gap", "gender stereotypes", "stem", "it-related programs",
    "information technology-related programs", "girls", "leadership in stem",
]

INDUSTRIAL_EXTENDED_STRONG_SIGNALS = [
    "business management", "business process management", "competitiveness",
    "commercial model", "sales", "crm", "lead handling", "demand management",
    "operational management", "production plant", "plant capacity", "technological feasibility",
    "occupational risk", "occupational risks", "safety rules", "service model", "nps",
    "retail", "commercial sector", "forecasting", "profile standardization", "poka yoke",
    "queueing", "workload balancing", "yamazumi", "jit", "rop", "ddmrp", "heijunka",
    "six sigma", "sweep method", "value stream mapping", "vsm",
]

CHEM_MATERIAL_PRODUCT_STRONG_SIGNALS = [
    "synthesis", "green synthesis", "chemical synthesis", "characterization",
    "computational characterization", "dft", "density functional theory", "molecular docking",
    "molecular dynamics", "metabolites", "secondary metabolites", "hplc", "uhplc",
    "mass spectrometry", "orbitrap", "compound", "compounds", "derivatives",
    "antitubercular", "isoniazid", "hydrazone", "phenylpyrazole", "cellulose",
    "nanofibers", "nanofibres", "nanocellulose", "chitosan", "films", "biopolymer",
    "activated carbon", "adsorption", "photoelectrocatalysis", "electrode", "photoanode",
]



# Extensiones de scoring activadas para el run posterior al 68.
CAREER_INFERENCE_SIGNALS["Ingeniería Industrial"].extend(
    INDUSTRIAL_EXTENDED_STRONG_SIGNALS + CHEM_MATERIAL_PRODUCT_STRONG_SIGNALS
)
CAREER_INFERENCE_SIGNALS["Ingeniería Civil"].extend(
    CIVIL_STRUCTURAL_STRONG_SIGNALS
    + CIVIL_CONSTRUCTION_MATERIALS_STRONG_SIGNALS
    + CIVIL_BIM_VDC_STRONG_SIGNALS
    + CIVIL_WATER_STRONG_SIGNALS
    + CIVIL_TRANSPORT_GEOTECH_STRONG_SIGNALS
)
CAREER_INFERENCE_SIGNALS["Ingeniería de Sistemas"].extend(
    SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS
    + SYSTEMS_ML_STRONG_SIGNALS
    + SYSTEMS_CYBER_STRONG_SIGNALS
    + SYSTEMS_GAME_STRONG_SIGNALS
    + SYSTEMS_ARVR_STRONG_SIGNALS
    + SYSTEMS_AGENT_STRONG_SIGNALS
    + SYSTEMS_SOFTWARE_STRONG_SIGNALS
    + SYSTEMS_GENDER_TECH_STRONG_SIGNALS
)

CAREER_LINE_HINTS["Ingeniería Industrial"]["Planeamiento y Gestión de Operaciones"].extend(
    INDUSTRIAL_EXTENDED_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería Industrial"]["Desarrollo de producto"].extend(
    CHEM_MATERIAL_PRODUCT_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería Industrial"]["Diseño y desarrollo de modelos para el análisis y predicción de las variables de un proceso"].extend(
    ["forecasting", "demand forecasting", "predictive analytics", "regression", "classification"]
)

CAREER_LINE_HINTS["Ingeniería Civil"]["Técnicas de experimentación en estructuras"].extend(
    CIVIL_STRUCTURAL_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería Civil"]["Materiales de Construcción"].extend(
    CIVIL_CONSTRUCTION_MATERIALS_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería Civil"]["Metodología BIM"].extend(
    CIVIL_BIM_VDC_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería Civil"]["Calidad del Agua"].extend(
    ["water treatment", "arsenic", "pollutant", "contaminant", "adsorption", "wastewater"]
)

CAREER_LINE_HINTS["Ingeniería de Sistemas"]["Visión computacional"].extend(
    SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería de Sistemas"]["Aprendizaje automático"].extend(
    SYSTEMS_ML_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería de Sistemas"]["Redes y ciberseguridad"].extend(
    SYSTEMS_CYBER_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería de Sistemas"]["Construcción de juegos y gamificación"].extend(
    SYSTEMS_GAME_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería de Sistemas"]["Realidad virtual y aumentada"].extend(
    SYSTEMS_ARVR_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería de Sistemas"]["Agentes virtuales"].extend(
    SYSTEMS_AGENT_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería de Sistemas"]["Ingeniería de software"].extend(
    SYSTEMS_SOFTWARE_STRONG_SIGNALS
)
CAREER_LINE_HINTS["Ingeniería de Sistemas"]["Liderazgo, género y tecnología"].extend(
    SYSTEMS_GENDER_TECH_STRONG_SIGNALS
)

IDIC_LINE_HINTS["Visión computacional"].extend(SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS)
IDIC_LINE_HINTS["Machine learning y deep learning"].extend(SYSTEMS_ML_STRONG_SIGNALS)
IDIC_LINE_HINTS["Ciberseguridad y privacidad"].extend(SYSTEMS_CYBER_STRONG_SIGNALS)
IDIC_LINE_HINTS["Realidad virtual y aumentada"].extend(SYSTEMS_ARVR_STRONG_SIGNALS)
IDIC_LINE_HINTS["Gestión de la innovación"].extend(INDUSTRIAL_EXTENDED_STRONG_SIGNALS)
IDIC_LINE_HINTS["Materiales avanzados"].extend(CHEM_MATERIAL_PRODUCT_STRONG_SIGNALS)
IDIC_LINE_HINTS["Gestión de residuos"].extend(["arsenic", "water treatment", "pollutant", "contaminant", "photoelectrocatalysis", "adsorption"])



def contains_any_phrase_in_text_fields(text_fields: dict[str, str], phrases: list[str]) -> bool:
    for phrase in phrases:
        norm_phrase = normalize_generic_text(phrase)
        if not norm_phrase:
            continue
        for text_value in text_fields.values():
            if text_value and phrase_in_text(text_value, norm_phrase):
                return True
    return False


def count_phrase_hits_in_text(value: str | None, phrases: list[str]) -> int:
    norm_value = normalize_generic_text(value)
    if not norm_value:
        return 0

    hits = 0
    seen: set[str] = set()
    for phrase in phrases:
        norm_phrase = normalize_generic_text(phrase)
        if not norm_phrase or norm_phrase in seen:
            continue
        if phrase_in_text(norm_value, norm_phrase):
            hits += 1
            seen.add(norm_phrase)
    return hits


def is_line_eligible_by_domain_rules(carrera: str, linea: str, text_fields: dict[str, str]) -> bool:
    if carrera != "Ingeniería Industrial":
        return True

    if linea in (
        "Modelamiento matemático a la mejora de procesos como soporte a la toma de decisiones",
        "Simulación para la mejora del diseño de procesos",
        "Diseño y desarrollo de modelos para el análisis y predicción de las variables de un proceso",
    ):
        return contains_any_phrase_in_text_fields(text_fields, INDUSTRIAL_ORA_STRONG_SIGNALS)

    return True


def apply_industrial_line_bonus(carrera: str, linea: str, text_fields: dict[str, str], base_score: int) -> int:
    if carrera != "Ingeniería Industrial":
        return base_score

    title_text = text_fields.get("title")
    abstract_text = text_fields.get("abstract")
    keyword_text = " | ".join(
        [
            text_fields.get("author_keywords", ""),
            text_fields.get("index_keywords", ""),
        ]
    )

    if linea in (
        "Modelamiento matemático a la mejora de procesos como soporte a la toma de decisiones",
        "Simulación para la mejora del diseño de procesos",
        "Diseño y desarrollo de modelos para el análisis y predicción de las variables de un proceso",
    ):
        bonus = 0
        bonus += 3 * count_phrase_hits_in_text(title_text, INDUSTRIAL_ORA_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, INDUSTRIAL_ORA_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(keyword_text, INDUSTRIAL_ORA_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Gestión de Inventarios, Almacenes y Transportes":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, INDUSTRIAL_SCM_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, INDUSTRIAL_SCM_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, INDUSTRIAL_SCM_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Gestión de la cadena de suministro":
        bonus = 0
        bonus += 3 * count_phrase_hits_in_text(title_text, INDUSTRIAL_SCM_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, INDUSTRIAL_SCM_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, INDUSTRIAL_SCM_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Planeamiento y Gestión de Operaciones":
        bonus = 0
        bonus += 3 * count_phrase_hits_in_text(title_text, INDUSTRIAL_OEM_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, INDUSTRIAL_OEM_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, INDUSTRIAL_OEM_STRONG_SIGNALS)
        return base_score + bonus

    if linea in ("Diseño de producto", "Desarrollo de producto"):
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, INDUSTRIAL_PDD_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, INDUSTRIAL_PDD_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, INDUSTRIAL_PDD_STRONG_SIGNALS)
        return base_score + bonus

    return base_score


def choose_best_line_with_rules(
    carrera: str,
    score_map: dict[str, int],
    text_fields: dict[str, str],
    min_score: int,
    min_margin: int = 1,
    relaxed: bool = False,
) -> str | None:
    filtered: list[tuple[str, int]] = []

    for linea, score in score_map.items():
        if score < min_score:
            continue
        if not is_line_eligible_by_domain_rules(carrera, linea, text_fields):
            continue
        adjusted_score = apply_industrial_line_bonus(carrera, linea, text_fields, score)
        filtered.append((linea, adjusted_score))

    if not filtered:
        return None

    filtered.sort(key=lambda x: (-x[1], normalize_generic_text(x[0])))

    if relaxed:
        return filtered[0][0]

    best_linea, best_score = filtered[0]
    tied = [linea for linea, score in filtered if score == best_score]
    if len(tied) > 1:
        return None

    if len(filtered) > 1:
        second_score = filtered[1][1]
        if (best_score - second_score) < min_margin:
            return None

    return best_linea





def career_fallback_from_domain_rules(
    carrera: str,
    text_fields: dict[str, str],
) -> tuple[str | None, str | None, str | None]:
    """
    Fallback de carrera controlado por dominio.

    Evita que el fallback alfabético mande Civil a 'Calidad del Agua' o Sistemas
    a 'Agentes virtuales' cuando no existe evidencia real para esas líneas.
    """
    if carrera == "Ingeniería Civil":
        if contains_any_phrase_in_text_fields(text_fields, CIVIL_BIM_VDC_STRONG_SIGNALS):
            return "Construcción", "Metodología BIM", "career_guardrail_civil_bim"

        if contains_any_phrase_in_text_fields(text_fields, CIVIL_STRUCTURAL_STRONG_SIGNALS):
            return "Estructuras", "Técnicas de experimentación en estructuras", "career_guardrail_civil_structures"

        if contains_any_phrase_in_text_fields(text_fields, CIVIL_CONSTRUCTION_MATERIALS_STRONG_SIGNALS):
            return "Construcción", "Materiales de Construcción", "career_guardrail_civil_construction_materials"

        if contains_any_phrase_in_text_fields(text_fields, CIVIL_TRANSPORT_GEOTECH_STRONG_SIGNALS):
            if contains_any_phrase_in_text_fields(text_fields, ["pavement", "pavimento"]):
                return "Transporte y Geotecnia", "Diseño estructural de Pavimentos", "career_guardrail_civil_pavements"
            if contains_any_phrase_in_text_fields(text_fields, ["asphalt", "stone mastic asphalt", "sma"]):
                return "Transporte y Geotecnia", "Tecnologia de mezclas asfalticas", "career_guardrail_civil_asphalt"
            return "Transporte y Geotecnia", "Geotecnia experimental", "career_guardrail_civil_geotech"

        if contains_any_phrase_in_text_fields(text_fields, CIVIL_WATER_STRONG_SIGNALS):
            if contains_any_phrase_in_text_fields(text_fields, ["drainage", "irrigation", "riego", "drenaje"]):
                return "Hidráulica", "Riego y Drenaje", "career_guardrail_civil_drainage_irrigation"
            if contains_any_phrase_in_text_fields(text_fields, ["hydrogeology", "groundwater", "aquifer", "hidrogeologia"]):
                return "Hidráulica", "Hidrogeología", "career_guardrail_civil_hydrogeology"
            if contains_any_phrase_in_text_fields(text_fields, ["water quality", "calidad del agua", "water treatment", "arsenic", "pollutant", "contaminant", "adsorption"]):
                return "Hidráulica", "Calidad del Agua", "career_guardrail_civil_water_quality"
            return "Hidráulica", "Hidrología e Hidráulica", "career_guardrail_civil_hydraulics"

        if contains_any_phrase_in_text_fields(text_fields, ["urban regeneration", "informal settlement", "informal settlements", "sustainable urban", "housing"]):
            return "Construcción", "Sostenibilidad", "career_guardrail_civil_sustainability"

        # Fallback seguro de Civil: evita 'Calidad del Agua' sin evidencia hídrica.
        return "Construcción", "Sostenibilidad", "career_guardrail_civil_safe_default"

    if carrera == "Ingeniería de Sistemas":
        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS):
            return "Aplicaciones en inteligencia artificial", "Visión computacional", "career_guardrail_systems_computer_vision"

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_CYBER_STRONG_SIGNALS):
            return "Sistemas de Tecnologías de Información (TI)", "Redes y ciberseguridad", "career_guardrail_systems_cybersecurity"

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ARVR_STRONG_SIGNALS):
            return "Interacción humano-computadora", "Realidad virtual y aumentada", "career_guardrail_systems_arvr"

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_GAME_STRONG_SIGNALS):
            return "Interacción humano-computadora", "Construcción de juegos y gamificación", "career_guardrail_systems_games"

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_AGENT_STRONG_SIGNALS):
            return "Interacción humano-computadora", "Agentes virtuales", "career_guardrail_systems_virtual_agents"

        if contains_any_phrase_in_text_fields(text_fields, ["nlp", "natural language processing", "large language model", "llm"]):
            return "Aplicaciones en inteligencia artificial", "NLP", "career_guardrail_systems_nlp"

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ML_STRONG_SIGNALS):
            return "Aplicaciones en inteligencia artificial", "Aprendizaje automático", "career_guardrail_systems_machine_learning"

        if contains_any_phrase_in_text_fields(text_fields, ["iot", "internet of things"]):
            return "Sistemas de Tecnologías de Información (TI)", "IoT", "career_guardrail_systems_iot"

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_SOFTWARE_STRONG_SIGNALS):
            return "Algoritmos y sistemas computacionales", "Ingeniería de software", "career_guardrail_systems_software"

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_GENDER_TECH_STRONG_SIGNALS):
            return "Tecnologías y gestión de la información", "Liderazgo, género y tecnología", "career_guardrail_systems_gender_tech"

        # Fallback seguro de Sistemas: evita 'Agentes virtuales' sin evidencia.
        return "Tecnologías y gestión de la información", "Computación aplicada", "career_guardrail_systems_safe_default"

    if carrera == "Ingeniería Industrial":
        if contains_any_phrase_in_text_fields(text_fields, CHEM_MATERIAL_PRODUCT_STRONG_SIGNALS):
            return "Product Design & Development", "Desarrollo de producto", "career_guardrail_industrial_product_development_chem_materials"

        if contains_any_phrase_in_text_fields(text_fields, INDUSTRIAL_SCM_STRONG_SIGNALS):
            if contains_any_phrase_in_text_fields(text_fields, ["supply chain", "cadena de suministro"]):
                return "Supply Chain Management", "Gestión de la cadena de suministro", "career_guardrail_industrial_supply_chain"
            return "Supply Chain Management", "Gestión de Inventarios, Almacenes y Transportes", "career_guardrail_industrial_inventory_warehouse"

        if contains_any_phrase_in_text_fields(text_fields, INDUSTRIAL_OEM_STRONG_SIGNALS + INDUSTRIAL_EXTENDED_STRONG_SIGNALS):
            return "Operations Engineering & Management", "Planeamiento y Gestión de Operaciones", "career_guardrail_industrial_operations"

        if contains_any_phrase_in_text_fields(text_fields, INDUSTRIAL_ORA_STRONG_SIGNALS):
            return "Operations Research & Analysis", "Diseño y desarrollo de modelos para el análisis y predicción de las variables de un proceso", "career_guardrail_industrial_ora"

        return "Operations Engineering & Management", "Planeamiento y Gestión de Operaciones", "career_guardrail_industrial_safe_default"

    return None, None, None


def apply_final_career_guardrails(
    carrera: str | None,
    area_carrera: str | None,
    linea_carrera: str | None,
    text_fields: dict[str, str],
) -> tuple[str | None, str | None, str | None]:
    """
    Corrige clasificaciones de carrera que son formalmente válidas pero semánticamente
    engañosas para los patrones observados en la carga Scopus ULima.
    """
    if not carrera or carrera not in CAREER_AREA_LINE_CATALOG:
        return area_carrera, linea_carrera, None

    fallback_area, fallback_line, fallback_source = career_fallback_from_domain_rules(carrera, text_fields)

    if carrera == "Ingeniería Civil":
        if linea_carrera == "Calidad del Agua" and not contains_any_phrase_in_text_fields(text_fields, CIVIL_WATER_STRONG_SIGNALS):
            return fallback_area, fallback_line, fallback_source or "career_guardrail_civil_no_water_evidence"

        if linea_carrera in ("Hidrología e Hidráulica", "Riego y Drenaje", "Hidrogeología", "Transporte de Sedimentos") and not contains_any_phrase_in_text_fields(text_fields, CIVIL_WATER_STRONG_SIGNALS):
            return fallback_area, fallback_line, fallback_source or "career_guardrail_civil_no_hydraulic_evidence"

    if carrera == "Ingeniería de Sistemas":
        if linea_carrera == "Visión computacional" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS):
            return fallback_area, fallback_line, fallback_source or "career_guardrail_systems_no_computer_vision_evidence"

        if linea_carrera == "Realidad virtual y aumentada" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ARVR_STRONG_SIGNALS):
            return fallback_area, fallback_line, fallback_source or "career_guardrail_systems_no_arvr_evidence"

        if linea_carrera == "Ingeniería de software" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_SOFTWARE_STRONG_SIGNALS):
            return fallback_area, fallback_line, fallback_source or "career_guardrail_systems_no_software_evidence"

        if linea_carrera == "Aprendizaje automático" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ML_STRONG_SIGNALS):
            return fallback_area, fallback_line, fallback_source or "career_guardrail_systems_no_ml_evidence"

        if linea_carrera == "Construcción de juegos y gamificación" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_GAME_STRONG_SIGNALS):
            return fallback_area, fallback_line, fallback_source or "career_guardrail_systems_no_game_evidence"

        if linea_carrera == "Agentes virtuales" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_AGENT_STRONG_SIGNALS):
            return fallback_area, fallback_line, fallback_source or "career_guardrail_systems_no_agent_evidence"

        if linea_carrera == "Redes y ciberseguridad":
            cyber_present = contains_any_phrase_in_text_fields(text_fields, SYSTEMS_CYBER_STRONG_SIGNALS)
            vision_present = contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS)
            if vision_present and not cyber_present:
                return "Aplicaciones en inteligencia artificial", "Visión computacional", "career_guardrail_systems_vision_over_security"

    if carrera == "Ingeniería Industrial":
        # Priorizar SCM cuando la evidencia central es inventario, almacén, logística, EOQ/MRP, slotting, fill rate o supply chain.
        # Esto evita que papers logísticos caigan por defecto en Operations Engineering & Management.
        if contains_any_phrase_in_text_fields(text_fields, INDUSTRIAL_SCM_STRONG_SIGNALS):
            if contains_any_phrase_in_text_fields(text_fields, ["supply chain", "cadena de suministro"]):
                return "Supply Chain Management", "Gestión de la cadena de suministro", "career_guardrail_industrial_supply_chain_over_operations"
            return "Supply Chain Management", "Gestión de Inventarios, Almacenes y Transportes", "career_guardrail_industrial_inventory_warehouse_over_operations"

        if area_carrera == "Operations Research & Analysis" and contains_any_phrase_in_text_fields(text_fields, CHEM_MATERIAL_PRODUCT_STRONG_SIGNALS):
            return "Product Design & Development", "Desarrollo de producto", "career_guardrail_industrial_product_over_ora"

    if not area_carrera or not linea_carrera:
        return fallback_area, fallback_line, fallback_source

    return area_carrera, linea_carrera, None

def has_climate_adaptation_evidence(text_fields: dict[str, str]) -> bool:
    if contains_any_phrase_in_text_fields(text_fields, IDIC_CLIMATE_ADAPTATION_EXACT_SIGNALS):
        return True

    climate_present = contains_any_phrase_in_text_fields(text_fields, IDIC_CLIMATE_TERMS)
    adaptation_context_present = contains_any_phrase_in_text_fields(
        text_fields,
        IDIC_ADAPTATION_CONTEXT_TERMS,
    )
    return climate_present and adaptation_context_present


def choose_non_climate_sustainability_idic_alternative(
    text_fields: dict[str, str],
) -> tuple[str | None, str | None, str | None, str | None]:
    if contains_any_phrase_in_text_fields(text_fields, IDIC_CIRCULAR_ECONOMY_STRONG_SIGNALS):
        return (
            "Desarrollo sostenible y medioambiente",
            "Sostenibilidad y cambio climático",
            "Economía circular",
            "idic_guardrail_circular_economy_over_climate_adaptation",
        )

    if contains_any_phrase_in_text_fields(text_fields, IDIC_CLEAN_TECH_STRONG_SIGNALS):
        return (
            "Desarrollo sostenible y medioambiente",
            "Tecnología y ecosistemas",
            "Tecnologías limpias",
            "idic_guardrail_clean_tech_over_climate_adaptation",
        )

    if contains_any_phrase_in_text_fields(text_fields, IDIC_WASTE_MANAGEMENT_STRONG_SIGNALS):
        return (
            "Desarrollo sostenible y medioambiente",
            "Tecnología y ecosistemas",
            "Gestión de residuos",
            "idic_guardrail_waste_over_climate_adaptation",
        )

    if contains_any_phrase_in_text_fields(text_fields, IDIC_ADVANCED_MATERIALS_STRONG_SIGNALS):
        return (
            "Desarrollo sostenible y medioambiente",
            "Tecnología y ecosistemas",
            "Materiales avanzados",
            "idic_guardrail_materials_over_climate_adaptation",
        )

    if contains_any_phrase_in_text_fields(text_fields, IDIC_SUSTAINABLE_INFRASTRUCTURE_STRONG_SIGNALS):
        return (
            "Desarrollo sostenible y medioambiente",
            "Ciudades inteligentes y sostenibles",
            "Infraestructura sostenible",
            "idic_guardrail_infrastructure_over_climate_adaptation",
        )

    return None, None, None, None


def choose_general_idic_fallback(
    text_fields: dict[str, str],
) -> tuple[str, str, str, str]:
    """
    Fallback final obligatorio para evitar NULLs en IDIC.

    Esta función solo se usa cuando LLM + hints + guardrails no logran producir
    una clasificación IDIC completa. Mantiene el criterio general del proyecto:
    elegir la alternativa más cercana dentro del catálogo, sin usar
    'Adaptación al cambio climático' salvo que exista evidencia climática explícita.
    """
    alternative_category, alternative_area, alternative_line, source = choose_non_climate_sustainability_idic_alternative(text_fields)
    if alternative_category and alternative_area and alternative_line:
        return alternative_category, alternative_area, alternative_line, source or "idic_general_fallback_sustainability"

    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS):
        return (
            "Innovación y tecnología digital",
            "Inteligencia artificial y computación avanzada",
            "Visión computacional",
            "idic_general_fallback_computer_vision",
        )

    if contains_any_phrase_in_text_fields(text_fields, IDIC_MACHINE_LEARNING_STRONG_SIGNALS + SYSTEMS_ML_STRONG_SIGNALS):
        return (
            "Innovación y tecnología digital",
            "Inteligencia artificial y computación avanzada",
            "Machine learning y deep learning",
            "idic_general_fallback_machine_learning",
        )

    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_CYBER_STRONG_SIGNALS):
        return (
            "Innovación y tecnología digital",
            "Transformación digital",
            "Ciberseguridad y privacidad",
            "idic_general_fallback_cybersecurity",
        )

    if contains_any_phrase_in_text_fields(text_fields, ["iot", "internet of things"]):
        return (
            "Innovación y tecnología digital",
            "Transformación digital",
            "Internet de las cosas (IoT)",
            "idic_general_fallback_iot",
        )

    if contains_any_phrase_in_text_fields(text_fields, CIVIL_BIM_VDC_STRONG_SIGNALS):
        return (
            "Innovación y tecnología digital",
            "Transformación digital",
            "Diseño y construcción virtual",
            "idic_general_fallback_bim_vdc",
        )

    if contains_any_phrase_in_text_fields(text_fields, IDIC_INNOVATION_MANAGEMENT_STRONG_SIGNALS + INDUSTRIAL_EXTENDED_STRONG_SIGNALS):
        return (
            "Gestión y economía del conocimiento",
            "Innovación empresarial",
            "Gestión de la innovación",
            "idic_general_fallback_innovation_management",
        )

    if contains_any_phrase_in_text_fields(text_fields, IDIC_EMERGING_TECH_STRONG_SIGNALS):
        return (
            "Innovación y tecnología digital",
            "Transformación digital",
            "Tecnologías emergentes",
            "idic_general_fallback_emerging_technologies",
        )

    # Último recurso institucional: nunca dejar IDIC en NULL.
    # Se elige gestión de innovación como comodín de menor riesgo para papers de mejora/propuesta,
    # evitando inflar 'Tecnologías emergentes' sin evidencia.
    return (
        "Gestión y economía del conocimiento",
        "Innovación empresarial",
        "Gestión de la innovación",
        "idic_general_fallback_default_innovation_management",
    )


def is_idic_line_eligible_by_domain_rules(linea: str, text_fields: dict[str, str]) -> bool:
    innovation_present = contains_any_phrase_in_text_fields(
        text_fields,
        IDIC_INNOVATION_MANAGEMENT_STRONG_SIGNALS,
    )
    transformation_present = contains_any_phrase_in_text_fields(
        text_fields,
        IDIC_ORG_TRANSFORMATION_STRONG_SIGNALS,
    )

    if linea == "Transformación organizacional" and innovation_present and not transformation_present:
        return False

    if linea == "Adaptación al cambio climático" and not has_climate_adaptation_evidence(text_fields):
        return False

    return True


def apply_idic_line_bonus(linea: str, text_fields: dict[str, str], base_score: int) -> int:
    title_text = text_fields.get("title")
    abstract_text = text_fields.get("abstract")
    keyword_text = " | ".join(
        [
            text_fields.get("author_keywords", ""),
            text_fields.get("index_keywords", ""),
        ]
    )

    if linea == "Gestión de la innovación":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, IDIC_INNOVATION_MANAGEMENT_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, IDIC_INNOVATION_MANAGEMENT_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_INNOVATION_MANAGEMENT_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Transformación organizacional":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, IDIC_ORG_TRANSFORMATION_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, IDIC_ORG_TRANSFORMATION_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_ORG_TRANSFORMATION_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Machine learning y deep learning":
        bonus = 0
        bonus += 6 * count_phrase_hits_in_text(title_text, IDIC_MACHINE_LEARNING_STRONG_SIGNALS)
        bonus += 3 * count_phrase_hits_in_text(keyword_text, IDIC_MACHINE_LEARNING_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(abstract_text, IDIC_MACHINE_LEARNING_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Adaptación al cambio climático":
        bonus = 0
        bonus += 5 * count_phrase_hits_in_text(title_text, IDIC_CLIMATE_ADAPTATION_EXACT_SIGNALS)
        bonus += 3 * count_phrase_hits_in_text(keyword_text, IDIC_CLIMATE_ADAPTATION_EXACT_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(abstract_text, IDIC_CLIMATE_ADAPTATION_EXACT_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(title_text, IDIC_CLIMATE_TERMS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_CLIMATE_TERMS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_ADAPTATION_CONTEXT_TERMS)
        return base_score + bonus

    if linea == "Economía circular":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, IDIC_CIRCULAR_ECONOMY_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, IDIC_CIRCULAR_ECONOMY_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_CIRCULAR_ECONOMY_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Tecnologías limpias":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, IDIC_CLEAN_TECH_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, IDIC_CLEAN_TECH_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_CLEAN_TECH_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Gestión de residuos":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, IDIC_WASTE_MANAGEMENT_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, IDIC_WASTE_MANAGEMENT_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_WASTE_MANAGEMENT_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Materiales avanzados":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, IDIC_ADVANCED_MATERIALS_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, IDIC_ADVANCED_MATERIALS_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_ADVANCED_MATERIALS_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Infraestructura sostenible":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, IDIC_SUSTAINABLE_INFRASTRUCTURE_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, IDIC_SUSTAINABLE_INFRASTRUCTURE_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_SUSTAINABLE_INFRASTRUCTURE_STRONG_SIGNALS)
        return base_score + bonus

    if linea == "Tecnologías emergentes":
        bonus = 0
        bonus += 4 * count_phrase_hits_in_text(title_text, IDIC_EMERGING_TECH_STRONG_SIGNALS)
        bonus += 2 * count_phrase_hits_in_text(keyword_text, IDIC_EMERGING_TECH_STRONG_SIGNALS)
        bonus += 1 * count_phrase_hits_in_text(abstract_text, IDIC_EMERGING_TECH_STRONG_SIGNALS)
        return base_score + bonus

    return base_score


def choose_best_idic_line_with_rules(
    score_map: dict[str, int],
    text_fields: dict[str, str],
    min_score: int,
    min_margin: int = 1,
    relaxed: bool = False,
) -> str | None:
    filtered: list[tuple[str, int]] = []

    for linea, score in score_map.items():
        if score < min_score:
            continue
        if not is_idic_line_eligible_by_domain_rules(linea, text_fields):
            continue
        adjusted_score = apply_idic_line_bonus(linea, text_fields, score)
        filtered.append((linea, adjusted_score))

    if not filtered:
        return None

    filtered.sort(key=lambda x: (-x[1], normalize_generic_text(x[0])))

    if relaxed:
        return filtered[0][0]

    best_linea, best_score = filtered[0]
    tied = [linea for linea, score in filtered if score == best_score]
    if len(tied) > 1:
        return None

    if len(filtered) > 1:
        second_score = filtered[1][1]
        if (best_score - second_score) < min_margin:
            return None

    return best_linea


def apply_final_idic_guardrails(
    category_tematica: str | None,
    area_idic: str | None,
    linea_idic: str | None,
    text_fields: dict[str, str],
) -> tuple[str | None, str | None, str | None, str | None]:
    if not category_tematica or not area_idic or not linea_idic:
        return category_tematica, area_idic, linea_idic, None

    innovation_present = contains_any_phrase_in_text_fields(
        text_fields,
        IDIC_INNOVATION_MANAGEMENT_STRONG_SIGNALS,
    )
    transformation_present = contains_any_phrase_in_text_fields(
        text_fields,
        IDIC_ORG_TRANSFORMATION_STRONG_SIGNALS,
    )

    if (
        category_tematica == "Gestión y economía del conocimiento"
        and area_idic == "Innovación empresarial"
        and linea_idic in ("Gestión de la innovación", "Transformación organizacional")
        and contains_any_phrase_in_text_fields(text_fields, IDIC_ADVANCED_MATERIALS_STRONG_SIGNALS)
        and not innovation_present
    ):
        return (
            "Desarrollo sostenible y medioambiente",
            "Tecnología y ecosistemas",
            "Materiales avanzados",
            "idic_guardrail_materials_over_generic_innovation",
        )

    if (
        category_tematica == "Gestión y economía del conocimiento"
        and area_idic == "Innovación empresarial"
        and linea_idic == "Transformación organizacional"
        and innovation_present
        and not transformation_present
    ):
        return (
            category_tematica,
            area_idic,
            "Gestión de la innovación",
            "idic_guardrail_innovation_over_transformation",
        )

    if linea_idic == "Adaptación al cambio climático" and not has_climate_adaptation_evidence(text_fields):
        alternative_category, alternative_area, alternative_line, source = choose_non_climate_sustainability_idic_alternative(
            text_fields
        )
        if alternative_category and alternative_area and alternative_line:
            return alternative_category, alternative_area, alternative_line, source

        fallback_category, fallback_area, fallback_line, fallback_source = choose_general_idic_fallback(text_fields)
        return fallback_category, fallback_area, fallback_line, fallback_source

    # Evitar que "Tecnologías emergentes" opere como comodín cuando existe una
    # señal temática más específica y defendible.
    if (
        category_tematica == "Innovación y tecnología digital"
        and area_idic == "Transformación digital"
        and linea_idic == "Tecnologías emergentes"
    ):
        alternative_category, alternative_area, alternative_line, source = choose_non_climate_sustainability_idic_alternative(
            text_fields
        )
        if alternative_category and alternative_area and alternative_line:
            return alternative_category, alternative_area, alternative_line, source or "idic_guardrail_specific_over_emerging"

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS):
            return (
                "Innovación y tecnología digital",
                "Inteligencia artificial y computación avanzada",
                "Visión computacional",
                "idic_guardrail_computer_vision_over_emerging",
            )

        if contains_any_phrase_in_text_fields(text_fields, IDIC_MACHINE_LEARNING_STRONG_SIGNALS + SYSTEMS_ML_STRONG_SIGNALS):
            return (
                "Innovación y tecnología digital",
                "Inteligencia artificial y computación avanzada",
                "Machine learning y deep learning",
                "idic_guardrail_ml_over_emerging",
            )

        if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_CYBER_STRONG_SIGNALS):
            return (
                "Innovación y tecnología digital",
                "Transformación digital",
                "Ciberseguridad y privacidad",
                "idic_guardrail_cyber_over_emerging",
            )

        if contains_any_phrase_in_text_fields(text_fields, ["iot", "internet of things"]):
            return (
                "Innovación y tecnología digital",
                "Transformación digital",
                "Internet de las cosas (IoT)",
                "idic_guardrail_iot_over_emerging",
            )

        if innovation_present:
            return (
                "Gestión y economía del conocimiento",
                "Innovación empresarial",
                "Gestión de la innovación",
                "idic_guardrail_innovation_over_generic_emerging",
            )

    return category_tematica, area_idic, linea_idic, None

def append_classification_source(result: dict, source_name: str) -> None:
    current = result.get("classification_mode")
    if not current:
        result["classification_mode"] = source_name
        return

    parts = [p.strip() for p in str(current).split("|") if p.strip()]
    if source_name not in parts:
        parts.append(source_name)
    result["classification_mode"] = "|".join(parts)


def set_first_justification(result: dict, justification: str | None) -> None:
    if justification and not result.get("justification"):
        result["justification"] = justification


def classify_career_dimensions_force_best(
    carrera: str | None,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict:
    if not carrera or carrera not in CAREER_AREA_LINE_CATALOG:
        return {"area_carrera_raw": None, "linea_carrera_raw": None}

    text_fields = build_thematic_text_fields(
        title_value,
        abstract_value,
        author_keywords_value,
        index_keywords_value,
        source_title_value,
    )

    line_scores: dict[str, int] = {}
    for area_name, lineas in CAREER_AREA_LINE_CATALOG[carrera].items():
        for linea in lineas:
            primary_aliases = [linea]
            support_aliases = [area_name]
            support_aliases.extend(CAREER_LINE_HINTS.get(carrera, {}).get(linea, []))
            score = build_weighted_candidate_score(
                text_fields=text_fields,
                primary_aliases=primary_aliases,
                support_aliases=support_aliases,
            )
            score = apply_industrial_line_bonus(carrera, linea, text_fields, score)
            line_scores[linea] = score

    fallback_area, fallback_line, _fallback_source = career_fallback_from_domain_rules(carrera, text_fields)

    # Si no hay evidencia temática suficiente, usar fallback controlado por dominio,
    # no orden alfabético del catálogo.
    max_score = max(line_scores.values()) if line_scores else 0
    if max_score < THEMATIC_APPROX_MIN_SCORE and fallback_area and fallback_line:
        return {"area_carrera_raw": fallback_area, "linea_carrera_raw": fallback_line}

    eligible = [
        linea for linea in line_scores.keys()
        if is_line_eligible_by_domain_rules(carrera, linea, text_fields)
    ]
    candidate_pool = eligible or list(line_scores.keys())
    if not candidate_pool:
        return {"area_carrera_raw": fallback_area, "linea_carrera_raw": fallback_line}

    candidate_pool.sort(key=lambda x: (-line_scores.get(x, 0), normalize_generic_text(x)))
    best_linea = candidate_pool[0]
    best_area = coerce_area_carrera_from_linea(carrera, best_linea)

    best_area, best_linea, _guardrail_source = apply_final_career_guardrails(
        carrera,
        best_area,
        best_linea,
        text_fields,
    )

    return {"area_carrera_raw": best_area, "linea_carrera_raw": best_linea}


def classify_idic_dimensions_force_best(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict:
    text_fields = build_thematic_text_fields(
        title_value,
        abstract_value,
        author_keywords_value,
        index_keywords_value,
        source_title_value,
    )

    line_scores: dict[str, int] = {}
    for category_name, areas in IDIC_CATEGORY_AREA_LINE_CATALOG.items():
        for area_name, lineas in areas.items():
            for linea in lineas:
                primary_aliases = [linea]
                support_aliases = [area_name, category_name]
                support_aliases.extend(IDIC_LINE_HINTS.get(linea, []))
                base_score = build_weighted_candidate_score(
                    text_fields=text_fields,
                    primary_aliases=primary_aliases,
                    support_aliases=support_aliases,
                )
                if not is_idic_line_eligible_by_domain_rules(linea, text_fields):
                    continue
                line_scores[linea] = apply_idic_line_bonus(linea, text_fields, base_score)

    if not line_scores:
        return {
            "category_tematica_raw": None,
            "area_idic_raw": None,
            "linea_idic_raw": None,
        }

    ranked = sorted(line_scores.keys(), key=lambda x: (-line_scores.get(x, 0), normalize_generic_text(x)))
    best_linea = ranked[0]
    best_score = line_scores.get(best_linea, 0)

    if best_score < THEMATIC_APPROX_MIN_SCORE:
        fallback_category, fallback_area, fallback_line, _fallback_source = choose_general_idic_fallback(text_fields)
        return {
            "category_tematica_raw": fallback_category,
            "area_idic_raw": fallback_area,
            "linea_idic_raw": fallback_line,
        }

    best_area = coerce_area_idic_from_linea(best_linea)
    best_category = coerce_category_tematica_from_area(best_area)

    (
        best_category,
        best_area,
        best_linea,
        _guardrail_source,
    ) = apply_final_idic_guardrails(best_category, best_area, best_linea, text_fields)

    if not is_valid_idic_triplet(best_category, best_area, best_linea):
        fallback_category, fallback_area, fallback_line, _fallback_source = choose_general_idic_fallback(text_fields)
        best_category, best_area, best_linea = fallback_category, fallback_area, fallback_line

    return {
        "category_tematica_raw": best_category,
        "area_idic_raw": best_area,
        "linea_idic_raw": best_linea,
    }


def classify_career_dimensions_by_hints(
    carrera: str | None,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict:
    if not carrera or carrera not in CAREER_AREA_LINE_CATALOG:
        return {"area_carrera_raw": None, "linea_carrera_raw": None}

    text_fields = build_thematic_text_fields(
        title_value,
        abstract_value,
        author_keywords_value,
        index_keywords_value,
        source_title_value,
    )

    if not any(text_fields.values()):
        return {"area_carrera_raw": None, "linea_carrera_raw": None}

    line_scores: dict[str, int] = {}
    for area_name, lineas in CAREER_AREA_LINE_CATALOG[carrera].items():
        for linea in lineas:
            primary_aliases = [linea]
            support_aliases = [area_name]
            support_aliases.extend(CAREER_LINE_HINTS.get(carrera, {}).get(linea, []))
            line_scores[linea] = build_weighted_candidate_score(
                text_fields=text_fields,
                primary_aliases=primary_aliases,
                support_aliases=support_aliases,
            )

    best_linea = choose_best_line_with_rules(
        carrera=carrera,
        score_map=line_scores,
        text_fields=text_fields,
        min_score=THEMATIC_STRICT_MIN_SCORE,
        min_margin=THEMATIC_STRICT_MIN_MARGIN,
        relaxed=False,
    )
    best_area = coerce_area_carrera_from_linea(carrera, best_linea) if best_linea else None

    return {"area_carrera_raw": best_area, "linea_carrera_raw": best_linea}


def classify_career_dimensions_by_hints_approx(
    carrera: str | None,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict:
    if not carrera or carrera not in CAREER_AREA_LINE_CATALOG:
        return {"area_carrera_raw": None, "linea_carrera_raw": None}

    text_fields = build_thematic_text_fields(
        title_value,
        abstract_value,
        author_keywords_value,
        index_keywords_value,
        source_title_value,
    )

    if not any(text_fields.values()):
        return {"area_carrera_raw": None, "linea_carrera_raw": None}

    line_scores: dict[str, int] = {}
    for area_name, lineas in CAREER_AREA_LINE_CATALOG[carrera].items():
        for linea in lineas:
            primary_aliases = [linea]
            support_aliases = [area_name]
            support_aliases.extend(CAREER_LINE_HINTS.get(carrera, {}).get(linea, []))
            line_scores[linea] = build_weighted_candidate_score(
                text_fields=text_fields,
                primary_aliases=primary_aliases,
                support_aliases=support_aliases,
            )

    best_linea = choose_best_line_with_rules(
        carrera=carrera,
        score_map=line_scores,
        text_fields=text_fields,
        min_score=THEMATIC_APPROX_MIN_SCORE,
        min_margin=1,
        relaxed=True,
    )
    best_area = coerce_area_carrera_from_linea(carrera, best_linea) if best_linea else None

    return {"area_carrera_raw": best_area, "linea_carrera_raw": best_linea}


def classify_idic_dimensions_by_hints(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict:
    text_fields = build_thematic_text_fields(
        title_value,
        abstract_value,
        author_keywords_value,
        index_keywords_value,
        source_title_value,
    )

    if not any(text_fields.values()):
        return {
            "category_tematica_raw": None,
            "area_idic_raw": None,
            "linea_idic_raw": None,
        }

    line_scores: dict[str, int] = {}
    for category_name, areas in IDIC_CATEGORY_AREA_LINE_CATALOG.items():
        for area_name, lineas in areas.items():
            for linea in lineas:
                primary_aliases = [linea]
                support_aliases = [area_name, category_name]
                support_aliases.extend(IDIC_LINE_HINTS.get(linea, []))
                line_scores[linea] = build_weighted_candidate_score(
                    text_fields=text_fields,
                    primary_aliases=primary_aliases,
                    support_aliases=support_aliases,
                )

    best_linea = choose_best_idic_line_with_rules(
        score_map=line_scores,
        text_fields=text_fields,
        min_score=THEMATIC_STRICT_MIN_SCORE,
        min_margin=THEMATIC_STRICT_MIN_MARGIN,
        relaxed=False,
    )
    best_area = coerce_area_idic_from_linea(best_linea) if best_linea else None
    best_category = coerce_category_tematica_from_area(best_area) if best_area else None

    if best_category and best_area and best_linea:
        if not is_valid_idic_triplet(best_category, best_area, best_linea):
            best_category = None
            best_area = None
            best_linea = None

    return {
        "category_tematica_raw": best_category,
        "area_idic_raw": best_area,
        "linea_idic_raw": best_linea,
    }


def classify_idic_dimensions_by_hints_approx(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict:
    text_fields = build_thematic_text_fields(
        title_value,
        abstract_value,
        author_keywords_value,
        index_keywords_value,
        source_title_value,
    )

    if not any(text_fields.values()):
        return {
            "category_tematica_raw": None,
            "area_idic_raw": None,
            "linea_idic_raw": None,
        }

    line_scores: dict[str, int] = {}
    for category_name, areas in IDIC_CATEGORY_AREA_LINE_CATALOG.items():
        for area_name, lineas in areas.items():
            for linea in lineas:
                primary_aliases = [linea]
                support_aliases = [area_name, category_name]
                support_aliases.extend(IDIC_LINE_HINTS.get(linea, []))
                line_scores[linea] = build_weighted_candidate_score(
                    text_fields=text_fields,
                    primary_aliases=primary_aliases,
                    support_aliases=support_aliases,
                )

    best_linea = choose_best_idic_line_with_rules(
        score_map=line_scores,
        text_fields=text_fields,
        min_score=THEMATIC_APPROX_MIN_SCORE,
        min_margin=1,
        relaxed=True,
    )
    best_area = coerce_area_idic_from_linea(best_linea) if best_linea else None
    best_category = coerce_category_tematica_from_area(best_area) if best_area else None

    if best_category and best_area and best_linea:
        if not is_valid_idic_triplet(best_category, best_area, best_linea):
            best_category = None
            best_area = None
            best_linea = None

    return {
        "category_tematica_raw": best_category,
        "area_idic_raw": best_area,
        "linea_idic_raw": best_linea,
    }


# =========================
# AZURE OPENAI THEMATIC CLASSIFIER
# =========================
_AZURE_OPENAI_CLIENT = None


def is_thematic_llm_configured() -> bool:
    return bool(
        os.environ.get("AZURE_OPENAI_API_KEY")
        and (os.environ.get("AZURE_OPENAI_BASE_URL") or os.environ.get("AZURE_OPENAI_ENDPOINT"))
        and os.environ.get("AZURE_OPENAI_RESPONSES_MODEL")
    )


def get_azure_openai_base_url() -> str:
    explicit_base_url = os.environ.get("AZURE_OPENAI_BASE_URL")
    if explicit_base_url:
        return explicit_base_url.rstrip("/")

    endpoint = get_env("AZURE_OPENAI_ENDPOINT").rstrip("/")
    return f"{endpoint}/openai/v1"


def get_azure_openai_client():
    global _AZURE_OPENAI_CLIENT

    if _AZURE_OPENAI_CLIENT is None:
        from openai import OpenAI

        _AZURE_OPENAI_CLIENT = OpenAI(
            api_key=get_env("AZURE_OPENAI_API_KEY"),
            base_url=get_azure_openai_base_url(),
        )

    return _AZURE_OPENAI_CLIENT


def get_thematic_llm_min_confidence(mode: str = "strict") -> float:
    env_name = "THEMATIC_LLM_MIN_CONFIDENCE"
    default_value = "0.80"

    if mode == "approx":
        env_name = "THEMATIC_LLM_MIN_CONFIDENCE_APPROX"
        default_value = "0.55"

    raw = os.environ.get(env_name, default_value)
    try:
        value = float(raw)
    except Exception:
        value = float(default_value)

    if value < 0:
        return 0.0
    if value > 1:
        return 1.0
    return value


def build_thematic_empty_result() -> dict:
    return {
        "area_carrera_raw": None,
        "linea_carrera_raw": None,
        "category_tematica_raw": None,
        "area_idic_raw": None,
        "linea_idic_raw": None,
        "confidence": None,
        "justification": None,
        "classification_mode": None,
    }


def build_career_classification_prompt(
    carrera: str,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    mode: str = "strict",
) -> str:
    career_catalog = CAREER_AREA_LINE_CATALOG.get(carrera, {})

    rules = [
        "Clasifica SOLO area_carrera_raw y linea_carrera_raw.",
        "No clasifiques campos IDIC en esta tarea.",
        "Usa abstract_scopus como evidencia principal y criterio dominante para decidir el foco real del artículo.",
        "Usa title, author_keywords e index_keywords solo como apoyo o corroboración.",
        "Usa source_title solo como evidencia débil.",
        "Debes elegir solo valores existentes en el catálogo proporcionado.",
        "No inventes áreas ni líneas.",
        "Si eliges linea_carrera_raw, debe pertenecer a la carrera dada y ser consistente con area_carrera_raw.",
        "Para Ingeniería Industrial, NO elijas 'Operations Research & Analysis' salvo que el aporte central sea modelamiento, optimización, predicción, benchmarking, simulación o algoritmos.",
        "Para Ingeniería Industrial, favorece 'Supply Chain Management' cuando el foco sea inventarios, almacenes, logística, transporte, proveedores, cadena de suministro u OTIF.",
        "Para Ingeniería Industrial, favorece 'Operations Engineering & Management' cuando el foco sea 5S, standard work, lean, mejora de procesos, eficiencia operativa o gestión de operaciones.",
        "Para Ingeniería Industrial, favorece 'Product Design & Development' cuando el foco sea síntesis, materiales, fibras, nanopartículas, formulación, caracterización o desarrollo de producto/material.",
        "Evita decidir solo por palabras amplias si el título y el resumen apuntan a un tema más específico.",
    ]

    if mode == "strict":
        rules.append("Si no hay evidencia suficiente, devuelve null en los campos inciertos.")
    else:
        rules.append("Si no hay evidencia exacta, elige la opción más cercana y defendible dentro del catálogo.")

    payload = {
        "mode": mode,
        "task": "career_only",
        "carrera": carrera,
        "career_catalog": career_catalog,
        "article": {
            "title": clip_text(title_value, 2000),
            "abstract_scopus": clip_text(abstract_value, 9000),
            "author_keywords": clip_text(author_keywords_value, 2000),
            "index_keywords": clip_text(index_keywords_value, 2000),
            "source_title": clip_text(source_title_value, 1000),
        },
        "output_schema": {
            "area_carrera_raw": "string|null",
            "linea_carrera_raw": "string|null",
            "confidence": "number 0..1",
            "justification": "short string",
        },
        "rules": rules,
    }

    return (
        "Eres un clasificador temático académico. Analiza el artículo y clasifica solo la dimensión Carrera.\n\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )


def build_idic_classification_prompt(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    mode: str = "strict",
) -> str:
    rules = [
        "Clasifica SOLO category_tematica_raw, area_idic_raw y linea_idic_raw.",
        "No clasifiques campos de Carrera en esta tarea.",
        "Usa abstract_scopus como evidencia principal y criterio dominante para decidir el foco real del artículo.",
        "Usa title, author_keywords e index_keywords solo como apoyo o corroboración.",
        "Usa source_title solo como evidencia débil.",
        "Debes elegir solo valores existentes en el catálogo proporcionado.",
        "No inventes categorías, áreas ni líneas.",
        "Si eliges linea_idic_raw, debe ser consistente con area_idic_raw y category_tematica_raw.",
        "Dentro de 'Gestión y economía del conocimiento' > 'Innovación empresarial', favorece 'Gestión de la innovación' cuando el foco sea 5S, standard work, lean, mejora continua, mejora de procesos, eficiencia operativa o productividad.",
        "Dentro de 'Gestión y economía del conocimiento' > 'Innovación empresarial', usa 'Transformación organizacional' solo cuando el aporte central trate explícitamente de transformación, rediseño o cambio organizacional.",
        "Distingue claramente entre enfoques de tecnología digital, sostenibilidad y comportamiento humano según el aporte central del artículo.",
        "Usa 'Adaptación al cambio climático' solo si el artículo menciona explícitamente cambio climático, adaptación, resiliencia, vulnerabilidad, riesgo climático o impactos climáticos.",
        "Si el artículo trata materiales, síntesis, nanopartículas, concreto, adsorción, valorización de residuos o tecnologías limpias sin evidencia climática explícita, no uses 'Adaptación al cambio climático'.",
        "Evita decidir solo por palabras amplias si el título y el resumen apuntan a un tema más específico.",
    ]

    if mode == "strict":
        rules.append("Si no hay evidencia suficiente, devuelve null en los campos inciertos.")
    else:
        rules.append("Si no hay evidencia exacta, elige la opción más cercana y defendible dentro del catálogo.")

    payload = {
        "mode": mode,
        "task": "idic_only",
        "idic_catalog": IDIC_CATEGORY_AREA_LINE_CATALOG,
        "article": {
            "title": clip_text(title_value, 2000),
            "abstract_scopus": clip_text(abstract_value, 9000),
            "author_keywords": clip_text(author_keywords_value, 2000),
            "index_keywords": clip_text(index_keywords_value, 2000),
            "source_title": clip_text(source_title_value, 1000),
        },
        "output_schema": {
            "category_tematica_raw": "string|null",
            "area_idic_raw": "string|null",
            "linea_idic_raw": "string|null",
            "confidence": "number 0..1",
            "justification": "short string",
        },
        "rules": rules,
    }

    return (
        "Eres un clasificador temático académico. Analiza el artículo y clasifica solo la dimensión IDIC.\n\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )


def validate_llm_career_output(
    carrera: str | None,
    raw_output: dict | None,
    mode: str = "strict",
) -> dict:
    result = build_thematic_empty_result()

    if not carrera or carrera not in CAREER_AREA_LINE_CATALOG or not raw_output:
        return result

    confidence = parse_float_or_none(raw_output.get("confidence"))
    result["confidence"] = confidence
    result["justification"] = raw_output.get("justification")
    result["classification_mode"] = f"career_llm_{mode}"

    area_carrera = coerce_choice(raw_output.get("area_carrera_raw"), get_allowed_career_areas(carrera))
    linea_carrera = coerce_choice(raw_output.get("linea_carrera_raw"), get_allowed_career_lines(carrera))

    if linea_carrera and not area_carrera:
        area_carrera = coerce_area_carrera_from_linea(carrera, linea_carrera)

    if not is_valid_career_area_line(carrera, area_carrera, linea_carrera):
        area_carrera = None
        linea_carrera = None

    min_confidence = get_thematic_llm_min_confidence(mode=mode)
    if confidence is not None and confidence < min_confidence:
        return build_thematic_empty_result()

    result["area_carrera_raw"] = area_carrera
    result["linea_carrera_raw"] = linea_carrera
    return result


def validate_llm_idic_output(
    raw_output: dict | None,
    mode: str = "strict",
) -> dict:
    result = build_thematic_empty_result()

    if not raw_output:
        return result

    confidence = parse_float_or_none(raw_output.get("confidence"))
    result["confidence"] = confidence
    result["justification"] = raw_output.get("justification")
    result["classification_mode"] = f"idic_llm_{mode}"

    category_tematica = coerce_choice(raw_output.get("category_tematica_raw"), get_allowed_idic_categories())
    area_idic = coerce_choice(raw_output.get("area_idic_raw"), get_allowed_idic_areas(category_tematica))

    if not area_idic:
        area_idic = coerce_choice(raw_output.get("area_idic_raw"), get_allowed_idic_areas())

    linea_idic = coerce_choice(raw_output.get("linea_idic_raw"), get_allowed_idic_lines(category_tematica, area_idic))

    if not linea_idic:
        linea_idic = coerce_choice(raw_output.get("linea_idic_raw"), get_allowed_idic_lines())

    if linea_idic and not area_idic:
        area_idic = coerce_area_idic_from_linea(linea_idic)

    if area_idic and not category_tematica:
        category_tematica = coerce_category_tematica_from_area(area_idic)

    if not is_valid_idic_triplet(category_tematica, area_idic, linea_idic):
        category_tematica = None
        area_idic = None
        linea_idic = None

    min_confidence = get_thematic_llm_min_confidence(mode=mode)
    if confidence is not None and confidence < min_confidence:
        return build_thematic_empty_result()

    result["category_tematica_raw"] = category_tematica
    result["area_idic_raw"] = area_idic
    result["linea_idic_raw"] = linea_idic
    return result


def classify_career_with_llm(
    carrera: str | None,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    mode: str = "strict",
) -> dict:
    empty_result = build_thematic_empty_result()

    if not carrera or carrera not in CAREER_AREA_LINE_CATALOG:
        return empty_result

    if not is_thematic_llm_configured():
        return empty_result

    prompt = build_career_classification_prompt(
        carrera=carrera,
        title_value=title_value,
        abstract_value=abstract_value,
        author_keywords_value=author_keywords_value,
        index_keywords_value=index_keywords_value,
        source_title_value=source_title_value,
        mode=mode,
    )

    try:
        client = get_azure_openai_client()
        response = client.responses.create(
            model=get_env("AZURE_OPENAI_RESPONSES_MODEL"),
            input=prompt,
        )

        raw_output = extract_json_object(getattr(response, "output_text", None))
        if not raw_output:
            logging.warning("Career LLM (%s) returned non-JSON output.", mode)
            return empty_result

        return validate_llm_career_output(carrera, raw_output, mode=mode)

    except Exception as exc:
        logging.warning("Career LLM classification failed (%s): %s", mode, str(exc))
        return empty_result


def classify_idic_with_llm(
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    mode: str = "strict",
) -> dict:
    empty_result = build_thematic_empty_result()

    if not is_thematic_llm_configured():
        return empty_result

    prompt = build_idic_classification_prompt(
        title_value=title_value,
        abstract_value=abstract_value,
        author_keywords_value=author_keywords_value,
        index_keywords_value=index_keywords_value,
        source_title_value=source_title_value,
        mode=mode,
    )

    try:
        client = get_azure_openai_client()
        response = client.responses.create(
            model=get_env("AZURE_OPENAI_RESPONSES_MODEL"),
            input=prompt,
        )

        raw_output = extract_json_object(getattr(response, "output_text", None))
        if not raw_output:
            logging.warning("IDIC LLM (%s) returned non-JSON output.", mode)
            return empty_result

        return validate_llm_idic_output(raw_output, mode=mode)

    except Exception as exc:
        logging.warning("IDIC LLM classification failed (%s): %s", mode, str(exc))
        return empty_result



# =========================
# THEMATIC REVIEW AI + DETERMINISTIC OVERRIDES (run80)
# =========================
KNOWLEDGE_MANAGEMENT_SIGNALS = [
    "knowledge management", "gestion del conocimiento", "gestión del conocimiento",
    "absorptive capacity", "absorptive capacities", "capacidad de absorcion", "capacidad de absorción",
    "knowledge acquisition", "knowledge sharing", "knowledge creation", "knowledge exchange",
    "unlearning", "organizational learning", "aprendizaje organizacional",
    "organizational performance", "innovation performance", "value chain", "value chains",
    "microenterprises", "quality management", "telecommuting", "shadow information technology", "shadow it",
]

COMPUTING_EDUCATION_SIGNALS = [
    "computer science education", "computing education", "information systems education",
    "engineering education", "programming education", "introduction to programming",
    "computational thinking", "yupana", "mooc", "moocs", "online education",
    "higher education", "curricular design", "research competencies", "undergraduate students",
]

IT_GOVERNANCE_SIGNALS = [
    "it governance", "cobit", "iso 38500", "information systems", "shadow it",
    "information technology adoption", "benefit dependency network", "information technology",
]

BLOCKCHAIN_SIGNALS = [
    "blockchain", "smart contract", "smart contracts", "ethereum", "proof of work", "traceability",
]

SYSTEMS_ALGORITHM_SIGNALS = [
    "compiler", "source to source", "automatic parallelization", "parallelization", "mutation testing",
    "genetic algorithm", "functional programming", "algorithm", "algorithms", "dynamic programming",
    "pattern recognition", "route planning", "vehicle routing", "ant colony", "k-means", "k means",
]

EDUCATION_IDIC_SIGNALS = COMPUTING_EDUCATION_SIGNALS + [
    "education", "educational", "learning", "teaching", "students", "teachers", "school", "curriculum",
]

ENVIRONMENTAL_WATER_AIR_SIGNALS = [
    "air pollution", "air pollutants", "indoor air quality", "water quality", "water treatment",
    "wastewater", "groundwater", "aquifer", "pathogen migration", "subsurface", "sediments",
    "suspended sediments", "marine pollution", "phosphogypsum", "naproxen", "paracetamol",
    "activated carbon", "adsorption", "contaminant", "contaminants", "pollutant", "pollutants",
    "nanoparticles", "hazardous elements", "particulate matter", "pm1", "pm2.5", "pm10",
]

COMMERCIAL_PROCESS_INDUSTRIAL_RESCUE_SIGNALS = [
    "commercial model", "sales cycle", "lead handling", "crm", "bpm", "scrum", "sales",
    "retail", "automotive retailer", "business management model", "lean methodologies",
    "kanban", "tpm", "5s", "delivery time", "textile factory", "factory", "process improvement",
    "productivity", "operations", "operational", "poka yoke", "forecasting",
]


def classify_systems_specific_line_from_domain(text_fields: dict[str, str]) -> tuple[str | None, str | None, str | None]:
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS):
        return "Aplicaciones en inteligencia artificial", "Visión computacional", "run80_systems_computer_vision"
    if contains_any_phrase_in_text_fields(text_fields, ["nlp", "natural language processing", "large language model", "llm"]):
        return "Aplicaciones en inteligencia artificial", "NLP", "run80_systems_nlp"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ML_STRONG_SIGNALS):
        return "Aplicaciones en inteligencia artificial", "Aprendizaje automático", "run80_systems_ml"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_CYBER_STRONG_SIGNALS + BLOCKCHAIN_SIGNALS):
        return "Sistemas de Tecnologías de Información (TI)", "Redes y ciberseguridad", "run80_systems_security_blockchain"
    if contains_any_phrase_in_text_fields(text_fields, ["iot", "internet of things", "wireless sensor network", "sensor network"]):
        return "Sistemas de Tecnologías de Información (TI)", "IoT", "run80_systems_iot"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ARVR_STRONG_SIGNALS):
        return "Interacción humano-computadora", "Realidad virtual y aumentada", "run80_systems_arvr"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_GAME_STRONG_SIGNALS):
        return "Interacción humano-computadora", "Construcción de juegos y gamificación", "run80_systems_gamification"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_AGENT_STRONG_SIGNALS):
        return "Interacción humano-computadora", "Agentes virtuales", "run80_systems_virtual_agents"
    if contains_any_phrase_in_text_fields(text_fields, KNOWLEDGE_MANAGEMENT_SIGNALS):
        return "Tecnologías y gestión de la información", "Sistemas de gestión del conocimiento", "run80_systems_knowledge_management"
    if contains_any_phrase_in_text_fields(text_fields, IT_GOVERNANCE_SIGNALS):
        return "Tecnologías y gestión de la información", "Gestión de procesos tecnológicos", "run80_systems_it_governance"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ALGORITHM_SIGNALS):
        return "Algoritmos y sistemas computacionales", "Diseño de algoritmos", "run80_systems_algorithms"
    if contains_any_phrase_in_text_fields(text_fields, COMPUTING_EDUCATION_SIGNALS):
        return "Tecnologías y gestión de la información", "Computación aplicada", "run80_systems_computing_education"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_SOFTWARE_STRONG_SIGNALS):
        return "Algoritmos y sistemas computacionales", "Ingeniería de software", "run80_systems_software"
    return None, None, None


def classify_idic_specific_line_from_domain(text_fields: dict[str, str]) -> tuple[str | None, str | None, str | None, str | None]:
    if contains_any_phrase_in_text_fields(text_fields, ENVIRONMENTAL_WATER_AIR_SIGNALS):
        if contains_any_phrase_in_text_fields(text_fields, IDIC_ADVANCED_MATERIALS_STRONG_SIGNALS):
            return "Desarrollo sostenible y medioambiente", "Tecnología y ecosistemas", "Materiales avanzados", "run80_idic_environment_materials"
        return "Desarrollo sostenible y medioambiente", "Tecnología y ecosistemas", "Gestión de residuos", "run80_idic_environment_waste"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS):
        return "Innovación y tecnología digital", "Inteligencia artificial y computación avanzada", "Visión computacional", "run80_idic_computer_vision"
    if contains_any_phrase_in_text_fields(text_fields, IDIC_MACHINE_LEARNING_STRONG_SIGNALS + SYSTEMS_ML_STRONG_SIGNALS):
        return "Innovación y tecnología digital", "Inteligencia artificial y computación avanzada", "Machine learning y deep learning", "run80_idic_ml"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_CYBER_STRONG_SIGNALS + BLOCKCHAIN_SIGNALS):
        return "Innovación y tecnología digital", "Transformación digital", "Ciberseguridad y privacidad", "run80_idic_cyber_blockchain"
    if contains_any_phrase_in_text_fields(text_fields, ["iot", "internet of things", "wireless sensor network", "sensor network"]):
        return "Innovación y tecnología digital", "Transformación digital", "Internet de las cosas (IoT)", "run80_idic_iot"
    if contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ARVR_STRONG_SIGNALS):
        return "Innovación y tecnología digital", "Experiencia digital humana", "Realidad virtual y aumentada", "run80_idic_arvr"
    if contains_any_phrase_in_text_fields(text_fields, EDUCATION_IDIC_SIGNALS):
        return "Sociedad y comportamiento humano", "Bienestar y desarrollo humano", "Educación, desarrollo cognitivo y socioafectivo", "run80_idic_education"
    if contains_any_phrase_in_text_fields(text_fields, KNOWLEDGE_MANAGEMENT_SIGNALS):
        return "Gestión y economía del conocimiento", "Gestión del conocimiento", "Aprendizaje organizacional", "run80_idic_knowledge_management"
    if contains_any_phrase_in_text_fields(text_fields, COMMERCIAL_PROCESS_INDUSTRIAL_RESCUE_SIGNALS + IDIC_INNOVATION_MANAGEMENT_STRONG_SIGNALS):
        return "Gestión y economía del conocimiento", "Innovación empresarial", "Gestión de la innovación", "run80_idic_innovation_management"
    return None, None, None, None


def apply_run80_domain_specific_overrides(carrera: str | None, merged: dict, text_fields: dict[str, str]) -> tuple[dict, list[str]]:
    result = dict(merged)
    sources: list[str] = []

    if carrera == "Ingeniería de Sistemas":
        sensitive = result.get("linea_carrera_raw") in {
            "Realidad virtual y aumentada", "Visión computacional", "Ingeniería de software", "Agentes virtuales",
        }
        if sensitive:
            area, line, source = classify_systems_specific_line_from_domain(text_fields)
            if area and line and is_valid_career_area_line(carrera, area, line):
                if area != result.get("area_carrera_raw") or line != result.get("linea_carrera_raw"):
                    result["area_carrera_raw"] = area
                    result["linea_carrera_raw"] = line
                    sources.append(source or "run80_systems_sensitive_override")

        if contains_any_phrase_in_text_fields(text_fields, KNOWLEDGE_MANAGEMENT_SIGNALS):
            area, line = "Tecnologías y gestión de la información", "Sistemas de gestión del conocimiento"
            if is_valid_career_area_line(carrera, area, line):
                result["area_carrera_raw"] = area
                result["linea_carrera_raw"] = line
                sources.append("run80_systems_knowledge_management_override")

    if carrera == "Ingeniería Civil":
        if result.get("linea_carrera_raw") == "Calidad del Agua" or contains_any_phrase_in_text_fields(text_fields, ENVIRONMENTAL_WATER_AIR_SIGNALS):
            cat, area, line, source = classify_idic_specific_line_from_domain(text_fields)
            if cat and area and line and is_valid_idic_triplet(cat, area, line):
                result["category_tematica_raw"] = cat
                result["area_idic_raw"] = area
                result["linea_idic_raw"] = line
                sources.append(source or "run80_civil_environment_idic_override")

    if carrera == "Ingeniería Industrial":
        if contains_any_phrase_in_text_fields(text_fields, COMMERCIAL_PROCESS_INDUSTRIAL_RESCUE_SIGNALS):
            area, line = "Operations Engineering & Management", "Planeamiento y Gestión de Operaciones"
            if contains_any_phrase_in_text_fields(text_fields, ["inventory", "warehouse", "supply chain", "logistics"]):
                area, line = "Supply Chain Management", "Gestión de Inventarios, Almacenes y Transportes"
            if is_valid_career_area_line(carrera, area, line):
                result["area_carrera_raw"] = area
                result["linea_carrera_raw"] = line
                sources.append("run80_industrial_process_override")

    current_idic_line = result.get("linea_idic_raw")
    idic_is_sensitive_wrong = (
        (current_idic_line == "Realidad virtual y aumentada" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ARVR_STRONG_SIGNALS))
        or (current_idic_line == "Visión computacional" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS))
        or (current_idic_line == "Tecnologías emergentes" and not contains_any_phrase_in_text_fields(text_fields, IDIC_EMERGING_TECH_STRONG_SIGNALS))
    )
    if idic_is_sensitive_wrong:
        cat, area, line, source = classify_idic_specific_line_from_domain(text_fields)
        if cat and area and line and is_valid_idic_triplet(cat, area, line):
            result["category_tematica_raw"] = cat
            result["area_idic_raw"] = area
            result["linea_idic_raw"] = line
            sources.append(source or "run80_idic_sensitive_override")

    return result, unique_keep_order([s for s in sources if s])


def detect_thematic_review_reasons(carrera: str | None, merged: dict, text_fields: dict[str, str]) -> list[str]:
    reasons: list[str] = []

    if carrera == "Ingeniería de Sistemas":
        if merged.get("linea_carrera_raw") == "Visión computacional" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS):
            reasons.append("SYSTEMS_COMPUTER_VISION_WITHOUT_EVIDENCE")
        if merged.get("linea_carrera_raw") == "Realidad virtual y aumentada" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ARVR_STRONG_SIGNALS):
            reasons.append("SYSTEMS_ARVR_WITHOUT_EVIDENCE")
        if merged.get("linea_carrera_raw") == "Ingeniería de software" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_SOFTWARE_STRONG_SIGNALS + SYSTEMS_ALGORITHM_SIGNALS):
            reasons.append("SYSTEMS_SOFTWARE_WITHOUT_EVIDENCE")
        if contains_any_phrase_in_text_fields(text_fields, KNOWLEDGE_MANAGEMENT_SIGNALS) and merged.get("linea_carrera_raw") in {"Visión computacional", "Realidad virtual y aumentada", "Ingeniería de software"}:
            reasons.append("KNOWLEDGE_MANAGEMENT_MISROUTED_TO_SYSTEMS_SENSITIVE_LINE")

    if carrera == "Ingeniería Civil":
        if merged.get("linea_carrera_raw") == "Calidad del Agua" and merged.get("linea_idic_raw") in {"Realidad virtual y aumentada", "Visión computacional", "Tecnologías emergentes"}:
            reasons.append("CIVIL_ENVIRONMENT_WITH_DIGITAL_IDIC_MISMATCH")

    if merged.get("linea_idic_raw") == "Realidad virtual y aumentada" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_ARVR_STRONG_SIGNALS):
        reasons.append("IDIC_ARVR_WITHOUT_EVIDENCE")
    if merged.get("linea_idic_raw") == "Visión computacional" and not contains_any_phrase_in_text_fields(text_fields, SYSTEMS_COMPUTER_VISION_STRONG_SIGNALS):
        reasons.append("IDIC_COMPUTER_VISION_WITHOUT_EVIDENCE")

    return unique_keep_order(reasons)


def build_thematic_review_ai_prompt(
    carrera: str,
    current_result: dict,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    review_reasons: list[str],
) -> str:
    payload = {
        "task": "review_and_correct_thematic_classification_for_ulima_engineering_publication",
        "allowed_decisions": ["KEEP", "CORRECT", "REVIEW"],
        "career": carrera,
        "career_catalog_for_this_career": CAREER_AREA_LINE_CATALOG.get(carrera, {}),
        "idic_catalog": IDIC_CATEGORY_AREA_LINE_CATALOG,
        "current_classification": {
            "area_carrera_raw": current_result.get("area_carrera_raw"),
            "linea_carrera_raw": current_result.get("linea_carrera_raw"),
            "category_tematica_raw": current_result.get("category_tematica_raw"),
            "area_idic_raw": current_result.get("area_idic_raw"),
            "linea_idic_raw": current_result.get("linea_idic_raw"),
        },
        "review_reasons": review_reasons,
        "article": {
            "title": clip_text(title_value, 2200),
            "abstract_scopus": clip_text(abstract_value, 9000),
            "author_keywords": clip_text(author_keywords_value, 2500),
            "index_keywords": clip_text(index_keywords_value, 2500),
            "source_title": clip_text(source_title_value, 1200),
        },
        "output_schema": {
            "decision": "KEEP|CORRECT|REVIEW",
            "area_carrera_raw": "string|null",
            "linea_carrera_raw": "string|null",
            "category_tematica_raw": "string|null",
            "area_idic_raw": "string|null",
            "linea_idic_raw": "string|null",
            "confidence": "number 0..1",
            "rationale": "short Spanish explanation, max 45 words",
        },
        "rules": [
            "Corrige SOLO si el catálogo permite una alternativa claramente mejor.",
            "No cambies la carrera; solo revisa área/línea de carrera e IDIC.",
            "Usa el abstract como evidencia principal; título y keywords como apoyo.",
            "No uses Visión computacional sin evidencia de imágenes, video, detección, visión, YOLO, GAN, keypoints u homografía.",
            "No uses Realidad virtual y aumentada sin evidencia explícita de AR/VR/mixed reality.",
            "No uses Ingeniería de software solo por palabras genéricas como plataforma, online o educación.",
            "Para knowledge management/absorptive capacity/organizational learning, evita CV/VR/software salvo evidencia técnica fuerte.",
            "Para contaminación, agua, aire, sedimentos, adsorción o nanopartículas ambientales, favorece Desarrollo sostenible y medioambiente.",
            "Si la mejor corrección no es clara, devuelve REVIEW.",
            "Devuelve SOLO JSON válido, sin markdown.",
        ],
    }
    return (
        "Eres un auditor académico senior de clasificación temática de producción científica ULima. "
        "Revisa clasificaciones sospechosas y corrige con criterio conservador.\\n\\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )


def review_thematic_classification_with_ai(
    carrera: str,
    current_result: dict,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    review_reasons: list[str],
) -> dict:
    empty = {"decision": "NO_AI_RESULT", "confidence": None, "raw": None, "rationale": None}

    if not is_thematic_llm_configured():
        empty["decision"] = "AI_NOT_CONFIGURED"
        return empty

    call_slot = reserve_thematic_review_ai_call_slot()
    if not call_slot.get("allowed"):
        empty["decision"] = f"AI_SKIPPED_MAX_CALLS_PER_RUN_{call_slot.get('calls_used')}_OF_{call_slot.get('max_calls')}"
        return empty

    prompt = build_thematic_review_ai_prompt(
        carrera=carrera,
        current_result=current_result,
        title_value=title_value,
        abstract_value=abstract_value,
        author_keywords_value=author_keywords_value,
        index_keywords_value=index_keywords_value,
        source_title_value=source_title_value,
        review_reasons=review_reasons,
    )
    raw = call_openai_responses_json(
        prompt,
        throttle_seconds=get_thematic_review_ai_call_delay_seconds(),
        call_label="thematic_review_ai",
    )
    if not raw:
        empty["decision"] = "AI_FAILED_OR_NON_JSON"
        return empty

    decision = str(raw.get("decision") or "").strip().upper()
    confidence = parse_float_or_none(raw.get("confidence"))
    result = {"decision": decision or "INVALID_DECISION", "confidence": confidence, "raw": raw, "rationale": raw.get("rationale")}

    if decision not in {"KEEP", "CORRECT", "REVIEW"}:
        result["decision"] = "INVALID_DECISION"
        return result

    return result


def apply_thematic_review_ai_result(carrera: str, merged: dict, ai_result: dict) -> tuple[dict, str | None]:
    if not ai_result or ai_result.get("decision") != "CORRECT":
        return merged, None

    confidence = ai_result.get("confidence")
    if confidence is None or float(confidence) < get_thematic_review_ai_min_confidence():
        return merged, None

    raw = ai_result.get("raw") or {}
    area_carrera = coerce_choice(raw.get("area_carrera_raw"), get_allowed_career_areas(carrera))
    linea_carrera = coerce_choice(raw.get("linea_carrera_raw"), get_allowed_career_lines(carrera))
    if linea_carrera and not area_carrera:
        area_carrera = coerce_area_carrera_from_linea(carrera, linea_carrera)

    category_tematica = coerce_choice(raw.get("category_tematica_raw"), get_allowed_idic_categories())
    area_idic = coerce_choice(raw.get("area_idic_raw"), get_allowed_idic_areas(category_tematica))
    if not area_idic:
        area_idic = coerce_choice(raw.get("area_idic_raw"), get_allowed_idic_areas())
    linea_idic = coerce_choice(raw.get("linea_idic_raw"), get_allowed_idic_lines(category_tematica, area_idic))
    if not linea_idic:
        linea_idic = coerce_choice(raw.get("linea_idic_raw"), get_allowed_idic_lines())
    if linea_idic and not area_idic:
        area_idic = coerce_area_idic_from_linea(linea_idic)
    if area_idic and not category_tematica:
        category_tematica = coerce_category_tematica_from_area(area_idic)

    corrected = dict(merged)
    changed = False
    if is_valid_career_area_line(carrera, area_carrera, linea_carrera):
        corrected["area_carrera_raw"] = area_carrera
        corrected["linea_carrera_raw"] = linea_carrera
        changed = True
    if is_valid_idic_triplet(category_tematica, area_idic, linea_idic):
        corrected["category_tematica_raw"] = category_tematica
        corrected["area_idic_raw"] = area_idic
        corrected["linea_idic_raw"] = linea_idic
        changed = True

    if changed:
        if ai_result.get("rationale"):
            set_first_justification(corrected, str(ai_result.get("rationale")))
        return corrected, "thematic_review_ai_corrected"

    return merged, None


def build_ai_taxonomy_prompt(
    carrera: str,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> str:
    """Prompt único para clasificar taxonomía completa con catálogo cerrado."""
    payload = {
        "task": "classify_ulima_engineering_publication_taxonomy",
        "allowed_decisions": ["ACCEPT", "REVIEW"],
        "fixed_career": carrera,
        "career_catalog_for_fixed_career": CAREER_AREA_LINE_CATALOG.get(carrera, {}),
        "idic_catalog": IDIC_CATEGORY_AREA_LINE_CATALOG,
        "article": {
            "title": clip_text(title_value, 2200),
            "abstract_scopus": clip_text(abstract_value, 9000),
            "author_keywords": clip_text(author_keywords_value, 2500),
            "index_keywords": clip_text(index_keywords_value, 2500),
            "source_title": clip_text(source_title_value, 1200),
        },
        "output_schema": {
            "decision": "ACCEPT|REVIEW",
            "area_carrera_raw": "string from career_catalog_for_fixed_career|null",
            "linea_carrera_raw": "string from career_catalog_for_fixed_career|null",
            "category_tematica_raw": "string from idic_catalog|null",
            "area_idic_raw": "string from idic_catalog|null",
            "linea_idic_raw": "string from idic_catalog|null",
            "confidence": "number 0..1",
            "rationale": "short Spanish explanation, max 55 words",
            "evidence": ["short strings from title/abstract/keywords"]
        },
        "rules": [
            "La carrera fija ya fue resuelta por otro proceso. No la cambies ni la discutas.",
            "Clasifica obligatoriamente dos dimensiones: (1) área/línea de carrera y (2) categoría/área/línea IDIC.",
            "Usa el abstract_scopus como evidencia principal. Usa título y keywords como apoyo. Usa source_title solo como evidencia débil.",
            "Devuelve ACCEPT solo si puedes elegir valores exactos del catálogo cerrado con evidencia suficiente.",
            "Devuelve REVIEW si la clasificación no es defendible, si el artículo es demasiado ambiguo o si no hay una línea claramente dominante.",
            "No inventes categorías, áreas ni líneas. Copia los valores exactamente como aparecen en los catálogos.",
            "La línea de carrera debe pertenecer al área elegida y a la carrera fija.",
            "La línea IDIC debe pertenecer al área IDIC elegida y a la categoría temática elegida.",
            "No uses Realidad virtual y aumentada si el artículo no trata explícitamente AR, VR, mixed reality o entornos inmersivos.",
            "No uses Visión computacional si el artículo no trata explícitamente imágenes, video, detección, visión, LiDAR, point cloud, fotogrametría o procesamiento visual.",
            "Para remediación ambiental, residuos, contaminación, agua, suelo, adsorción, petróleo o microorganismos, favorece Desarrollo sostenible y medioambiente cuando corresponda.",
            "Para Shadow IT, governance, knowledge management, telecommuting o adoption de sistemas de información, no uses Visión computacional ni AR/VR salvo evidencia explícita.",
            "Para Lean, 5S, TPM, Kanban, inventarios, almacenes, logística, supply chain o mejora de procesos, clasifica la carrera e IDIC según operaciones/innovación, no como tecnología visual.",
            "Devuelve SOLO JSON válido, sin markdown ni texto adicional."
        ],
    }
    return (
        "Eres un clasificador académico senior para producción científica de la Facultad de Ingeniería de la Universidad de Lima. "
        "Debes clasificar taxonomía con criterio conservador, usando únicamente catálogos cerrados.\n\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )


def build_ai_taxonomy_review_result(reason: str, raw: dict | None = None, confidence=None, rationale: str | None = None) -> dict:
    result = build_thematic_empty_result()
    result["classification_mode"] = "ai_taxonomy_review"
    result["confidence"] = confidence
    result["justification"] = rationale
    result["thematic_review_rejected"] = should_ai_taxonomy_review_reject()
    result["thematic_review_rejection_reason"] = (
        "REVIEW_AI_TAXONOMY_CLASSIFICATION: AI taxonomy classifier did not produce a safe accepted taxonomy. "
        f"reason={reason}; confidence={'NULL' if confidence is None else confidence}; "
        f"rationale={(rationale or '')[:500]}"
    )
    if raw is not None:
        result["ai_taxonomy_raw"] = raw
    return result


def validate_ai_taxonomy_output(carrera: str, raw_output: dict | None) -> dict:
    if not raw_output:
        return build_ai_taxonomy_review_result("AI_FAILED_OR_NON_JSON")

    decision = str(raw_output.get("decision") or "").strip().upper()
    confidence = parse_float_or_none(raw_output.get("confidence"))
    rationale = str(raw_output.get("rationale") or "").strip() or None

    if decision != "ACCEPT":
        return build_ai_taxonomy_review_result(
            reason=f"AI_{decision or 'INVALID_DECISION'}",
            raw=raw_output,
            confidence=confidence,
            rationale=rationale,
        )

    if confidence is None or confidence < get_ai_taxonomy_min_confidence():
        return build_ai_taxonomy_review_result(
            reason=f"AI_LOW_CONFIDENCE_BELOW_THRESHOLD_{get_ai_taxonomy_min_confidence():.2f}",
            raw=raw_output,
            confidence=confidence,
            rationale=rationale,
        )

    result = build_thematic_empty_result()
    result["confidence"] = confidence
    result["justification"] = rationale
    result["classification_mode"] = "ai_taxonomy_classifier"

    area_carrera = coerce_choice(raw_output.get("area_carrera_raw"), get_allowed_career_areas(carrera))
    linea_carrera = coerce_choice(raw_output.get("linea_carrera_raw"), get_allowed_career_lines(carrera))
    if linea_carrera and not area_carrera:
        area_carrera = coerce_area_carrera_from_linea(carrera, linea_carrera)

    category_tematica = coerce_choice(raw_output.get("category_tematica_raw"), get_allowed_idic_categories())
    area_idic = coerce_choice(raw_output.get("area_idic_raw"), get_allowed_idic_areas(category_tematica))
    if not area_idic:
        area_idic = coerce_choice(raw_output.get("area_idic_raw"), get_allowed_idic_areas())
    linea_idic = coerce_choice(raw_output.get("linea_idic_raw"), get_allowed_idic_lines(category_tematica, area_idic))
    if not linea_idic:
        linea_idic = coerce_choice(raw_output.get("linea_idic_raw"), get_allowed_idic_lines())
    if linea_idic and not area_idic:
        area_idic = coerce_area_idic_from_linea(linea_idic)
    if area_idic and not category_tematica:
        category_tematica = coerce_category_tematica_from_area(area_idic)

    if not is_valid_career_area_line(carrera, area_carrera, linea_carrera):
        return build_ai_taxonomy_review_result(
            reason="AI_INVALID_CAREER_AREA_LINE",
            raw=raw_output,
            confidence=confidence,
            rationale=rationale,
        )

    if not is_valid_idic_triplet(category_tematica, area_idic, linea_idic):
        return build_ai_taxonomy_review_result(
            reason="AI_INVALID_IDIC_TRIPLET",
            raw=raw_output,
            confidence=confidence,
            rationale=rationale,
        )

    result["area_carrera_raw"] = area_carrera
    result["linea_carrera_raw"] = linea_carrera
    result["category_tematica_raw"] = category_tematica
    result["area_idic_raw"] = area_idic
    result["linea_idic_raw"] = linea_idic
    return result


def classify_taxonomy_with_ai(
    carrera: str,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
) -> dict:
    if not is_thematic_llm_configured():
        return build_ai_taxonomy_review_result("AI_NOT_CONFIGURED")

    call_slot = reserve_ai_taxonomy_call_slot()
    if not call_slot.get("allowed"):
        return build_ai_taxonomy_review_result(
            f"AI_SKIPPED_MAX_CALLS_PER_RUN_{call_slot.get('calls_used')}_OF_{call_slot.get('max_calls')}"
        )

    prompt = build_ai_taxonomy_prompt(
        carrera=carrera,
        title_value=title_value,
        abstract_value=abstract_value,
        author_keywords_value=author_keywords_value,
        index_keywords_value=index_keywords_value,
        source_title_value=source_title_value,
    )
    raw_output = call_openai_responses_json(
        prompt,
        throttle_seconds=get_ai_taxonomy_call_delay_seconds(),
        call_label="ai_taxonomy_classifier",
    )
    return validate_ai_taxonomy_output(carrera, raw_output)


def classify_thematic_fields(
    carrera: str | None,
    title_value: str | None,
    abstract_value: str | None,
    author_keywords_value: str | None,
    index_keywords_value: str | None,
    source_title_value: str | None,
    use_llm: bool | None = None,
    use_thematic_review_ai: bool | None = None,
    use_ai_taxonomy: bool | None = None,
) -> dict:
    """
    Clasificación temática.

    Run88:
    - Si AI_TAXONOMY_CLASSIFIER_ENABLED / ai_taxonomy=true está activo, la IA
      clasifica TODOS los registros válidos con catálogo cerrado.
    - El clasificador por reglas queda como fallback solo cuando ai_taxonomy está apagado.
    """
    merged = build_thematic_empty_result()

    if not carrera or carrera not in CAREER_AREA_LINE_CATALOG:
        return merged

    if use_llm is None:
        use_llm = resolve_llm_enabled()
    if use_thematic_review_ai is None:
        use_thematic_review_ai = resolve_thematic_review_ai_enabled()
    if use_ai_taxonomy is None:
        use_ai_taxonomy = resolve_ai_taxonomy_classifier_enabled()

    if use_ai_taxonomy:
        return classify_taxonomy_with_ai(
            carrera=carrera,
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
        )

    career_fields = ["area_carrera_raw", "linea_carrera_raw"]
    idic_fields = ["category_tematica_raw", "area_idic_raw", "linea_idic_raw"]

    # -------------------------
    # A. CARRERA - hints primero
    # -------------------------
    career_hint = classify_career_dimensions_by_hints(
        carrera=carrera,
        title_value=title_value,
        abstract_value=abstract_value,
        author_keywords_value=author_keywords_value,
        index_keywords_value=index_keywords_value,
        source_title_value=source_title_value,
    )
    merged = merge_non_null_fields(merged, career_hint, career_fields)
    if any(career_hint.get(field) for field in career_fields):
        append_classification_source(merged, "career_hints_strict")

    if any(not merged.get(field) for field in career_fields):
        career_hint_approx = classify_career_dimensions_by_hints_approx(
            carrera=carrera,
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
        )
        merged = merge_non_null_fields(merged, career_hint_approx, career_fields)
        if any(career_hint_approx.get(field) for field in career_fields):
            append_classification_source(merged, "career_hints_approx")

    # LLM solo como fallback opcional, no como motor principal.
    if use_llm and any(not merged.get(field) for field in career_fields):
        career_strict = classify_career_with_llm(
            carrera=carrera,
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
            mode="strict",
        )
        merged = merge_non_null_fields(merged, career_strict, career_fields)
        if any(career_strict.get(field) for field in career_fields):
            append_classification_source(merged, "career_llm_strict")
            if career_strict.get("confidence") is not None and merged.get("confidence") is None:
                merged["confidence"] = career_strict.get("confidence")
            set_first_justification(merged, career_strict.get("justification"))

    if use_llm and any(not merged.get(field) for field in career_fields):
        career_approx = classify_career_with_llm(
            carrera=carrera,
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
            mode="approx",
        )
        merged = merge_non_null_fields(merged, career_approx, career_fields)
        if any(career_approx.get(field) for field in career_fields):
            append_classification_source(merged, "career_llm_approx")
            if career_approx.get("confidence") is not None and merged.get("confidence") is None:
                merged["confidence"] = career_approx.get("confidence")
            set_first_justification(merged, career_approx.get("justification"))

    # -------------------------
    # B. IDIC - hints primero
    # -------------------------
    idic_hint = classify_idic_dimensions_by_hints(
        title_value=title_value,
        abstract_value=abstract_value,
        author_keywords_value=author_keywords_value,
        index_keywords_value=index_keywords_value,
        source_title_value=source_title_value,
    )
    merged = merge_non_null_fields(merged, idic_hint, idic_fields)
    if any(idic_hint.get(field) for field in idic_fields):
        append_classification_source(merged, "idic_hints_strict")

    if any(not merged.get(field) for field in idic_fields):
        idic_hint_approx = classify_idic_dimensions_by_hints_approx(
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
        )
        merged = merge_non_null_fields(merged, idic_hint_approx, idic_fields)
        if any(idic_hint_approx.get(field) for field in idic_fields):
            append_classification_source(merged, "idic_hints_approx")

    # LLM solo como fallback opcional, no como motor principal.
    if use_llm and any(not merged.get(field) for field in idic_fields):
        idic_strict = classify_idic_with_llm(
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
            mode="strict",
        )
        merged = merge_non_null_fields(merged, idic_strict, idic_fields)
        if any(idic_strict.get(field) for field in idic_fields):
            append_classification_source(merged, "idic_llm_strict")
            if idic_strict.get("confidence") is not None and merged.get("confidence") is None:
                merged["confidence"] = idic_strict.get("confidence")
            set_first_justification(merged, idic_strict.get("justification"))

    if use_llm and any(not merged.get(field) for field in idic_fields):
        idic_approx = classify_idic_with_llm(
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
            mode="approx",
        )
        merged = merge_non_null_fields(merged, idic_approx, idic_fields)
        if any(idic_approx.get(field) for field in idic_fields):
            append_classification_source(merged, "idic_llm_approx")
            if idic_approx.get("confidence") is not None and merged.get("confidence") is None:
                merged["confidence"] = idic_approx.get("confidence")
            set_first_justification(merged, idic_approx.get("justification"))

    text_fields = build_thematic_text_fields(
        title_value,
        abstract_value,
        author_keywords_value,
        index_keywords_value,
        source_title_value,
    )

    (
        merged["category_tematica_raw"],
        merged["area_idic_raw"],
        merged["linea_idic_raw"],
        idic_guardrail_source,
    ) = apply_final_idic_guardrails(
        merged.get("category_tematica_raw"),
        merged.get("area_idic_raw"),
        merged.get("linea_idic_raw"),
        text_fields,
    )

    if idic_guardrail_source:
        append_classification_source(merged, idic_guardrail_source)

    (
        merged["area_carrera_raw"],
        merged["linea_carrera_raw"],
        career_guardrail_source,
    ) = apply_final_career_guardrails(
        carrera,
        merged.get("area_carrera_raw"),
        merged.get("linea_carrera_raw"),
        text_fields,
    )

    if career_guardrail_source:
        append_classification_source(merged, career_guardrail_source)

    # Validación final
    if not is_valid_career_area_line(
        carrera,
        merged.get("area_carrera_raw"),
        merged.get("linea_carrera_raw"),
    ):
        merged["area_carrera_raw"] = None
        merged["linea_carrera_raw"] = None

    if not is_valid_idic_triplet(
        merged.get("category_tematica_raw"),
        merged.get("area_idic_raw"),
        merged.get("linea_idic_raw"),
    ):
        merged["category_tematica_raw"] = None
        merged["area_idic_raw"] = None
        merged["linea_idic_raw"] = None

    # Fallback final obligatorio: no dejar NULLs temáticos
    if any(not merged.get(field) for field in career_fields):
        career_forced = classify_career_dimensions_force_best(
            carrera=carrera,
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
        )
        merged = merge_non_null_fields(merged, career_forced, career_fields)
        if any(career_forced.get(field) for field in career_fields):
            append_classification_source(merged, "career_force_best")
            set_first_justification(merged, "Fallback final obligatorio de carrera para evitar NULLs.")

    if any(not merged.get(field) for field in idic_fields):
        idic_forced = classify_idic_dimensions_force_best(
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
        )
        merged = merge_non_null_fields(merged, idic_forced, idic_fields)
        if any(idic_forced.get(field) for field in idic_fields):
            append_classification_source(merged, "idic_force_best")
            set_first_justification(merged, "Fallback final obligatorio de IDIC para evitar NULLs.")

    # Safety net: por regla del proyecto, ningún registro aceptado debe salir con IDIC en NULL.
    if any(not merged.get(field) for field in idic_fields):
        fallback_category, fallback_area, fallback_line, fallback_source = choose_general_idic_fallback(text_fields)
        merged["category_tematica_raw"] = fallback_category
        merged["area_idic_raw"] = fallback_area
        merged["linea_idic_raw"] = fallback_line
        append_classification_source(merged, fallback_source)
        set_first_justification(merged, "Fallback institucional obligatorio de IDIC para evitar NULLs.")

    # Run80 deterministic overrides after all fallbacks.
    merged, run80_sources = apply_run80_domain_specific_overrides(carrera, merged, text_fields)
    for source in run80_sources:
        append_classification_source(merged, source)

    # Run80 selective AI review: only suspicious accepted classifications.
    review_reasons = detect_thematic_review_reasons(carrera, merged, text_fields)
    if use_thematic_review_ai and review_reasons:
        ai_review = review_thematic_classification_with_ai(
            carrera=carrera,
            current_result=merged,
            title_value=title_value,
            abstract_value=abstract_value,
            author_keywords_value=author_keywords_value,
            index_keywords_value=index_keywords_value,
            source_title_value=source_title_value,
            review_reasons=review_reasons,
        )
        append_classification_source(merged, f"thematic_review_ai_{ai_review.get('decision', 'NO_RESULT')}")
        corrected, ai_source = apply_thematic_review_ai_result(carrera, merged, ai_review)
        merged = corrected
        if ai_source:
            append_classification_source(merged, ai_source)
        if ai_review.get("decision") == "REVIEW" and should_send_thematic_review_to_rejected():
            merged["thematic_review_rejected"] = True
            merged["thematic_review_rejection_reason"] = (
                "REVIEW_THEMATIC_CLASSIFICATION_UNSAFE: selective AI requested manual review. "
                f"reasons={';'.join(review_reasons)}; confidence={ai_review.get('confidence')}; "
                f"rationale={(ai_review.get('rationale') or '')[:500]}"
            )

    return merged


# =========================
# STORAGE HELPERS
# =========================
def validate_containers() -> dict:
    blob_service = get_blob_service()

    raw_container = get_env("RAW_CONTAINER")
    processed_container = get_env("PROCESSED_CONTAINER")
    logs_container = get_env("LOGS_CONTAINER")

    return {
        "raw_exists": blob_service.get_container_client(raw_container).exists(),
        "processed_exists": blob_service.get_container_client(processed_container).exists(),
        "logs_exists": blob_service.get_container_client(logs_container).exists(),
    }


def get_latest_csv_blob_name(container_name: str) -> str:
    blob_service = get_blob_service()
    container_client = blob_service.get_container_client(container_name)

    blobs = [b for b in container_client.list_blobs() if b.name.lower().endswith(".csv")]
    if not blobs:
        raise RuntimeError(f"No CSV files found in container: {container_name}")

    latest = max(blobs, key=lambda b: b.last_modified)
    return latest.name


def get_latest_excel_blob_name(container_name: str) -> str:
    blob_service = get_blob_service()
    container_client = blob_service.get_container_client(container_name)

    blobs = [
        b for b in container_client.list_blobs()
        if b.name.lower().endswith(".xlsx") or b.name.lower().endswith(".xlsm")
    ]
    if not blobs:
        raise RuntimeError(f"No Excel files found in container: {container_name}")

    latest = max(blobs, key=lambda b: b.last_modified)
    return latest.name


def download_blob_text(container_name: str, blob_name: str) -> str:
    blob_service = get_blob_service()
    blob_client = blob_service.get_blob_client(container=container_name, blob=blob_name)
    data = blob_client.download_blob().readall()
    return data.decode("utf-8-sig")


def download_blob_bytes(container_name: str, blob_name: str) -> bytes:
    blob_service = get_blob_service()
    blob_client = blob_service.get_blob_client(container=container_name, blob=blob_name)
    return blob_client.download_blob().readall()


def upload_text_blob(container_name: str, blob_name: str, content: str) -> None:
    blob_service = get_blob_service()
    blob_client = blob_service.get_blob_client(container=container_name, blob=blob_name)
    blob_client.upload_blob(content.encode("utf-8"), overwrite=True)


def upload_bytes_blob(container_name: str, blob_name: str, content: bytes) -> None:
    blob_service = get_blob_service()
    blob_client = blob_service.get_blob_client(container=container_name, blob=blob_name)
    blob_client.upload_blob(content, overwrite=True)


# =========================
# SQL HELPERS
# =========================
def create_pipeline_run(
    trigger_type: str,
    source_name: str,
    source_file_name: str,
    source_file_path: str,
) -> int:
    from mssql_python import connect

    pipeline_name = os.environ.get("PIPELINE_NAME", "scopus_monthly_pipeline")
    sql_conn_str = get_sql_connection_string()

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO meta.pipeline_runs (
                pipeline_name,
                source_name,
                trigger_type,
                source_file_name,
                source_file_path,
                status,
                records_read,
                records_inserted,
                records_updated,
                records_rejected
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                pipeline_name,
                source_name,
                trigger_type,
                source_file_name,
                source_file_path,
                "STARTED",
                0,
                0,
                0,
                0,
            ),
        )
        conn.commit()

        cursor.execute("SELECT TOP 1 run_id FROM meta.pipeline_runs ORDER BY run_id DESC")
        row = cursor.fetchone()
        cursor.close()

    try:
        return int(row.run_id)
    except AttributeError:
        return int(row[0])


def update_pipeline_run(
    run_id: int,
    status: str,
    records_read: int = 0,
    records_inserted: int = 0,
    records_updated: int = 0,
    records_rejected: int = 0,
    error_message: str | None = None,
) -> None:
    from mssql_python import connect

    sql_conn_str = get_sql_connection_string()

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            UPDATE meta.pipeline_runs
            SET
                ended_at_utc = SYSUTCDATETIME(),
                status = ?,
                records_read = ?,
                records_inserted = ?,
                records_updated = ?,
                records_rejected = ?,
                error_message = ?
            WHERE run_id = ?
            """,
            (
                status,
                records_read,
                records_inserted,
                records_updated,
                records_rejected,
                error_message,
                run_id,
            ),
        )
        conn.commit()
        cursor.close()


def insert_pipeline_run_smoke(trigger_type: str, status: str, source_name: str = "FUNCTION_SMOKE") -> int:
    return create_pipeline_run(
        trigger_type=trigger_type,
        source_name=source_name,
        source_file_name="N/A",
        source_file_path="N/A",
    )


def execute_upsert_from_staging(run_id: int) -> None:
    from mssql_python import connect

    sql_conn_str = get_sql_connection_string()

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute("EXEC dbo.usp_upsert_publications_from_stg @run_id = ?", (run_id,))
        conn.commit()
        cursor.close()



def reactivate_curated_publications_from_valid_stg(run_id: int) -> int:
    """
    Safety patch post-upsert.

    Si una publicación reaparece como válida en un run nuevo, debe quedar activa.
    Esto corrige casos donde un registro previamente desactivado se actualiza pero
    conserva is_active = 0 por la lógica del stored procedure.
    """
    from mssql_python import connect

    sql_conn_str = get_sql_connection_string()

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            UPDATE c
            SET
                c.is_active = 1,
                c.updated_at_utc = SYSUTCDATETIME()
            FROM curated.publications c
            INNER JOIN stg.scopus_raw_load s
                ON s.run_id = ?
               AND s.is_valid_for_curated = 1
               AND (
                    (s.eid IS NOT NULL AND c.eid = s.eid)
                 OR (s.eid IS NULL AND s.doi_link_raw IS NOT NULL AND c.doi_link = s.doi_link_raw)
                 OR (
                        s.eid IS NULL
                    AND s.doi_link_raw IS NULL
                    AND s.publication_title_raw IS NOT NULL
                    AND LOWER(LTRIM(RTRIM(c.publication_title))) = LOWER(LTRIM(RTRIM(s.publication_title_raw)))
                    AND c.publication_year = TRY_CONVERT(INT, s.publication_year_raw)
                    )
               )
            WHERE ISNULL(c.is_active, 0) = 0
            """,
            (run_id,),
        )
        affected = cursor.rowcount if cursor.rowcount is not None else 0
        conn.commit()
        cursor.close()

    return int(affected) if affected and affected > 0 else 0


def deactivate_curated_publications_from_rejected_stg(run_id: int) -> int:
    """
    Si un artículo fue aceptado en un run anterior pero en el run actual queda
    rechazado/REVIEW, se desactiva en curated para que no siga apareciendo en
    Power BI por arrastre histórico.
    """
    from mssql_python import connect

    sql_conn_str = get_sql_connection_string()

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            UPDATE c
            SET
                c.is_active = 0,
                c.updated_at_utc = SYSUTCDATETIME()
            FROM curated.publications c
            INNER JOIN stg.scopus_raw_load s
                ON s.run_id = ?
               AND s.is_valid_for_curated = 0
               AND (
                    (s.eid IS NOT NULL AND c.eid = s.eid)
                 OR (s.eid IS NULL AND s.doi_link_raw IS NOT NULL AND c.doi_link = s.doi_link_raw)
                 OR (
                        s.eid IS NULL
                    AND s.doi_link_raw IS NULL
                    AND s.publication_title_raw IS NOT NULL
                    AND LOWER(LTRIM(RTRIM(c.publication_title))) = LOWER(LTRIM(RTRIM(s.publication_title_raw)))
                    AND c.publication_year = TRY_CONVERT(INT, s.publication_year_raw)
                    )
               )
            WHERE ISNULL(c.is_active, 0) = 1
            """,
            (run_id,),
        )
        affected = cursor.rowcount if cursor.rowcount is not None else 0
        conn.commit()
        cursor.close()

    return int(affected) if affected and affected > 0 else 0


# =========================
# DOCENTES INGESTION
# =========================
DOCENTES_SHEET_CAREERS = {
    "Civil": "Ingeniería Civil",
    "Industrial": "Ingeniería Industrial",
    "Sistemas": "Ingeniería de Sistemas",
}


def normalize_docente_name(value: str | None) -> str:
    if not value:
        return ""
    norm = normalize_generic_text(value)
    norm = norm.replace("/", " ")
    norm = re.sub(r"\s+", " ", norm).strip()
    return norm


def to_title_or_none(value: str | None) -> str | None:
    if not value:
        return None
    value = str(value).strip()
    if not value:
        return None
    return value.title()


def build_initials(nombres: str | None) -> str | None:
    if not nombres:
        return None
    parts = [p.strip() for p in nombres.split() if p.strip()]
    if not parts:
        return None
    return "".join(part[0].upper() for part in parts if part)


def parse_docente_principal_raw(docente_raw: str) -> dict:
    raw = str(docente_raw).strip()
    parts = [p.strip() for p in raw.split("/") if p.strip()]

    apellido_1 = None
    apellido_2 = None
    nombres = None

    if len(parts) >= 3:
        apellido_1 = to_title_or_none(parts[0])
        apellido_2 = to_title_or_none(parts[1])
        nombres = to_title_or_none(" ".join(parts[2:]))
    elif len(parts) == 2:
        apellido_1 = to_title_or_none(parts[0])
        nombres = to_title_or_none(parts[1])
    else:
        nombres = to_title_or_none(raw)

    full_for_normalization = " ".join([x for x in [apellido_1, apellido_2, nombres] if x])

    return {
        "nombre_original": raw,
        "nombre_normalizado": normalize_docente_name(full_for_normalization),
        "apellido_1": apellido_1,
        "apellido_2": apellido_2,
        "nombres": nombres,
        "iniciales": build_initials(nombres),
    }


def find_docentes_header_indices(worksheet) -> tuple[int, int | None, int]:
    max_scan_rows = min(10, worksheet.max_row)

    for row_idx in range(1, max_scan_rows + 1):
        row_values = list(
            worksheet.iter_rows(
                min_row=row_idx,
                max_row=row_idx,
                values_only=True,
            )
        )[0]

        normalized = [normalize_generic_text(v) for v in row_values]

        docente_idx = None
        codigo_idx = None

        for idx, value in enumerate(normalized):
            if "docente principal" in value:
                docente_idx = idx
            if "codigo" in value and "docente" in value:
                codigo_idx = idx
            elif "codigo docente" in value:
                codigo_idx = idx
            elif "cod docente" in value:
                codigo_idx = idx

        if docente_idx is not None:
            return row_idx, codigo_idx, docente_idx

    raise RuntimeError(
        f"Could not find header row with 'Docente Principal' in sheet '{worksheet.title}'"
    )


def extract_docentes_from_workbook_bytes(
    workbook_bytes: bytes,
    source_file_name: str,
    periodo_academico: str,
) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=io.BytesIO(workbook_bytes), data_only=True)
    extracted_rows: list[dict] = []

    for sheet_name, carrera in DOCENTES_SHEET_CAREERS.items():
        if sheet_name not in workbook.sheetnames:
            logging.warning("Sheet '%s' not found in workbook. Skipping.", sheet_name)
            continue

        ws = workbook[sheet_name]
        header_row_idx, codigo_idx, docente_idx = find_docentes_header_indices(ws)

        for excel_row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_values = list(
                ws.iter_rows(
                    min_row=excel_row_idx,
                    max_row=excel_row_idx,
                    values_only=True,
                )
            )[0]

            docente_raw = None
            codigo_raw = None

            if docente_idx is not None and docente_idx < len(row_values):
                docente_raw = row_values[docente_idx]
            if codigo_idx is not None and codigo_idx < len(row_values):
                codigo_raw = row_values[codigo_idx]

            if docente_raw is None or str(docente_raw).strip() == "":
                continue

            parsed = parse_docente_principal_raw(str(docente_raw))

            extracted_rows.append(
                {
                    "periodo_academico": periodo_academico,
                    "source_file_name": source_file_name,
                    "source_sheet_name": sheet_name,
                    "source_row_number": excel_row_idx,
                    "carrera_fuente": carrera,
                    "codigo_docente_raw": str(codigo_raw).strip() if codigo_raw is not None else None,
                    "docente_principal_raw": parsed["nombre_original"],
                    "docente_principal_normalizado": parsed["nombre_normalizado"],
                    "apellido_1": parsed["apellido_1"],
                    "apellido_2": parsed["apellido_2"],
                    "nombres": parsed["nombres"],
                    "iniciales": parsed["iniciales"],
                }
            )

    if not extracted_rows:
        raise RuntimeError("No docentes extracted from workbook.")

    return extracted_rows


def insert_docentes_raw_rows(run_id: int, rows: list[dict]) -> int:
    from mssql_python import connect

    sql_conn_str = get_sql_connection_string()

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()

        for row in rows:
            cursor.execute(
                """
                INSERT INTO stg.docentes_ulima_raw (
                    run_id,
                    periodo_academico,
                    source_file_name,
                    source_sheet_name,
                    source_row_number,
                    carrera_fuente,
                    codigo_docente_raw,
                    docente_principal_raw,
                    docente_principal_normalizado
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run_id,
                    row["periodo_academico"],
                    row["source_file_name"],
                    row["source_sheet_name"],
                    row["source_row_number"],
                    row["carrera_fuente"],
                    row["codigo_docente_raw"],
                    row["docente_principal_raw"],
                    row["docente_principal_normalizado"],
                ),
            )

        conn.commit()
        cursor.close()

    return len(rows)


def rebuild_docentes_reference(run_id: int, rows: list[dict], periodo_academico: str) -> int:
    from mssql_python import connect

    unique_map: dict[tuple[str, str, str], dict] = {}

    for row in rows:
        key = (
            row["periodo_academico"],
            row["carrera_fuente"],
            row["docente_principal_normalizado"],
        )

        if key not in unique_map:
            unique_map[key] = {
                "periodo_academico": row["periodo_academico"],
                "carrera": row["carrera_fuente"],
                "codigo_docente": row["codigo_docente_raw"],
                "nombre_original": row["docente_principal_raw"],
                "nombre_normalizado": row["docente_principal_normalizado"],
                "apellido_1": row["apellido_1"],
                "apellido_2": row["apellido_2"],
                "nombres": row["nombres"],
                "iniciales": row["iniciales"],
                "activo": 1,
                "source_run_id": run_id,
            }
        else:
            if not unique_map[key]["codigo_docente"] and row["codigo_docente_raw"]:
                unique_map[key]["codigo_docente"] = row["codigo_docente_raw"]

    deduped_rows = list(unique_map.values())
    sql_conn_str = get_sql_connection_string()

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()

        cursor.execute(
            "DELETE FROM ref.docentes_ulima WHERE periodo_academico = ?",
            (periodo_academico,),
        )

        for row in deduped_rows:
            cursor.execute(
                """
                INSERT INTO ref.docentes_ulima (
                    periodo_academico,
                    carrera,
                    codigo_docente,
                    nombre_original,
                    nombre_normalizado,
                    apellido_1,
                    apellido_2,
                    nombres,
                    iniciales,
                    activo,
                    source_run_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["periodo_academico"],
                    row["carrera"],
                    row["codigo_docente"],
                    row["nombre_original"],
                    row["nombre_normalizado"],
                    row["apellido_1"],
                    row["apellido_2"],
                    row["nombres"],
                    row["iniciales"],
                    row["activo"],
                    row["source_run_id"],
                ),
            )

        conn.commit()
        cursor.close()

    return len(deduped_rows)


# =========================
# DOCENTES REF MATCHING / NORMALIZATION
# =========================
def build_docente_display_name(docente: dict) -> str:
    family = " ".join([x for x in [docente.get("apellido_1"), docente.get("apellido_2")] if x])
    names = docente.get("nombres")
    if family and names:
        return f"{family.title()}, {names.title()}"
    return (docente.get("nombre_original") or "").replace("/", " ").title()


def first_given_name(nombres: str | None) -> str:
    tokens = [t for t in normalize_generic_text(nombres).split() if t]
    return tokens[0] if tokens else ""


def initials_variants_from_tokens(tokens: list[str]) -> set[str]:
    if not tokens:
        return set()
    initials = "".join(token[0] for token in tokens if token)
    variants = {initials}
    if initials:
        variants.add(initials[0])
    return {v for v in variants if v}


def prepare_docente_reference_entry(docente: dict) -> dict:
    apellido_1_norm = normalize_generic_text(docente.get("apellido_1"))
    apellido_2_norm = normalize_generic_text(docente.get("apellido_2"))
    nombres_norm = normalize_generic_text(docente.get("nombres"))
    nombre_normalizado = normalize_person_name(docente.get("nombre_normalizado") or docente.get("nombre_original"))

    family_signature = " ".join([x for x in [apellido_1_norm, apellido_2_norm] if x]).strip()
    first_name = first_given_name(docente.get("nombres"))
    initials_full = normalize_generic_text(docente.get("iniciales"))

    strong_aliases: set[str] = set()
    weak_aliases: set[str] = set()

    if nombre_normalizado:
        strong_aliases.add(nombre_normalizado)

    initials_variants = {initials_full}
    if initials_full:
        initials_variants.add(initials_full[:1])

    if family_signature:
        for iv in initials_variants:
            if iv:
                strong_aliases.add(f"{family_signature} {iv}".strip())
        if first_name:
            strong_aliases.add(f"{family_signature} {first_name}".strip())

    if apellido_1_norm:
        for iv in initials_variants:
            if iv:
                weak_aliases.add(f"{apellido_1_norm} {iv}".strip())
        if first_name:
            weak_aliases.add(f"{apellido_1_norm} {first_name}".strip())

    docente_prepared = dict(docente)
    docente_prepared["apellido_1_norm"] = apellido_1_norm
    docente_prepared["apellido_2_norm"] = apellido_2_norm
    docente_prepared["nombres_norm"] = nombres_norm
    docente_prepared["family_signature"] = family_signature
    docente_prepared["first_name_norm"] = first_name
    docente_prepared["strong_aliases"] = strong_aliases
    docente_prepared["weak_aliases"] = weak_aliases

    return docente_prepared


def get_docentes_reference(periodo_academico: str) -> list[dict]:
    from mssql_python import connect

    sql_conn_str = get_sql_connection_string()
    docentes: list[dict] = []

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT
                docente_ref_id,
                periodo_academico,
                carrera,
                codigo_docente,
                nombre_original,
                nombre_normalizado,
                apellido_1,
                apellido_2,
                nombres,
                iniciales
            FROM ref.docentes_ulima
            WHERE periodo_academico = ?
              AND activo = 1
            """,
            (periodo_academico,),
        )

        rows = cursor.fetchall()
        cursor.close()

    for row in rows:
        docente = {
            "docente_ref_id": row_attr(row, "docente_ref_id", 0),
            "periodo_academico": row_attr(row, "periodo_academico", 1),
            "carrera": row_attr(row, "carrera", 2),
            "codigo_docente": row_attr(row, "codigo_docente", 3),
            "nombre_original": row_attr(row, "nombre_original", 4),
            "nombre_normalizado": row_attr(row, "nombre_normalizado", 5),
            "apellido_1": row_attr(row, "apellido_1", 6),
            "apellido_2": row_attr(row, "apellido_2", 7),
            "nombres": row_attr(row, "nombres", 8),
            "iniciales": row_attr(row, "iniciales", 9),
        }
        docentes.append(prepare_docente_reference_entry(docente))

    return docentes


def split_authors_with_affiliations_blocks(
    author_full_names_value: str | None,
    authors_with_affiliations_value: str | None,
) -> list[str]:
    text = str(authors_with_affiliations_value or "").strip()
    if not text:
        return []

    full_names = [clean_author_full_name(x) for x in split_semicolon_values(author_full_names_value)]
    if not full_names:
        return split_semicolon_values(authors_with_affiliations_value)

    positions: list[tuple[str, int]] = []
    search_start = 0

    for name in full_names:
        pos = text.find(name, search_start)
        if pos == -1:
            pos = text.find(name)
        if pos != -1:
            positions.append((name, pos))
            search_start = pos + len(name)

    if len(positions) < 1:
        return split_semicolon_values(authors_with_affiliations_value)

    blocks: list[str] = []
    for i, (_, start_pos) in enumerate(positions):
        end_pos = positions[i + 1][1] if i + 1 < len(positions) else len(text)
        block = text[start_pos:end_pos].strip(" ;")
        if block:
            blocks.append(block)

    return blocks


def extract_author_name_from_block(block: str | None) -> str | None:
    if not block:
        return None

    first_part = str(block).split(",", 1)[0].strip()
    if not first_part:
        return None

    return clean_author_full_name(first_part)


def get_preferred_author_name_for_block(
    idx: int,
    full_authors: list[str],
    short_authors: list[str],
    block: str | None,
) -> str | None:
    if idx < len(full_authors) and full_authors[idx]:
        return clean_author_full_name(full_authors[idx])

    if idx < len(short_authors) and short_authors[idx]:
        return clean_author_full_name(short_authors[idx])

    return extract_author_name_from_block(block)


def parse_scopus_author_name(scopus_author_name: str) -> dict:
    raw = clean_author_full_name(scopus_author_name)

    family_tokens: list[str] = []
    given_tokens: list[str] = []

    if "," in raw:
        family_part, given_part = [p.strip() for p in raw.split(",", 1)]
        family_tokens = [t for t in normalize_generic_text(family_part).split() if t]
        given_tokens = [t for t in normalize_generic_text(given_part.replace(".", " ")).split() if t]
    else:
        tokens = [t.strip() for t in raw.replace(".", " ").split() if t.strip()]
        normalized_tokens = [normalize_generic_text(t) for t in tokens]

        trailing_initials: list[str] = []
        idx = len(normalized_tokens) - 1
        while idx >= 0 and len(normalized_tokens[idx]) == 1:
            trailing_initials.insert(0, normalized_tokens[idx])
            idx -= 1

        if trailing_initials and idx >= 0:
            family_tokens = normalized_tokens[: idx + 1]
            given_tokens = trailing_initials
        elif len(normalized_tokens) >= 2:
            family_tokens = normalized_tokens[:-1]
            given_tokens = [normalized_tokens[-1]]
        else:
            family_tokens = normalized_tokens
            given_tokens = []

    apellido_1 = family_tokens[0] if family_tokens else None
    apellido_2 = " ".join(family_tokens[1:]) if len(family_tokens) > 1 else None
    initials = "".join(token[0] for token in given_tokens if token)
    normalized_full = normalize_person_name(raw)
    family_signature = " ".join(family_tokens).strip()
    first_name = given_tokens[0] if given_tokens else ""

    return {
        "raw": raw,
        "apellido_1": apellido_1,
        "apellido_2": apellido_2,
        "family_tokens": family_tokens,
        "family_signature": family_signature,
        "given_tokens": given_tokens,
        "initials": initials,
        "first_name": first_name,
        "normalized_full": normalized_full,
    }


def build_scopus_author_aliases(parsed: dict) -> dict:
    strong_aliases: set[str] = set()
    weak_aliases: set[str] = set()

    if parsed["normalized_full"]:
        strong_aliases.add(parsed["normalized_full"])

    initials_variants = initials_variants_from_tokens(parsed["given_tokens"])

    if parsed["family_signature"]:
        for iv in initials_variants:
            strong_aliases.add(f"{parsed['family_signature']} {iv}".strip())
        if parsed["first_name"]:
            strong_aliases.add(f"{parsed['family_signature']} {parsed['first_name']}".strip())

    if parsed["apellido_1"]:
        for iv in initials_variants:
            weak_aliases.add(f"{parsed['apellido_1']} {iv}".strip())
        if parsed["first_name"]:
            weak_aliases.add(f"{parsed['apellido_1']} {parsed['first_name']}".strip())

    return {
        "strong_aliases": {a for a in strong_aliases if a},
        "weak_aliases": {a for a in weak_aliases if a},
    }


def given_names_match(ref_nombres: str | None, scopus_given_tokens: list[str], scopus_initials: str) -> bool:
    ref_tokens = [t for t in normalize_generic_text(ref_nombres).split() if t]

    if not scopus_given_tokens and not scopus_initials:
        return True

    if scopus_initials and (not scopus_given_tokens or all(len(tok) == 1 for tok in scopus_given_tokens)):
        ref_initials = "".join(token[0] for token in ref_tokens if token)
        return ref_initials.startswith(scopus_initials)

    if scopus_given_tokens:
        if len(ref_tokens) < len(scopus_given_tokens):
            return False
        for idx, token in enumerate(scopus_given_tokens):
            if ref_tokens[idx] != token:
                return False
        return True

    return False


def filter_candidates_by_career_hint(candidates: list[dict], career_hint: str | None) -> list[dict]:
    if not career_hint:
        return candidates
    filtered = [c for c in candidates if c.get("carrera") == career_hint]
    return filtered if filtered else candidates


def unique_alias_match(
    scopus_aliases: set[str],
    docentes_ref: list[dict],
    alias_key: str,
    career_hint: str | None = None,
) -> dict:
    candidates = []
    for docente in docentes_ref:
        if scopus_aliases.intersection(docente.get(alias_key, set())):
            candidates.append(docente)

    candidates = filter_candidates_by_career_hint(candidates, career_hint)

    if len(candidates) == 1:
        return {"matched": True, "ambiguous": False, "docente": candidates[0]}

    if len(candidates) > 1:
        return {"matched": False, "ambiguous": True, "docente": None}

    return {"matched": False, "ambiguous": False, "docente": None}


def match_scopus_author_to_docente(
    scopus_author_name: str,
    docentes_ref: list[dict],
    career_hint: str | None = None,
) -> dict:
    parsed = parse_scopus_author_name(scopus_author_name)
    scopus_aliases = build_scopus_author_aliases(parsed)

    exact_matches = [d for d in docentes_ref if d["nombre_normalizado"] == parsed["normalized_full"]]
    exact_matches = filter_candidates_by_career_hint(exact_matches, career_hint)

    if len(exact_matches) == 1:
        return {
            "matched": True,
            "ambiguous": False,
            "match_method": "DOCENTES_REF_EXACT",
            "docente": exact_matches[0],
        }

    if len(exact_matches) > 1:
        return {
            "matched": False,
            "ambiguous": True,
            "match_method": "DOCENTES_REF_EXACT_AMBIGUOUS",
            "docente": None,
        }

    structured_matches = []
    for docente in docentes_ref:
        ref_ap1 = docente.get("apellido_1_norm")
        ref_ap2 = docente.get("apellido_2_norm")

        if parsed["apellido_1"] and ref_ap1 != parsed["apellido_1"]:
            continue
        if parsed["apellido_2"] and ref_ap2 != parsed["apellido_2"]:
            continue
        if not parsed["apellido_1"]:
            continue
        if not given_names_match(docente.get("nombres"), parsed["given_tokens"], parsed["initials"]):
            continue

        structured_matches.append(docente)

    structured_matches = filter_candidates_by_career_hint(structured_matches, career_hint)

    if len(structured_matches) == 1:
        return {
            "matched": True,
            "ambiguous": False,
            "match_method": "DOCENTES_REF_STRUCTURED",
            "docente": structured_matches[0],
        }

    if len(structured_matches) > 1:
        return {
            "matched": False,
            "ambiguous": True,
            "match_method": "DOCENTES_REF_STRUCTURED_AMBIGUOUS",
            "docente": None,
        }

    strong_match = unique_alias_match(
        scopus_aliases=scopus_aliases["strong_aliases"],
        docentes_ref=docentes_ref,
        alias_key="strong_aliases",
        career_hint=career_hint,
    )

    if strong_match["matched"]:
        return {
            "matched": True,
            "ambiguous": False,
            "match_method": "DOCENTES_REF_ALIAS_STRONG",
            "docente": strong_match["docente"],
        }

    if strong_match["ambiguous"]:
        return {
            "matched": False,
            "ambiguous": True,
            "match_method": "DOCENTES_REF_ALIAS_STRONG_AMBIGUOUS",
            "docente": None,
        }

    return {
        "matched": False,
        "ambiguous": False,
        "match_method": "DOCENTES_REF_NO_MATCH",
        "docente": None,
    }


# =========================
# AFFILIATION-FIRST + AUTHOR FALLBACK
# =========================
def get_publication_level_engineering_careers(affiliations_value: str | None) -> list[str]:
    publication_details = resolve_engineering_affiliation_details(affiliations_value)
    return filter_valid_engineering_careers(publication_details.get("careers"))



def get_block_engineering_careers(
    block_value: str | None,
    publication_level_careers: list[str],
) -> list[str]:
    block_details = resolve_engineering_affiliation_details(block_value)
    block_careers = filter_valid_engineering_careers(block_details.get("careers"))
    if block_careers:
        return block_careers

    if block_details.get("has_ulima_engineering_context") and len(publication_level_careers) == 1:
        return publication_level_careers

    return []



def enrich_ulima_fields_from_ref(
    authors_value: str | None,
    author_full_names_value: str | None,
    authors_with_affiliations_value: str | None,
    affiliations_value: str | None,
    docentes_ref: list[dict],
) -> dict:
    short_authors = split_semicolon_values(authors_value)
    full_authors = [clean_author_full_name(x) for x in split_semicolon_values(author_full_names_value)]
    author_blocks = split_authors_with_affiliations_blocks(
        author_full_names_value=author_full_names_value,
        authors_with_affiliations_value=authors_with_affiliations_value,
    )
    if not author_blocks and affiliations_value:
        author_blocks = split_semicolon_values(affiliations_value)
        if not author_blocks:
            author_blocks = [str(affiliations_value)]

    publication_details = resolve_engineering_affiliation_details(affiliations_value)
    publication_level_careers = filter_valid_engineering_careers(publication_details.get("careers"))

    ulima_authors_detected: list[str] = []
    publication_careers: list[str] = []
    docente_ref_careers: list[str] = []
    first_author_ulima = False
    matched_any = False
    affiliation_ulima_detected = False
    has_ulima_engineering_affiliation = False

    for idx, block in enumerate(author_blocks):
        block_details = resolve_engineering_affiliation_details(block)
        if not block_details.get("has_ulima_affiliation"):
            continue

        affiliation_ulima_detected = True
        if block_details.get("has_ulima_engineering_context"):
            has_ulima_engineering_affiliation = True

        preferred_author_name = get_preferred_author_name_for_block(
            idx=idx,
            full_authors=full_authors,
            short_authors=short_authors,
            block=block,
        )
        if preferred_author_name:
            ulima_authors_detected.append(preferred_author_name)

        if idx == 0:
            first_author_ulima = True

        block_careers = get_block_engineering_careers(
            block_value=block,
            publication_level_careers=publication_level_careers,
        )

        if idx < len(full_authors):
            scopus_author_name = full_authors[idx]
        elif idx < len(short_authors):
            scopus_author_name = short_authors[idx]
        else:
            scopus_author_name = extract_author_name_from_block(block) or ""

        resolved_career_hint = block_careers[0] if len(block_careers) == 1 else None

        match_result = match_scopus_author_to_docente(
            scopus_author_name=scopus_author_name,
            docentes_ref=docentes_ref,
            career_hint=resolved_career_hint,
        )

        if match_result["matched"]:
            matched_any = True
            docente = match_result["docente"]

            if docente.get("carrera") in VALID_ENGINEERING_CAREERS:
                docente_ref_careers.append(docente.get("carrera"))

            if (
                not block_careers
                and block_details.get("has_ulima_engineering_context")
                and docente.get("carrera") in VALID_ENGINEERING_CAREERS
            ):
                block_careers = [docente.get("carrera")]

        publication_careers.extend(filter_valid_engineering_careers(block_careers))

    if not publication_careers and publication_level_careers and publication_details.get("has_ulima_engineering_context"):
        publication_careers.extend(publication_level_careers)
        has_ulima_engineering_affiliation = True

    publication_careers = unique_keep_order([c for c in publication_careers if c])
    docente_ref_careers = unique_keep_order([c for c in docente_ref_careers if c])
    ulima_authors_detected = unique_keep_order([a for a in ulima_authors_detected if a])

    if matched_any and publication_careers:
        metodo = "AFFILIATION+DOCENTES_REF"
    elif publication_careers:
        metodo = "AFFILIATION_ONLY"
    elif matched_any:
        metodo = "DOCENTES_REF_ONLY"
    else:
        metodo = None

    return {
        "ulima_docentes_raw": "; ".join(ulima_authors_detected) if ulima_authors_detected else None,
        "first_author_ulima_raw": "True" if first_author_ulima else "False",
        "carrera_raw": "; ".join(publication_careers) if publication_careers else None,
        "docentes_ref_careers": docente_ref_careers,
        "metodo_cruce_scopus_raw": metodo,
        "es_ulima_raw_detected": affiliation_ulima_detected,
        "has_ulima_engineering_affiliation_raw": has_ulima_engineering_affiliation,
    }


# =========================
# CSV MAPPING
# =========================
COLUMN_ALIASES = {
    "eid": ["EID", "eid"],
    "publication_year_raw": ["Year", "Año"],
    "scopus_year_raw": ["Año (Scopus)", "Scopus Year", "Year"],
    "publication_type_raw": ["Tipo de publicación", "Document Type", "Source & document type", "Document Type (Scopus)"],
    "publication_title_raw": ["Título de la publicación", "Title", "Publication Title", "Document title"],
    "authors_raw": ["Authors", "Autor(es)", "Author(s)"],
    "conference_journal_raw": ["Conferencia/Journal", "Conference/Journal", "Conference Journal"],
    "indexation_raw": ["Indexación", "Indexation"],
    "editorial_publication_raw": ["Editorial de publicación", "Publisher"],
    "doi_link_raw": ["DOI /link", "DOI"],
    "publication_date_raw": ["Fecha de publicacion", "Publication date", "Date"],
    "publication_date_scopus_raw": ["Fecha de publicacion (Scopus)"],
    "revista_raw": ["Revista"],
    "conference_raw": ["Conferencia", "Conference name"],
    "publication_status_raw": ["Estado publicación", "Publication Stage", "Publication Status"],
    "issn_raw": ["ISSN", "Serial identifiers (e.g. ISSN)"],
    "isbn_raw": ["ISBN"],
    "scopus_link_raw": ["Link SCOPUS", "Link"],
    "alternative_link_raw": ["LINK REVISTA (ALTERNATIVO)", "Alternative Link"],
    "category_tematica_raw": ["Categoria temática", "Categoría temática"],
    "area_idic_raw": ["Area IDIC"],
    "linea_idic_raw": ["Linea IDIC"],
    "area_carrera_raw": ["Area de la carrera"],
    "linea_carrera_raw": ["Linea de la carrera"],
    "carrera_raw": ["Carrera"],
    "es_scopus_raw": ["es_scopus"],
    "retractado_raw": ["Retractado"],
    "descontinuado_scopus_raw": ["Descontinuado de SCOPUS", "Descontinuado de Scopus"],
    "metodo_cruce_scopus_raw": ["Método de cruce Scopus", "Metodo de cruce Scopus"],
    "abstract_scopus_raw": ["Abstract Scopus", "Abstract"],
    "author_keywords_raw": ["Author keywords", "Author Keywords"],
    "index_keywords_raw": ["Index keywords", "Index Keywords"],
    "source_title_raw": ["Source title", "Source Title"],
    "document_type_scopus_raw": ["Document Type (Scopus)", "Document Type"],
    "affiliation_raw": ["Affiliations", "Afiliaciones"],
}


def map_row_to_staging(
    row: dict,
    docentes_ref: list[dict],
    use_llm: bool | None = None,
    use_career_ai: bool = False,
    use_thematic_review_ai: bool = False,
    use_ai_taxonomy: bool = False,
) -> dict:
    mapped = {}

    for target_column, aliases in COLUMN_ALIASES.items():
        mapped[target_column] = safe_get(row, aliases)

    authors_value = safe_get(row, ["Authors", "Author(s)", "Autor(es)"])
    author_full_names_value = safe_get(row, ["Author full names"])
    authors_with_affiliations_value = safe_get(row, ["Authors with affiliations"])
    affiliations_value = safe_get(row, ["Affiliations", "Afiliaciones"])
    source_title_value = safe_get(row, ["Source title", "Source Title"])
    document_type_value = safe_get(row, ["Document Type", "Document Type (Scopus)", "Source & document type"])
    publication_stage_value = safe_get(row, ["Publication Stage", "Publication Status"])
    publisher_value = safe_get(row, ["Publisher"])
    doi_value = safe_get(row, ["DOI /link", "DOI"])

    enrichment = enrich_ulima_fields_from_ref(
        authors_value=authors_value,
        author_full_names_value=author_full_names_value,
        authors_with_affiliations_value=authors_with_affiliations_value,
        affiliations_value=affiliations_value,
        docentes_ref=docentes_ref,
    )

    authors_display = build_authors_display(author_full_names_value, authors_value)
    conference_journal_value = derive_conference_journal_value(document_type_value)
    revista_value, conference_value = derive_revista_and_conference(source_title_value, document_type_value)
    alternative_link_value = build_doi_url(doi_value)

    mapped["authors_raw"] = authors_display or mapped.get("authors_raw")
    mapped["affiliation_raw"] = authors_with_affiliations_value or affiliations_value or mapped.get("affiliation_raw")
    mapped["ulima_docentes_raw"] = enrichment["ulima_docentes_raw"]
    mapped["first_author_ulima_raw"] = enrichment["first_author_ulima_raw"]

    if enrichment["carrera_raw"]:
        mapped["carrera_raw"] = enrichment["carrera_raw"]

    if not mapped.get("conference_journal_raw"):
        mapped["conference_journal_raw"] = conference_journal_value

    mapped["revista_raw"] = revista_value
    mapped["conference_raw"] = conference_value

    if not mapped.get("editorial_publication_raw"):
        mapped["editorial_publication_raw"] = publisher_value

    if not mapped.get("publication_status_raw"):
        mapped["publication_status_raw"] = publication_stage_value

    if not mapped.get("source_title_raw"):
        mapped["source_title_raw"] = source_title_value

    if not mapped.get("document_type_scopus_raw"):
        mapped["document_type_scopus_raw"] = document_type_value

    if not mapped.get("doi_link_raw"):
        mapped["doi_link_raw"] = normalize_url_or_doi(doi_value)

    if not mapped.get("alternative_link_raw"):
        mapped["alternative_link_raw"] = alternative_link_value

    if not mapped.get("indexation_raw"):
        mapped["indexation_raw"] = "Scopus"

    if not mapped.get("es_scopus_raw"):
        mapped["es_scopus_raw"] = "True"

    if not mapped.get("retractado_raw"):
        mapped["retractado_raw"] = "No"

    if not mapped.get("descontinuado_scopus_raw"):
        mapped["descontinuado_scopus_raw"] = "No"

    if enrichment["metodo_cruce_scopus_raw"]:
        mapped["metodo_cruce_scopus_raw"] = enrichment["metodo_cruce_scopus_raw"]
    elif mapped.get("eid") and not mapped.get("metodo_cruce_scopus_raw"):
        mapped["metodo_cruce_scopus_raw"] = "EID"

    eligibility = determine_row_engineering_eligibility(
        document_type_value=document_type_value,
        authors_with_affiliations_value=authors_with_affiliations_value,
        affiliations_value=affiliations_value,
        enrichment=enrichment,
        title_value=mapped.get("publication_title_raw"),
        abstract_value=mapped.get("abstract_scopus_raw"),
        author_keywords_value=mapped.get("author_keywords_raw"),
        index_keywords_value=mapped.get("index_keywords_raw"),
        source_title_value=mapped.get("source_title_raw"),
        authors_value=mapped.get("authors_raw"),
        use_career_ai=use_career_ai,
    )

    if eligibility.get("carrera_raw"):
        mapped["carrera_raw"] = eligibility.get("carrera_raw")

    if eligibility.get("metodo_cruce_scopus_raw"):
        mapped["metodo_cruce_scopus_raw"] = eligibility.get("metodo_cruce_scopus_raw")

    if not eligibility.get("eligible"):
        mapped["area_carrera_raw"] = None
        mapped["linea_carrera_raw"] = None
        mapped["category_tematica_raw"] = None
        mapped["area_idic_raw"] = None
        mapped["linea_idic_raw"] = None
        mapped = sanitize_identifier_fields(mapped)
        mapped["record_hash"] = compute_record_hash(
            mapped.get("eid"),
            mapped.get("doi_link_raw"),
            mapped.get("publication_title_raw"),
        )
        mapped["is_valid_for_curated"] = 0
        mapped["rejection_reason"] = eligibility.get("reason")
        mapped["__skip_insert__"] = True
        return mapped

    # -------------------------
    # Thematic classification
    # -------------------------
    text_fields_for_primary_career = build_thematic_text_fields(
        mapped.get("publication_title_raw"),
        mapped.get("abstract_scopus_raw"),
        mapped.get("author_keywords_raw"),
        mapped.get("index_keywords_raw"),
        mapped.get("source_title_raw"),
    )
    carrera_for_classification = choose_primary_career_for_classification(
        mapped.get("carrera_raw"),
        text_fields_for_primary_career,
    )

    if carrera_for_classification:
        thematic = classify_thematic_fields(
            carrera=carrera_for_classification,
            title_value=mapped.get("publication_title_raw"),
            abstract_value=mapped.get("abstract_scopus_raw"),
            author_keywords_value=mapped.get("author_keywords_raw"),
            index_keywords_value=mapped.get("index_keywords_raw"),
            source_title_value=mapped.get("source_title_raw"),
            use_llm=use_llm,
            use_thematic_review_ai=use_thematic_review_ai,
            use_ai_taxonomy=use_ai_taxonomy,
        )

        if not mapped.get("area_carrera_raw") and thematic.get("area_carrera_raw"):
            mapped["area_carrera_raw"] = thematic["area_carrera_raw"]

        if not mapped.get("linea_carrera_raw") and thematic.get("linea_carrera_raw"):
            mapped["linea_carrera_raw"] = thematic["linea_carrera_raw"]

        if not mapped.get("category_tematica_raw") and thematic.get("category_tematica_raw"):
            mapped["category_tematica_raw"] = thematic["category_tematica_raw"]

        if not mapped.get("area_idic_raw") and thematic.get("area_idic_raw"):
            mapped["area_idic_raw"] = thematic["area_idic_raw"]

        # RUN84 FIX: copiar SIEMPRE la línea IDIC desde la clasificación temática.
        # En run83 faltaba este bloque y, por eso, los registros válidos quedaban con
        # linea_idic_raw = NULL; luego la validación del triplete anulaba también
        # category_tematica_raw y area_idic_raw.
        if not mapped.get("linea_idic_raw") and thematic.get("linea_idic_raw"):
            mapped["linea_idic_raw"] = thematic["linea_idic_raw"]

        if thematic.get("thematic_review_rejected"):
            mapped["area_carrera_raw"] = thematic.get("area_carrera_raw")
            mapped["linea_carrera_raw"] = thematic.get("linea_carrera_raw")
            mapped["category_tematica_raw"] = thematic.get("category_tematica_raw")
            mapped["area_idic_raw"] = thematic.get("area_idic_raw")
            mapped["linea_idic_raw"] = thematic.get("linea_idic_raw")
            mapped = sanitize_identifier_fields(mapped)
            mapped["record_hash"] = compute_record_hash(
                mapped.get("eid"),
                mapped.get("doi_link_raw"),
                mapped.get("publication_title_raw"),
            )
            mapped["is_valid_for_curated"] = 0
            mapped["rejection_reason"] = thematic.get("thematic_review_rejection_reason")
            mapped["__skip_insert__"] = True
            return mapped

    if not is_valid_career_area_line(
        carrera_for_classification,
        mapped.get("area_carrera_raw"),
        mapped.get("linea_carrera_raw"),
    ):
        mapped["area_carrera_raw"] = None
        mapped["linea_carrera_raw"] = None

    if not is_valid_idic_triplet(
        mapped.get("category_tematica_raw"),
        mapped.get("area_idic_raw"),
        mapped.get("linea_idic_raw"),
    ):
        mapped["category_tematica_raw"] = None
        mapped["area_idic_raw"] = None
        mapped["linea_idic_raw"] = None

    mapped = sanitize_identifier_fields(mapped)

    record_hash = compute_record_hash(
        mapped.get("eid"),
        mapped.get("doi_link_raw"),
        mapped.get("publication_title_raw"),
    )

    has_identity = any([
        mapped.get("eid"),
        mapped.get("doi_link_raw"),
        mapped.get("publication_title_raw"),
    ])

    mapped["record_hash"] = record_hash
    mapped["is_valid_for_curated"] = 1 if has_identity else 0
    mapped["rejection_reason"] = None if has_identity else "Missing EID/DOI/title"
    mapped["__skip_insert__"] = False

    return mapped


def parse_csv_text(csv_text: str) -> list[dict]:
    normalized_text = (csv_text or "").replace("\r\n", "\n").replace("\r", "\n")
    normalized_text, hinted_delimiter = strip_excel_separator_hint(normalized_text)

    candidate_delimiters = [hinted_delimiter] if hinted_delimiter else []
    for delimiter in CSV_DELIMITER_CANDIDATES:
        if delimiter not in candidate_delimiters:
            candidate_delimiters.append(delimiter)

    best_rows: list[list[str]] | None = None
    best_score: int | None = None
    best_delimiter: str | None = None

    for delimiter in candidate_delimiters:
        try:
            parsed_rows = parse_csv_rows_with_delimiter(normalized_text, delimiter)
        except Exception:
            continue

        candidate_score = score_csv_candidate(parsed_rows)
        if best_score is None or candidate_score > best_score:
            best_rows = parsed_rows
            best_score = candidate_score
            best_delimiter = delimiter

    if not best_rows:
        raise RuntimeError("Unable to parse CSV with supported delimiters.")

    logging.info(
        "CSV parsed using delimiter %r with score %s and %s data rows.",
        best_delimiter,
        best_score,
        max(len(best_rows) - 1, 0),
    )

    return build_dict_rows_from_csv_rows(best_rows)


def insert_scopus_raw_load_row(cursor, run_id: int, source_file_name: str, source_row_number: int, mapped: dict) -> None:
    cursor.execute(
        """
        INSERT INTO stg.scopus_raw_load (
            run_id,
            source_row_number,
            source_file_name,
            eid,
            publication_year_raw,
            scopus_year_raw,
            publication_type_raw,
            publication_title_raw,
            authors_raw,
            ulima_docentes_raw,
            ulima_estudiantes_raw,
            first_author_ulima_raw,
            conference_journal_raw,
            indexation_raw,
            editorial_publication_raw,
            doi_link_raw,
            publication_date_raw,
            publication_date_scopus_raw,
            revista_raw,
            conference_raw,
            publication_status_raw,
            issn_raw,
            isbn_raw,
            scopus_link_raw,
            alternative_link_raw,
            category_tematica_raw,
            area_idic_raw,
            linea_idic_raw,
            area_carrera_raw,
            linea_carrera_raw,
            carrera_raw,
            es_scopus_raw,
            retractado_raw,
            descontinuado_scopus_raw,
            metodo_cruce_scopus_raw,
            abstract_scopus_raw,
            author_keywords_raw,
            index_keywords_raw,
            source_title_raw,
            document_type_scopus_raw,
            affiliation_raw,
            record_hash,
            is_valid_for_curated,
            rejection_reason
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            run_id,
            source_row_number,
            source_file_name,
            mapped.get("eid"),
            mapped.get("publication_year_raw"),
            mapped.get("scopus_year_raw"),
            mapped.get("publication_type_raw"),
            mapped.get("publication_title_raw"),
            mapped.get("authors_raw"),
            mapped.get("ulima_docentes_raw"),
            mapped.get("ulima_estudiantes_raw"),
            mapped.get("first_author_ulima_raw"),
            mapped.get("conference_journal_raw"),
            mapped.get("indexation_raw"),
            mapped.get("editorial_publication_raw"),
            mapped.get("doi_link_raw"),
            mapped.get("publication_date_raw"),
            mapped.get("publication_date_scopus_raw"),
            mapped.get("revista_raw"),
            mapped.get("conference_raw"),
            mapped.get("publication_status_raw"),
            mapped.get("issn_raw"),
            mapped.get("isbn_raw"),
            mapped.get("scopus_link_raw"),
            mapped.get("alternative_link_raw"),
            mapped.get("category_tematica_raw"),
            mapped.get("area_idic_raw"),
            mapped.get("linea_idic_raw"),
            mapped.get("area_carrera_raw"),
            mapped.get("linea_carrera_raw"),
            mapped.get("carrera_raw"),
            mapped.get("es_scopus_raw"),
            mapped.get("retractado_raw"),
            mapped.get("descontinuado_scopus_raw"),
            mapped.get("metodo_cruce_scopus_raw"),
            mapped.get("abstract_scopus_raw"),
            mapped.get("author_keywords_raw"),
            mapped.get("index_keywords_raw"),
            mapped.get("source_title_raw"),
            mapped.get("document_type_scopus_raw"),
            mapped.get("affiliation_raw"),
            mapped.get("record_hash"),
            mapped.get("is_valid_for_curated"),
            mapped.get("rejection_reason"),
        ),
    )


def insert_rows_to_staging(
    run_id: int,
    source_file_name: str,
    rows: list[dict],
    docentes_ref: list[dict],
    source_row_start: int = 1,
    use_llm: bool | None = None,
    use_career_ai: bool = False,
    use_thematic_review_ai: bool = False,
    use_ai_taxonomy: bool = False,
) -> tuple[int, int, list[tuple[int, dict]]]:
    """
    Inserta SOLO filas válidas en staging antes del upsert y devuelve rechazados en memoria.

    Motivo: algunos procedimientos de upsert antiguos podrían no filtrar is_valid_for_curated=1.
    Por seguridad, los rechazados se guardan después del upsert, como auditoría, para no
    contaminar curated.publications.
    """
    from mssql_python import connect

    sql_conn_str = get_sql_connection_string()
    inserted_valid = 0
    rejected = 0
    rejected_audit_rows: list[tuple[int, dict]] = []

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()

        for idx, row in enumerate(rows, start=source_row_start):
            mapped = map_row_to_staging(
                row,
                docentes_ref=docentes_ref,
                use_llm=use_llm,
                use_career_ai=use_career_ai,
                use_thematic_review_ai=use_thematic_review_ai,
                use_ai_taxonomy=use_ai_taxonomy,
            )

            if mapped.get("is_valid_for_curated") == 1 and not mapped.get("__skip_insert__"):
                insert_scopus_raw_load_row(cursor, run_id, source_file_name, idx, mapped)
                inserted_valid += 1
            else:
                rejected += 1
                rejected_audit_rows.append((idx, mapped))

        conn.commit()
        cursor.close()

    return inserted_valid, rejected, rejected_audit_rows


def insert_rejected_rows_to_staging(
    run_id: int,
    source_file_name: str,
    rejected_audit_rows: list[tuple[int, dict]],
) -> int:
    """
    Inserta rechazados en stg.scopus_raw_load después del upsert.

    Esto permite auditar por SQL los falsos negativos de afiliación sin arriesgar que
    el procedimiento de upsert los inserte accidentalmente en curated.publications.
    """
    if not rejected_audit_rows:
        return 0

    from mssql_python import connect

    sql_conn_str = get_sql_connection_string()
    inserted_rejected = 0

    with connect(sql_conn_str) as conn:
        cursor = conn.cursor()

        for source_row_number, mapped in rejected_audit_rows:
            mapped["is_valid_for_curated"] = 0
            if not mapped.get("rejection_reason"):
                mapped["rejection_reason"] = "REJECT_UNSPECIFIED"
            insert_scopus_raw_load_row(cursor, run_id, source_file_name, source_row_number, mapped)
            inserted_rejected += 1

        conn.commit()
        cursor.close()

    return inserted_rejected


# =========================
# HTTP FUNCTIONS
# =========================
@app.route(route="health", methods=["GET"])
def health(req: func.HttpRequest) -> func.HttpResponse:
    try:
        container_status = validate_containers()

        payload = {
            "status": "ok",
            "utc_now": utc_now_iso(),
            "function_app": "funcprodcientificadev01",
            "sql_server": get_env("SQL_SERVER"),
            "sql_database": get_env("SQL_DATABASE"),
            "storage_account": get_env("DATA_STORAGE_ACCOUNT"),
            "containers": container_status,
            "thematic_llm_configured": is_thematic_llm_configured(),
            "thematic_llm_enabled": resolve_llm_enabled(),
            "thematic_review_ai_configured": is_thematic_llm_configured(),
            "thematic_review_ai_enabled": resolve_thematic_review_ai_enabled(),
            "thematic_review_ai_min_confidence": get_thematic_review_ai_min_confidence(),
            "thematic_review_ai_call_delay_seconds": get_thematic_review_ai_call_delay_seconds(),
            "thematic_review_ai_max_calls_per_run": get_thematic_review_ai_max_calls_per_run(),
            "thematic_review_send_unsafe_to_review": should_send_thematic_review_to_rejected(),
            "ai_taxonomy_classifier_configured": is_thematic_llm_configured(),
            "ai_taxonomy_classifier_enabled": resolve_ai_taxonomy_classifier_enabled(),
            "ai_taxonomy_min_confidence": get_ai_taxonomy_min_confidence(),
            "ai_taxonomy_call_delay_seconds": get_ai_taxonomy_call_delay_seconds(),
            "ai_taxonomy_max_calls_per_run": get_ai_taxonomy_max_calls_per_run(),
            "ai_taxonomy_review_to_rejected": should_ai_taxonomy_review_reject(),
            "career_ambiguity_ai_configured": is_thematic_llm_configured(),
            "career_ambiguity_ai_enabled": resolve_career_ambiguity_ai_enabled(),
            "career_ambiguity_ai_min_confidence": get_career_ambiguity_llm_min_confidence(),
            "career_ai_call_delay_seconds": get_career_ai_call_delay_seconds(),
            "career_ai_max_calls_per_run": get_career_ai_max_calls_per_run(),
            "azure_openai_max_retries": get_openai_max_retries(),
            "azure_openai_retry_base_seconds": get_openai_retry_base_seconds(),
            "career_ai_extra_sleep_after_429_seconds": get_career_ai_throttle_after_429_seconds(),
            "scopus_ingest_singleton_lock_enabled": get_ingest_singleton_lock_enabled(),
            "scopus_ingest_lock_timeout_ms": get_ingest_singleton_lock_timeout_ms(),
            "scopus_max_rows_per_run": get_default_max_rows_per_run(),
            "scopus_hard_max_rows_per_run": get_hard_max_rows_per_run(),
            "save_rejected_to_staging_default": resolve_save_rejected_enabled(),
            "engineering_generic_inference_enabled": True,
            "engineering_generic_inference_min_score": CAREER_INFERENCE_MIN_SCORE,
            "engineering_generic_inference_min_margin": CAREER_INFERENCE_MIN_MARGIN,
            "classification_guardrails_version": "v10_run88_ai_taxonomy_classifier_closed_catalog",
            "post_upsert_reactivation_enabled": True,
            "post_upsert_deactivation_from_rejected_enabled": True,
        }

        return func.HttpResponse(
            json.dumps(payload, ensure_ascii=False, indent=2),
            status_code=200,
            mimetype="application/json",
        )

    except Exception as exc:
        logging.exception("Health check failed")
        return func.HttpResponse(
            json.dumps({"status": "error", "message": str(exc)}, ensure_ascii=False, indent=2),
            status_code=500,
            mimetype="application/json",
        )


@app.route(route="run-smoke", methods=["GET", "POST"])
def run_smoke(req: func.HttpRequest) -> func.HttpResponse:
    run_id = None
    try:
        container_status = validate_containers()
        run_id = insert_pipeline_run_smoke(trigger_type="MANUAL", status="STARTED")
        update_pipeline_run(run_id, status="SUCCESS")

        payload = {
            "status": "ok",
            "message": "Smoke test executed successfully.",
            "run_id": run_id,
            "containers": container_status,
            "utc_now": utc_now_iso(),
        }

        return func.HttpResponse(
            json.dumps(payload, ensure_ascii=False, indent=2),
            status_code=200,
            mimetype="application/json",
        )

    except Exception as exc:
        logging.exception("Smoke run failed")
        if run_id is not None:
            update_pipeline_run(run_id, status="FAILED", error_message=str(exc))
        return func.HttpResponse(
            json.dumps({"status": "error", "message": str(exc)}, ensure_ascii=False, indent=2),
            status_code=500,
            mimetype="application/json",
        )


@app.route(route="run-ingest-docentes", methods=["GET", "POST"])
def run_ingest_docentes(req: func.HttpRequest) -> func.HttpResponse:
    run_id = None

    try:
        docentes_container = get_env("DOCENTES_CONTAINER")
        periodo_academico = get_env("DOCENTES_ACTIVE_PERIOD")
        processed_container = get_env("PROCESSED_CONTAINER")
        logs_container = get_env("LOGS_CONTAINER")

        requested_blob = req.params.get("blob_name")
        blob_name = requested_blob or get_latest_excel_blob_name(docentes_container)

        run_id = create_pipeline_run(
            trigger_type="MANUAL",
            source_name="DOCENTES_ULIMA",
            source_file_name=blob_name,
            source_file_path=f"{docentes_container}/{blob_name}",
        )

        workbook_bytes = download_blob_bytes(docentes_container, blob_name)

        extracted_rows = extract_docentes_from_workbook_bytes(
            workbook_bytes=workbook_bytes,
            source_file_name=blob_name,
            periodo_academico=periodo_academico,
        )

        raw_rows_loaded = insert_docentes_raw_rows(run_id, extracted_rows)
        docentes_unique_loaded = rebuild_docentes_reference(
            run_id=run_id,
            rows=extracted_rows,
            periodo_academico=periodo_academico,
        )

        processed_name = f"docentes/{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{os.path.basename(blob_name)}"
        upload_bytes_blob(processed_container, processed_name, workbook_bytes)

        log_payload = {
            "run_id": run_id,
            "source_blob": blob_name,
            "processed_blob": processed_name,
            "periodo_academico": periodo_academico,
            "raw_rows_loaded": raw_rows_loaded,
            "docentes_unique_loaded": docentes_unique_loaded,
            "utc_now": utc_now_iso(),
            "status": "SUCCESS",
        }

        log_name = f"docentes_run_{run_id}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}.json"
        upload_text_blob(logs_container, log_name, json.dumps(log_payload, ensure_ascii=False, indent=2))

        update_pipeline_run(
            run_id=run_id,
            status="SUCCESS",
            records_read=raw_rows_loaded,
            records_inserted=docentes_unique_loaded,
            records_updated=0,
            records_rejected=0,
            error_message=None,
        )

        return func.HttpResponse(
            json.dumps(
                {
                    "status": "ok",
                    "message": "Docentes ingestion completed successfully.",
                    "run_id": run_id,
                    "source_blob": blob_name,
                    "processed_blob": processed_name,
                    "periodo_academico": periodo_academico,
                    "raw_rows_loaded": raw_rows_loaded,
                    "docentes_unique_loaded": docentes_unique_loaded,
                    "utc_now": utc_now_iso(),
                },
                ensure_ascii=False,
                indent=2,
            ),
            status_code=200,
            mimetype="application/json",
        )

    except Exception as exc:
        logging.exception("Docentes ingestion failed")

        if run_id is not None:
            update_pipeline_run(
                run_id=run_id,
                status="FAILED",
                error_message=str(exc),
            )

        return func.HttpResponse(
            json.dumps(
                {
                    "status": "error",
                    "message": str(exc),
                    "run_id": run_id,
                    "utc_now": utc_now_iso(),
                },
                ensure_ascii=False,
                indent=2,
            ),
            status_code=500,
            mimetype="application/json",
        )


@app.route(route="run-ingest-scopus", methods=["GET", "POST"])
def run_ingest_scopus(req: func.HttpRequest) -> func.HttpResponse:
    run_id = None
    ingest_lock_conn = None
    try:
        raw_container = get_env("RAW_CONTAINER")
        processed_container = get_env("PROCESSED_CONTAINER")
        logs_container = get_env("LOGS_CONTAINER")
        periodo_academico = get_env("DOCENTES_ACTIVE_PERIOD")
        docentes_ref = get_docentes_reference(periodo_academico)

        request_payload = get_request_json_payload(req)

        requested_blob = get_request_value(req, request_payload, ["blob_name", "blobName", "source_blob"])
        blob_name = requested_blob or get_latest_csv_blob_name(raw_container)

        # Parámetros para carga masiva segura.
        # start_row/end_row/max_rows se refieren a filas de datos del CSV, no a la cabecera.
        start_row = parse_positive_int(get_request_value(req, request_payload, ["start_row", "startRow"]), "start_row")
        end_row = parse_positive_int(get_request_value(req, request_payload, ["end_row", "endRow"]), "end_row")
        max_rows = parse_positive_int(get_request_value(req, request_payload, ["max_rows", "maxRows"]), "max_rows")

        # RUN84: procesamiento controlado por lotes. Si no se define rango,
        # se procesa automáticamente un máximo de 40 filas. Si se pide más,
        # se detiene con error explícito.
        hard_max_rows = get_hard_max_rows_per_run()
        if max_rows is None and end_row is None:
            max_rows = get_default_max_rows_per_run() or hard_max_rows
        if max_rows is not None and max_rows > hard_max_rows:
            raise ValueError(
                f"max_rows={max_rows} exceeds SCOPUS_HARD_MAX_ROWS_PER_RUN={hard_max_rows}. "
                "Process Scopus in controlled batches."
            )
        if start_row is not None and end_row is not None and (end_row - start_row + 1) > hard_max_rows:
            raise ValueError(
                f"Requested row range has {end_row - start_row + 1} rows, exceeding "
                f"SCOPUS_HARD_MAX_ROWS_PER_RUN={hard_max_rows}."
            )

        llm_value = get_request_value(req, request_payload, ["llm", "thematic_llm", "use_llm", "useLlm"])
        career_ai_value = get_request_value(
            req,
            request_payload,
            ["career_ai", "careerAi", "career_ambiguity_ai", "career_llm", "ai_career", "use_career_ai", "useCareerAi"],
        )
        thematic_review_ai_value = get_request_value(
            req,
            request_payload,
            ["theme_review_ai", "thematic_review_ai", "thematicReviewAi", "review_ai", "use_thematic_review_ai"],
        )
        ai_taxonomy_value = get_request_value(
            req,
            request_payload,
            ["ai_taxonomy", "taxonomy_ai", "aiTaxonomy", "ai_taxonomy_classifier", "use_ai_taxonomy"],
        )
        save_rejected_value = get_request_value(
            req,
            request_payload,
            ["save_rejected", "saveRejected", "save_rejected_to_staging", "saveRejectedToStaging"],
        )

        use_llm = resolve_llm_enabled(llm_value)
        use_career_ai = resolve_career_ambiguity_ai_enabled(career_ai_value)
        use_thematic_review_ai = resolve_thematic_review_ai_enabled(thematic_review_ai_value)
        use_ai_taxonomy = resolve_ai_taxonomy_classifier_enabled(ai_taxonomy_value)
        save_rejected_to_staging = resolve_save_rejected_enabled(save_rejected_value)

        reset_ai_runtime_counters()
        validate_career_ai_runtime_configuration(use_career_ai)
        validate_ai_taxonomy_runtime_configuration(use_ai_taxonomy)

        # Candado distribuido: impide corridas simultáneas que saturan Azure OpenAI.
        ingest_lock_conn = acquire_scopus_ingest_singleton_lock()

        run_id = create_pipeline_run(
            trigger_type="MANUAL",
            source_name="SCOPUS",
            source_file_name=blob_name,
            source_file_path=f"{raw_container}/{blob_name}",
        )

        csv_text = download_blob_text(raw_container, blob_name)
        all_rows = parse_csv_text(csv_text)
        total_rows_in_file = len(all_rows)

        rows, effective_start_row, effective_end_row = slice_rows_by_1_based_range(
            all_rows,
            start_row=start_row,
            end_row=end_row,
            max_rows=max_rows,
        )

        records_read = len(rows)

        # RUN84: control RAW antes de clasificar. Si el CSV no trae afiliación o
        # abstract, no se debe marcar todo como REJECT_NO_ULIMA_AFFILIATION; el
        # archivo está incompleto para este pipeline.
        if records_read > 0:
            affiliation_missing = sum(
                1
                for r in rows
                if not (safe_get(r, ["Authors with affiliations"]) or safe_get(r, ["Affiliations", "Afiliaciones"]))
            )
            abstract_missing = sum(
                1
                for r in rows
                if not safe_get(r, ["Abstract Scopus", "Abstract", "Resumen"])
            )
            if affiliation_missing / records_read >= 0.80:
                raise RuntimeError(
                    "RAW_INCOMPLETE_HEADERS: more than 80% of selected rows have no "
                    "Affiliations/Authors with affiliations. Re-export Scopus with full metadata."
                )
            if abstract_missing / records_read >= 0.80:
                raise RuntimeError(
                    "RAW_INCOMPLETE_HEADERS: more than 80% of selected rows have no Abstract. "
                    "Re-export Scopus with full metadata."
                )

        records_inserted, records_rejected, rejected_audit_rows = insert_rows_to_staging(
            run_id=run_id,
            source_file_name=blob_name,
            rows=rows,
            docentes_ref=docentes_ref,
            source_row_start=effective_start_row,
            use_llm=use_llm,
            use_career_ai=use_career_ai,
            use_thematic_review_ai=use_thematic_review_ai,
            use_ai_taxonomy=use_ai_taxonomy,
        )

        execute_upsert_from_staging(run_id)
        records_reactivated_in_curated = reactivate_curated_publications_from_valid_stg(run_id)

        records_rejected_saved_to_staging = 0
        records_deactivated_from_rejected = 0
        if save_rejected_to_staging:
            records_rejected_saved_to_staging = insert_rejected_rows_to_staging(
                run_id=run_id,
                source_file_name=blob_name,
                rejected_audit_rows=rejected_audit_rows,
            )
            records_deactivated_from_rejected = deactivate_curated_publications_from_rejected_stg(run_id)

        total_rows_inserted_to_staging = records_inserted + records_rejected_saved_to_staging

        processed_name = f"processed/{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{os.path.basename(blob_name)}"
        upload_text_blob(processed_container, processed_name, csv_text)

        log_payload = {
            "run_id": run_id,
            "source_blob": blob_name,
            "processed_blob": processed_name,
            "periodo_academico_docentes": periodo_academico,
            "total_rows_in_file": total_rows_in_file,
            "start_row": effective_start_row,
            "end_row": effective_end_row,
            "records_read": records_read,
            "records_inserted_to_staging": total_rows_inserted_to_staging,
            "records_valid_for_curated": records_inserted,
            "records_rejected": records_rejected,
            "records_rejected_saved_to_staging": records_rejected_saved_to_staging,
            "records_reactivated_in_curated": records_reactivated_in_curated,
            "records_deactivated_from_rejected": records_deactivated_from_rejected,
            "save_rejected_to_staging": save_rejected_to_staging,
            "thematic_llm_enabled": use_llm,
            "thematic_review_ai_enabled": use_thematic_review_ai,
            "thematic_review_ai_requested_value": thematic_review_ai_value,
            "thematic_review_ai_min_confidence": get_thematic_review_ai_min_confidence(),
            "thematic_review_ai_call_delay_seconds": get_thematic_review_ai_call_delay_seconds(),
            "thematic_review_ai_max_calls_per_run": get_thematic_review_ai_max_calls_per_run(),
            "thematic_review_ai_calls_this_run": get_thematic_review_ai_calls_this_run(),
            "ai_taxonomy_classifier_enabled": use_ai_taxonomy,
            "ai_taxonomy_requested_value": ai_taxonomy_value,
            "ai_taxonomy_min_confidence": get_ai_taxonomy_min_confidence(),
            "ai_taxonomy_call_delay_seconds": get_ai_taxonomy_call_delay_seconds(),
            "ai_taxonomy_max_calls_per_run": get_ai_taxonomy_max_calls_per_run(),
            "ai_taxonomy_calls_this_run": get_ai_taxonomy_calls_this_run(),
            "career_ambiguity_ai_enabled": use_career_ai,
            "career_ambiguity_ai_requested_value": career_ai_value,
            "career_ambiguity_ai_configured": is_thematic_llm_configured(),
            "career_ambiguity_ai_min_confidence": get_career_ambiguity_llm_min_confidence(),
            "career_ai_call_delay_seconds": get_career_ai_call_delay_seconds(),
            "career_ai_max_calls_per_run": get_career_ai_max_calls_per_run(),
            "career_ai_calls_this_run": get_career_ai_calls_this_run(),
            "azure_openai_max_retries": get_openai_max_retries(),
            "azure_openai_retry_base_seconds": get_openai_retry_base_seconds(),
            "career_ai_extra_sleep_after_429_seconds": get_career_ai_throttle_after_429_seconds(),
            "scopus_ingest_singleton_lock_enabled": get_ingest_singleton_lock_enabled(),
            "scopus_ingest_lock_timeout_ms": get_ingest_singleton_lock_timeout_ms(),
            "utc_now": utc_now_iso(),
            "status": "SUCCESS",
        }

        log_name = f"run_{run_id}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}.json"
        upload_text_blob(logs_container, log_name, json.dumps(log_payload, ensure_ascii=False, indent=2))

        update_pipeline_run(
            run_id=run_id,
            status="SUCCESS",
            records_read=records_read,
            records_inserted=records_inserted,
            records_updated=0,
            records_rejected=records_rejected,
            error_message=None,
        )

        release_scopus_ingest_singleton_lock(ingest_lock_conn)
        ingest_lock_conn = None

        return func.HttpResponse(
            json.dumps(
                {
                    "status": "ok",
                    "message": "Scopus ingestion completed successfully.",
                    "run_id": run_id,
                    "source_blob": blob_name,
                    "processed_blob": processed_name,
                    "periodo_academico_docentes": periodo_academico,
                    "total_rows_in_file": total_rows_in_file,
                    "start_row": effective_start_row,
                    "end_row": effective_end_row,
                    "records_read": records_read,
                    "records_inserted_to_staging": total_rows_inserted_to_staging,
                    "records_valid_for_curated": records_inserted,
                    "records_rejected": records_rejected,
                    "records_rejected_saved_to_staging": records_rejected_saved_to_staging,
                    "save_rejected_to_staging": save_rejected_to_staging,
                    "thematic_llm_enabled": use_llm,
                    "ai_taxonomy_classifier_enabled": use_ai_taxonomy,
                    "ai_taxonomy_min_confidence": get_ai_taxonomy_min_confidence(),
                    "ai_taxonomy_call_delay_seconds": get_ai_taxonomy_call_delay_seconds(),
                    "ai_taxonomy_max_calls_per_run": get_ai_taxonomy_max_calls_per_run(),
                    "ai_taxonomy_calls_this_run": get_ai_taxonomy_calls_this_run(),
                    "career_ambiguity_ai_enabled": use_career_ai,
                    "career_ambiguity_ai_min_confidence": get_career_ambiguity_llm_min_confidence(),
                    "career_ai_call_delay_seconds": get_career_ai_call_delay_seconds(),
                    "career_ai_max_calls_per_run": get_career_ai_max_calls_per_run(),
                    "career_ai_calls_this_run": get_career_ai_calls_this_run(),
                    "azure_openai_max_retries": get_openai_max_retries(),
                    "azure_openai_retry_base_seconds": get_openai_retry_base_seconds(),
                    "career_ai_extra_sleep_after_429_seconds": get_career_ai_throttle_after_429_seconds(),
                    "utc_now": utc_now_iso(),
                },
                ensure_ascii=False,
                indent=2,
            ),
            status_code=200,
            mimetype="application/json",
        )

    except Exception as exc:
        logging.exception("Scopus ingestion failed")

        if ingest_lock_conn is not None:
            release_scopus_ingest_singleton_lock(ingest_lock_conn)
            ingest_lock_conn = None

        if run_id is not None:
            update_pipeline_run(
                run_id=run_id,
                status="FAILED",
                error_message=str(exc),
            )

        return func.HttpResponse(
            json.dumps(
                {
                    "status": "error",
                    "message": str(exc),
                    "run_id": run_id,
                    "utc_now": utc_now_iso(),
                },
                ensure_ascii=False,
                indent=2,
            ),
            status_code=500,
            mimetype="application/json",
        )
